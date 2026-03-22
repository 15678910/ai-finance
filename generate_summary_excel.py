"""
generate_summary_excel.py
일일 시황 종합 보고서 Excel 파일 생성 스크립트.
daily_sector_analysis.py가 종합요약.txt를 생성한 후 자동 호출되거나
독립 실행 모두 지원.
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다. pip install openpyxl 실행 후 재시도하세요.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 색상 상수
# ---------------------------------------------------------------------------
C_HEADER_DARK  = "0D1B2A"
C_HEADER_BLUE  = "1B4F72"
C_ACCENT       = "2E86C1"
C_LIGHT_BLUE   = "D6EAF8"
C_WHITE        = "FFFFFF"
C_BLACK        = "000000"

C_GREEN        = "EAFAF1"
C_YELLOW       = "FEF9E7"
C_ORANGE       = "FDEBD0"
C_RED          = "FDEDEC"
C_DARK_GREEN   = "1E8449"
C_DARK_RED     = "C0392B"

FONT_NAME = "맑은 고딕"

# ---------------------------------------------------------------------------
# 헬퍼 함수
# ---------------------------------------------------------------------------

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def font(bold=False, size=11, color=C_BLACK, italic=False) -> Font:
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)


def center_align(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left_align(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def set_col_width(ws, col_letter: str, width: float):
    ws.column_dimensions[col_letter].width = width


def set_row_height(ws, row: int, height: float):
    ws.row_dimensions[row].height = height


def write_cell(ws, row, col, value, bold=False, size=11, color=C_BLACK,
               bg=None, align="left", wrap=False, border=False, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(bold=bold, size=size, color=color, italic=italic)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = center_align(wrap) if align == "center" else left_align(wrap)
    if border:
        cell.border = thin_border()
    return cell


def section_header(ws, row, col_start, col_end, title):
    """섹션 제목 행 — 파란 배경 + 흰 글씨."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font = font(bold=True, size=12, color=C_WHITE)
    cell.fill = fill(C_HEADER_BLUE)
    cell.alignment = left_align()
    set_row_height(ws, row, 22)


def table_header_row(ws, row, columns, col_start=2):
    """테이블 헤더 행 — 진한 배경 + 흰 글씨."""
    for i, col_name in enumerate(columns):
        cell = ws.cell(row=row, column=col_start + i, value=col_name)
        cell.font = font(bold=True, size=10, color=C_WHITE)
        cell.fill = fill(C_HEADER_DARK)
        cell.alignment = center_align()
        cell.border = thin_border()
    set_row_height(ws, row, 20)


# ---------------------------------------------------------------------------
# 레짐 → 배경색 매핑
# ---------------------------------------------------------------------------
REGIME_COLOR = {
    "강한상승": C_GREEN,
    "상승":    C_GREEN,
    "약한상승": C_YELLOW,
    "횡보":    C_YELLOW,
    "약한하락": C_ORANGE,
    "하락":    C_RED,
    "강한하락": C_RED,
}


def regime_bg(regime_str: str) -> str:
    for key, color in REGIME_COLOR.items():
        if key in regime_str:
            return color
    return C_WHITE


# ---------------------------------------------------------------------------
# Sheet 1: 대시보드
# ---------------------------------------------------------------------------

def build_dashboard(wb, date_label: str):
    ws = wb.create_sheet("대시보드")
    ws.sheet_view.showGridLines = False

    # 열 너비
    for col, width in [("A", 3), ("B", 22), ("C", 22), ("D", 18),
                        ("E", 18), ("F", 18), ("G", 18), ("H", 5)]:
        set_col_width(ws, col, width)

    # ── 타이틀 ────────────────────────────────────────────────
    ws.merge_cells("B2:H2")
    title_cell = ws.cell(row=2, column=2,
                          value=f"일일 시황 종합 보고서  ({date_label})")
    title_cell.font = Font(name=FONT_NAME, bold=True, size=16, color=C_WHITE)
    title_cell.fill = fill(C_HEADER_DARK)
    title_cell.alignment = center_align()
    set_row_height(ws, 2, 36)

    ws.merge_cells("B3:H3")
    sub_cell = ws.cell(row=3, column=2,
                        value="AI 기반 일일 금융 시황 분석 시스템 | OhMyOpenCode Finance")
    sub_cell.font = font(size=10, color=C_WHITE, italic=True)
    sub_cell.fill = fill(C_ACCENT)
    sub_cell.alignment = center_align()
    set_row_height(ws, 3, 18)

    set_row_height(ws, 4, 10)  # 여백

    # ── 섹션 1: 매크로 환경 ──────────────────────────────────
    section_header(ws, 5, 2, 7, "  매크로 환경")

    macro_headers = ["항목", "현재값"]
    table_header_row(ws, 6, macro_headers, col_start=2)

    macro_data = [
        ("경기 사이클", "확장기"),
        ("금리 사이클", "인하기 (FFR 3.64%)"),
        ("CPI YoY",    "2.43% (적정)"),
        ("VIX",        "26.8 (경계)"),
        ("실질금리",   "1.26%"),
    ]
    for i, (item, val) in enumerate(macro_data):
        r = 7 + i
        bg = C_LIGHT_BLUE if i % 2 == 0 else C_WHITE
        write_cell(ws, r, 2, item, bold=True, bg=bg, border=True)
        write_cell(ws, r, 3, val, bg=bg, border=True)
        # 나머지 열 채우기 (테두리만)
        for c in range(4, 8):
            ws.cell(row=r, column=c).fill = fill(bg)
        set_row_height(ws, r, 18)

    set_row_height(ws, 12, 10)  # 여백

    # ── 섹션 2: 지정학 리스크 ────────────────────────────────
    section_header(ws, 13, 2, 7, "  지정학 리스크")

    geo_headers = ["항목", "현재값"]
    table_header_row(ws, 14, geo_headers, col_start=2)

    geo_data = [
        ("리스크 온도계",  "주의 (58.9/100)"),
        ("전쟁/군사충돌", "높음"),
        ("관세/무역분쟁", "안정"),
        ("지역갈등",      "높음"),
        ("금(Gold)",      "$4,574.9  (-8.4% 1주)"),
        ("유가(WTI)",     "$98.2  (+5.1% 1주)"),
        ("USD/KRW",       "1,505"),
    ]
    for i, (item, val) in enumerate(geo_data):
        r = 15 + i
        bg = C_LIGHT_BLUE if i % 2 == 0 else C_WHITE
        write_cell(ws, r, 2, item, bold=True, bg=bg, border=True)
        write_cell(ws, r, 3, val, bg=bg, border=True)
        for c in range(4, 8):
            ws.cell(row=r, column=c).fill = fill(bg)
        set_row_height(ws, r, 18)

    set_row_height(ws, 22, 10)  # 여백

    # ── 섹션 3: 자산별 전망 ──────────────────────────────────
    section_header(ws, 23, 2, 7, "  자산별 전망")

    asset_headers = ["자산", "영향", "전략"]
    table_header_row(ws, 24, asset_headers, col_start=2)
    # 전략 열은 넓게 — 병합 (D~G)
    ws.merge_cells("D24:G24")

    asset_data = [
        ("미국 성장주", "긍정", "금리 인하기 밸류에이션 상승 기대"),
        ("미국 가치주", "긍정", "경기 확장기 수혜"),
        ("장기 채권",   "긍정", "듀레이션 확대 유리"),
        ("금/귀금속",   "긍정", "인플레이션 헤지"),
        ("미국 달러",   "부정", "달러 약세 가능성"),
        ("암호화폐",    "긍정", "유동성 확대기 위험자산 선호"),
    ]
    for i, (asset, effect, strategy) in enumerate(asset_data):
        r = 25 + i
        bg = C_LIGHT_BLUE if i % 2 == 0 else C_WHITE
        write_cell(ws, r, 2, asset, bold=True, bg=bg, border=True)

        effect_color = C_DARK_GREEN if effect == "긍정" else C_DARK_RED
        eff_cell = ws.cell(row=r, column=3, value=effect)
        eff_cell.font = font(bold=True, color=effect_color)
        eff_cell.fill = fill(bg)
        eff_cell.alignment = center_align()
        eff_cell.border = thin_border()

        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        strat_cell = ws.cell(row=r, column=4, value=strategy)
        strat_cell.font = font(size=10)
        strat_cell.fill = fill(bg)
        strat_cell.alignment = left_align()
        strat_cell.border = thin_border()
        set_row_height(ws, r, 18)


# ---------------------------------------------------------------------------
# Sheet 2: 섹터별 현황
# ---------------------------------------------------------------------------

SECTOR_DATA = [
    # (섹터, 종목, 레짐, 심리점수, 심리라벨, 종합판단)
    ("IT/반도체", "삼성전자",       "횡보",    "+23.2", "약간긍정", "관망. 눌림목 매수 기회 모색"),
    ("IT/반도체", "SK하이닉스",     "약한하락", "+33.2", "긍정",    "레짐-심리 괴리. 반등 가능성 주목"),
    ("IT/반도체", "네이버",         "약한상승", "-21.8", "약세",    "상승세이나 심리 부정적. 선별적 접근"),
    ("IT/반도체", "카카오",         "횡보",    "-31.8", "공포",    "약세 지속. 관망"),
    ("방산",      "한화에어로스페이스", "강한하락", "+18.2", "약간긍정", "지정학 수혜 기대. 저가 매수 관점 검토"),
    ("방산",      "LIG넥스원",      "하락",    "+23.2", "약간긍정", "하락세 지속. 지정학 모멘텀 주시"),
    ("방산",      "현대로템",       "하락",    "-26.8", "약세",    "약세 지속"),
    ("배터리",    "LG에너지솔루션", "하락",    "-31.8", "공포",    "전 종목 약세. 회피"),
    ("배터리",    "삼성SDI",        "하락",    "-4.5",  "중립",    "하락세 지속"),
    ("배터리",    "에코프로비엠",   "약한하락", "-26.8", "약세",    "약세"),
    ("암호화폐",  "비트코인",       "약한상승", "-10.7", "약세",    "레짐 상승이나 심리 부정적"),
    ("암호화폐",  "이더리움",       "약한하락", "-9.1",  "중립",    "방향성 없음"),
    ("암호화폐",  "리플",           "약한상승", "-5.2",  "중립",    "소폭 긍정"),
]


def build_sector_sheet(wb):
    ws = wb.create_sheet("섹터별 현황")
    ws.sheet_view.showGridLines = False

    # 열 너비
    for col, width in [("A", 3), ("B", 16), ("C", 20), ("D", 14),
                        ("E", 12), ("F", 14), ("G", 40), ("H", 3)]:
        set_col_width(ws, col, width)

    # 타이틀
    ws.merge_cells("B2:G2")
    tc = ws.cell(row=2, column=2, value="섹터별 종목 현황")
    tc.font = Font(name=FONT_NAME, bold=True, size=14, color=C_WHITE)
    tc.fill = fill(C_HEADER_DARK)
    tc.alignment = center_align()
    set_row_height(ws, 2, 30)

    set_row_height(ws, 3, 8)

    # 헤더
    headers = ["섹터", "종목", "현재 레짐", "심리 점수", "심리 라벨", "종합 판단"]
    table_header_row(ws, 4, headers, col_start=2)
    set_row_height(ws, 4, 22)

    for i, row_data in enumerate(SECTOR_DATA):
        r = 5 + i
        sector, ticker, regime, score, label, judgment = row_data
        bg = regime_bg(regime)

        for c_idx, val in enumerate(row_data):
            col = 2 + c_idx
            bold = (c_idx == 0)  # 섹터 열 볼드
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = font(bold=bold, size=10)
            cell.fill = fill(bg)
            cell.alignment = left_align(wrap=True) if c_idx == 5 else center_align()
            cell.border = thin_border()

        set_row_height(ws, r, 20)

    # 범례
    legend_row = 5 + len(SECTOR_DATA) + 2
    set_row_height(ws, legend_row - 1, 10)
    section_header(ws, legend_row, 2, 7, "  레짐 색상 범례")
    legends = [
        (C_GREEN,  "강한상승 / 상승"),
        (C_YELLOW, "약한상승 / 횡보"),
        (C_ORANGE, "약한하락"),
        (C_RED,    "하락 / 강한하락"),
    ]
    for j, (color, label) in enumerate(legends):
        r2 = legend_row + 1 + j
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=7)
        lc = ws.cell(row=r2, column=2, value=f"    {label}")
        lc.font = font(size=10)
        lc.fill = fill(color)
        lc.border = thin_border()
        set_row_height(ws, r2, 18)


# ---------------------------------------------------------------------------
# Sheet 3: 핵심 포인트
# ---------------------------------------------------------------------------

KEY_POINTS = [
    (1, "매크로 긍정적",
     "금리 인하기 + 경기 확장기로 자산시장에 유리한 환경. "
     "유동성 확대 국면에서 성장주·채권·금 등 다양한 자산의 상승 여건이 형성되어 있음."),
    (2, "배터리 섹터 최악",
     "LG에너지솔루션·삼성SDI·에코프로비엠 전 종목 하락 레짐 + 공포/약세 심리. "
     "당분간 회피 권장. 섹터 전반적 수요 둔화 및 공급 과잉 우려."),
    (3, "SK하이닉스 주목",
     "레짐은 약한하락이지만 심리 점수 +33.2(긍정). 레짐-심리 괴리가 반등 시그널 가능. "
     "HBM 수요 견조 등 펀더멘털 개선 기대 시 저점 매수 검토."),
    (4, "방산 괴리",
     "지정학 리스크 온도계 58.9(주의) 수준으로 방산 섹터 수혜 기대되나, "
     "주가는 하락 레짐 지속 중. 모멘텀 확인 후 저가 매수 관점에서 접근 권장."),
    (5, "암호화폐 중립",
     "비트코인·리플 약한상승 레짐이나 심리 지표는 여전히 부정적. "
     "금리 인하 환경에서 점진적 회복 가능하나 변동성 리스크 유의."),
]


def build_key_points_sheet(wb):
    ws = wb.create_sheet("핵심 포인트")
    ws.sheet_view.showGridLines = False

    for col, width in [("A", 3), ("B", 8), ("C", 28), ("D", 70), ("E", 3)]:
        set_col_width(ws, col, width)

    # 타이틀
    ws.merge_cells("B2:D2")
    tc = ws.cell(row=2, column=2, value="핵심 인사이트 5선")
    tc.font = Font(name=FONT_NAME, bold=True, size=14, color=C_WHITE)
    tc.fill = fill(C_HEADER_DARK)
    tc.alignment = center_align()
    set_row_height(ws, 2, 30)

    set_row_height(ws, 3, 8)

    # 헤더
    for col, label in [(2, "번호"), (3, "제목"), (4, "상세 설명")]:
        cell = ws.cell(row=4, column=col, value=label)
        cell.font = font(bold=True, size=10, color=C_WHITE)
        cell.fill = fill(C_HEADER_DARK)
        cell.alignment = center_align()
        cell.border = thin_border()
    set_row_height(ws, 4, 22)

    row_colors = [C_LIGHT_BLUE, C_WHITE]
    for i, (num, title, desc) in enumerate(KEY_POINTS):
        r = 5 + i
        bg = row_colors[i % 2]

        num_cell = ws.cell(row=r, column=2, value=num)
        num_cell.font = Font(name=FONT_NAME, bold=True, size=14, color=C_WHITE)
        num_cell.fill = fill(C_ACCENT)
        num_cell.alignment = center_align()
        num_cell.border = thin_border()

        title_cell = ws.cell(row=r, column=3, value=title)
        title_cell.font = font(bold=True, size=12)
        title_cell.fill = fill(bg)
        title_cell.alignment = center_align(wrap=True)
        title_cell.border = thin_border()

        desc_cell = ws.cell(row=r, column=4, value=desc)
        desc_cell.font = font(size=10)
        desc_cell.fill = fill(bg)
        desc_cell.alignment = left_align(wrap=True)
        desc_cell.border = thin_border()

        set_row_height(ws, r, 54)


# ---------------------------------------------------------------------------
# 메인
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(description="일일 시황 종합 Excel 보고서 생성")
    parser.add_argument(
        "--date",
        default=datetime.now().strftime("%Y%m%d"),
        help="분석 날짜 (YYYYMMDD). 기본값: 오늘",
    )
    parser.add_argument(
        "--daily-dir",
        default=None,
        help="출력 디렉토리 경로. 기본값: output/daily/{date}",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    date_str = args.date

    # 경로 결정
    base_dir = Path(__file__).parent
    if args.daily_dir:
        daily_dir = Path(args.daily_dir)
    else:
        daily_dir = base_dir / "output" / "daily" / date_str

    daily_dir.mkdir(parents=True, exist_ok=True)

    date_label = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    output_path = daily_dir / f"종합보고서_{date_str}.xlsx"

    print(f"Excel 종합보고서 생성 중 ({date_label})...", end=" ", flush=True)

    wb = Workbook()
    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_dashboard(wb, date_label)
    build_sector_sheet(wb)
    build_key_points_sheet(wb)

    # 첫 번째 시트를 활성화
    wb.active = wb["대시보드"]

    wb.save(str(output_path))
    print("완료")
    print(f"  저장 위치: {output_path}")


if __name__ == "__main__":
    main()
