"""
복합 자산 포트폴리오 분석 및 최적화 시스템
==========================================
사용법:
  1. pip install yfinance pandas numpy openpyxl
  2. python portfolio_analyzer.py --tickers 005930,000660,035420
  3. python portfolio_analyzer.py --tickers 005930,AAPL,BTC-USD --weights 0.5,0.3,0.2
  4. python portfolio_analyzer.py --tickers 005930,AAPL,BTC-USD --period 2y

생성 파일:
  - output/포트폴리오분석_[날짜].xlsx
"""

import os
import sys
import argparse
import warnings
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')

# -- 필수 라이브러리 설치 확인 ----------------------------------------
def check_and_install():
    required = {
        'yfinance': 'yfinance',
        'pandas': 'pandas',
        'numpy': 'numpy',
        'openpyxl': 'openpyxl',
    }
    missing = []
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[설치 필요] pip install {' '.join(missing)}")
        sys.exit(1)

check_and_install()

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- 출력 폴더 생성 ---------------------------------------------------
Path("output").mkdir(exist_ok=True)


# ====================================================================
# 1. 포트폴리오 데이터 수집기
# ====================================================================
class PortfolioDataCollector:

    def __init__(self, tickers: list, period: str = '2y'):
        self.tickers_raw = tickers
        self.period = period
        self.tickers = []           # yfinance용 티커 (감지 완료)
        self.ticker_names = {}      # ticker -> 표시 이름
        self.ticker_markets = {}    # ticker -> 'KR' / 'US' / 'CRYPTO'
        self.price_data = {}        # ticker -> DataFrame
        self.aligned_prices = None  # 날짜 정렬된 종가 DataFrame
        self.daily_returns = None   # 일간 수익률 DataFrame
        self.failed_tickers = []

        for t in tickers:
            yf_ticker, market, display_name = self._resolve_ticker(t)
            if yf_ticker:
                self.tickers.append(yf_ticker)
                self.ticker_names[yf_ticker] = display_name
                self.ticker_markets[yf_ticker] = market

    @staticmethod
    def _detect_market(ticker_code):
        """코스피(.KS) / 코스닥(.KQ) 자동 감지"""
        for suffix in [".KS", ".KQ"]:
            try:
                t = yf.Ticker(ticker_code + suffix)
                hist = t.history(period="5d")
                if not hist.empty and len(hist) > 0:
                    info = t.info or {}
                    name = info.get('longName') or info.get('shortName', '')
                    if name and not name.startswith(ticker_code):
                        return ticker_code + suffix
            except Exception:
                pass
        print(f"  [!] 시장 자동 감지 실패, 코스피(.KS)로 시도합니다.")
        return ticker_code + ".KS"

    def _resolve_ticker(self, raw: str) -> tuple:
        """원본 티커 -> (yfinance 티커, 시장, 표시 이름)"""
        raw = raw.strip()
        if raw.isdigit():
            # 한국 주식
            yf_ticker = self._detect_market(raw)
            stock = yf.Ticker(yf_ticker)
            name = stock.info.get('longName') or stock.info.get('shortName', raw)
            return yf_ticker, 'KR', name
        elif '-' in raw and raw.upper().endswith('USD'):
            # 암호화폐 (BTC-USD, ETH-USD 등)
            stock = yf.Ticker(raw.upper())
            name = stock.info.get('longName') or stock.info.get('shortName', raw.upper())
            return raw.upper(), 'CRYPTO', name
        else:
            # 미국/해외 주식
            stock = yf.Ticker(raw.upper())
            name = stock.info.get('longName') or stock.info.get('shortName', raw.upper())
            return raw.upper(), 'US', name

    def fetch_all(self) -> bool:
        """모든 자산의 가격 데이터 수집 및 정렬"""
        print(f"\n[수집] 포트폴리오 데이터 수집 시작 ({len(self.tickers)}개 자산)")

        valid_tickers = []
        for ticker in self.tickers:
            display = self.ticker_names.get(ticker, ticker)
            print(f"  |- {display} ({ticker}) 수집 중...")
            try:
                stock = yf.Ticker(ticker)
                df = stock.history(period=self.period)
                if df.empty or len(df) < 10:
                    print(f"  |  [경고] {display} - 데이터 부족, 건너뜁니다.")
                    self.failed_tickers.append(ticker)
                    continue
                # 타임존 제거
                if df.index.tz is not None:
                    df.index = df.index.tz_localize(None)
                self.price_data[ticker] = df
                valid_tickers.append(ticker)
                print(f"  |  수집 완료: {len(df)}일")
            except Exception as e:
                print(f"  |  [경고] {display} 수집 실패: {e}")
                self.failed_tickers.append(ticker)

        self.tickers = valid_tickers

        if len(self.tickers) == 0:
            print("  [오류] 유효한 자산이 없습니다.")
            return False

        # 종가 데이터 정렬 (inner join)
        print(f"\n  |- 날짜 정렬 중 (inner join)...")
        close_frames = {}
        for ticker in self.tickers:
            series = self.price_data[ticker]['Close'].copy()
            series.name = ticker
            close_frames[ticker] = series

        # 합치기 - 캘린더 날짜 기준으로 forward fill 후 inner join
        aligned = pd.DataFrame(close_frames)

        # 암호화폐가 포함된 경우 calendar day reindex 후 ffill
        has_crypto = any(self.ticker_markets.get(t) == 'CRYPTO' for t in self.tickers)
        if has_crypto:
            # 전체 날짜 범위로 리인덱스
            full_range = pd.date_range(aligned.index.min(), aligned.index.max(), freq='D')
            aligned = aligned.reindex(full_range)
            aligned = aligned.ffill()

        aligned = aligned.dropna()

        if len(aligned) < 60:
            print(f"  [경고] 정렬 후 데이터 {len(aligned)}일 - 최소 60일 필요")
            if len(aligned) < 20:
                print(f"  [오류] 분석 불가능한 데이터 수준입니다.")
                return False

        self.aligned_prices = aligned
        self.daily_returns = aligned.pct_change().dropna()
        print(f"  |  정렬 완료: {len(aligned)}일, 자산 {len(self.tickers)}개")
        return True

    def get_majority_market(self) -> str:
        """자산 중 다수 시장 반환 (벤치마크 선택용)"""
        markets = [self.ticker_markets.get(t, 'US') for t in self.tickers]
        kr_count = markets.count('KR')
        us_count = markets.count('US') + markets.count('CRYPTO')
        return 'KR' if kr_count > us_count else 'US'


# ====================================================================
# 2. 포트폴리오 분석기
# ====================================================================
class PortfolioAnalyzer:

    RISK_FREE_RATE = 0.035  # 연간 무위험 수익률 3.5%
    TRADING_DAYS = 252

    def __init__(self, aligned_prices: pd.DataFrame, daily_returns: pd.DataFrame,
                 weights: list = None, ticker_names: dict = None):
        self.prices = aligned_prices
        self.returns = daily_returns
        self.tickers = list(aligned_prices.columns)
        self.n_assets = len(self.tickers)
        self.ticker_names = ticker_names or {}

        # 가중치 설정 (기본: 동일비중)
        if weights and len(weights) == self.n_assets:
            self.weights = np.array(weights)
        else:
            self.weights = np.ones(self.n_assets) / self.n_assets

        # 가중치 정규화
        self.weights = self.weights / self.weights.sum()

    def _display_name(self, ticker: str) -> str:
        """표시용 이름 반환"""
        name = self.ticker_names.get(ticker, ticker)
        # 이름이 너무 길면 축약
        if len(name) > 20:
            name = name[:18] + '..'
        return name

    # ----------------------------------------------------------------
    # a) 개별 자산 분석
    # ----------------------------------------------------------------
    def analyze_individual_assets(self) -> list:
        """각 자산의 핵심 지표 계산"""
        results = []
        for ticker in self.tickers:
            ret = self.returns[ticker]
            price = self.prices[ticker]

            # 연간 수익률
            total_return = (price.iloc[-1] / price.iloc[0]) - 1
            n_years = len(price) / self.TRADING_DAYS
            ann_return = (1 + total_return) ** (1 / max(n_years, 0.01)) - 1

            # 연간 변동성
            ann_vol = ret.std() * np.sqrt(self.TRADING_DAYS)

            # Sharpe 비율
            sharpe = (ann_return - self.RISK_FREE_RATE) / ann_vol if ann_vol > 0 else 0

            # Sortino 비율 (하락 변동성만 고려)
            downside_ret = ret[ret < 0]
            downside_vol = downside_ret.std() * np.sqrt(self.TRADING_DAYS) if len(downside_ret) > 0 else 0
            sortino = (ann_return - self.RISK_FREE_RATE) / downside_vol if downside_vol > 0 else 0

            # 최대 낙폭 (Max Drawdown)
            cummax = price.cummax()
            drawdown = (price - cummax) / cummax
            max_dd = drawdown.min()

            # Calmar 비율
            calmar = ann_return / abs(max_dd) if abs(max_dd) > 0 else 0

            # 최고/최악일
            best_day = ret.max()
            worst_day = ret.min()
            best_day_date = ret.idxmax().strftime('%Y-%m-%d')
            worst_day_date = ret.idxmin().strftime('%Y-%m-%d')

            # 월간 수익률
            monthly = price.resample('ME').last().pct_change().dropna()
            best_month = monthly.max() if len(monthly) > 0 else 0
            worst_month = monthly.min() if len(monthly) > 0 else 0

            results.append({
                'ticker': ticker,
                'name': self._display_name(ticker),
                'ann_return': ann_return,
                'ann_volatility': ann_vol,
                'sharpe': sharpe,
                'sortino': sortino,
                'max_drawdown': max_dd,
                'calmar': calmar,
                'best_day': best_day,
                'best_day_date': best_day_date,
                'worst_day': worst_day,
                'worst_day_date': worst_day_date,
                'best_month': best_month,
                'worst_month': worst_month,
                'total_return': total_return,
            })
        return results

    # ----------------------------------------------------------------
    # b) 상관관계 분석
    # ----------------------------------------------------------------
    def analyze_correlations(self) -> dict:
        """상관관계 매트릭스 및 롤링 상관관계"""
        corr_matrix = self.returns.corr()

        # 롤링 60일 상관관계 (상위/하위 쌍)
        rolling_corr = {}
        if self.n_assets >= 2:
            for i in range(self.n_assets):
                for j in range(i + 1, self.n_assets):
                    t1, t2 = self.tickers[i], self.tickers[j]
                    rc = self.returns[t1].rolling(60).corr(self.returns[t2]).dropna()
                    if len(rc) > 0:
                        rolling_corr[f"{self._display_name(t1)} / {self._display_name(t2)}"] = {
                            'current': rc.iloc[-1] if len(rc) > 0 else 0,
                            'mean': rc.mean(),
                            'min': rc.min(),
                            'max': rc.max(),
                        }

        # 가장 높은/낮은 상관관계 쌍 찾기
        pairs = []
        for i in range(self.n_assets):
            for j in range(i + 1, self.n_assets):
                pairs.append((self.tickers[i], self.tickers[j],
                              corr_matrix.loc[self.tickers[i], self.tickers[j]]))

        pairs.sort(key=lambda x: x[2], reverse=True)
        highest_pair = pairs[0] if pairs else None
        lowest_pair = pairs[-1] if pairs else None

        return {
            'matrix': corr_matrix,
            'rolling': rolling_corr,
            'highest_pair': highest_pair,
            'lowest_pair': lowest_pair,
        }

    # ----------------------------------------------------------------
    # c) 포트폴리오 성과
    # ----------------------------------------------------------------
    def analyze_portfolio_performance(self) -> dict:
        """현재 가중치 기준 포트폴리오 성과"""
        return self._calc_portfolio_metrics(self.weights)

    def _calc_portfolio_metrics(self, weights: np.ndarray) -> dict:
        """주어진 가중치로 포트폴리오 지표 계산"""
        weights = np.array(weights)

        # 포트폴리오 일간 수익률
        port_returns = (self.returns * weights).sum(axis=1)

        # 누적 수익률
        cumulative = (1 + port_returns).cumprod()
        total_return = cumulative.iloc[-1] - 1

        # 연간 수익률
        n_years = len(port_returns) / self.TRADING_DAYS
        ann_return = (1 + total_return) ** (1 / max(n_years, 0.01)) - 1

        # 연간 변동성
        ann_vol = port_returns.std() * np.sqrt(self.TRADING_DAYS)

        # Sharpe
        sharpe = (ann_return - self.RISK_FREE_RATE) / ann_vol if ann_vol > 0 else 0

        # Sortino (하락 변동성만 고려)
        downside_ret = port_returns[port_returns < 0]
        downside_vol = downside_ret.std() * np.sqrt(self.TRADING_DAYS) if len(downside_ret) > 0 else 0
        sortino = (ann_return - self.RISK_FREE_RATE) / downside_vol if downside_vol > 0 else 0

        # 최대 낙폭
        cummax = cumulative.cummax()
        drawdown = (cumulative - cummax) / cummax
        max_dd = drawdown.min()

        return {
            'weights': weights,
            'total_return': total_return,
            'ann_return': ann_return,
            'ann_volatility': ann_vol,
            'sharpe': sharpe,
            'sortino': sortino,
            'max_drawdown': max_dd,
            'port_returns': port_returns,
            'cumulative': cumulative,
            'drawdown_series': drawdown,
        }

    def compare_portfolios(self) -> dict:
        """현재 vs 동일비중 vs 개별자산 비교"""
        current = self._calc_portfolio_metrics(self.weights)
        equal_w = np.ones(self.n_assets) / self.n_assets
        equal = self._calc_portfolio_metrics(equal_w)

        return {
            'current': current,
            'equal_weight': equal,
        }

    # ----------------------------------------------------------------
    # d) 최적 포트폴리오 (Monte Carlo)
    # ----------------------------------------------------------------
    def optimize_monte_carlo(self, n_simulations: int = 10000) -> dict:
        """몬테카를로 시뮬레이션으로 최적 포트폴리오 탐색"""
        print(f"  |- 몬테카를로 시뮬레이션 ({n_simulations:,}회)...")

        mean_returns = self.returns.mean() * self.TRADING_DAYS
        cov_matrix = self.returns.cov() * self.TRADING_DAYS

        results = np.zeros((n_simulations, 3 + self.n_assets))
        # columns: return, volatility, sharpe, weight_1, ..., weight_n

        np.random.seed(42)
        for i in range(n_simulations):
            # 랜덤 가중치 생성
            w = np.random.random(self.n_assets)
            w = w / w.sum()

            # 포트폴리오 수익률
            port_return = np.dot(w, mean_returns)
            # 포트폴리오 변동성
            port_vol = np.sqrt(np.dot(w.T, np.dot(cov_matrix, w)))
            # Sharpe
            port_sharpe = (port_return - self.RISK_FREE_RATE) / port_vol if port_vol > 0 else 0

            results[i, 0] = port_return
            results[i, 1] = port_vol
            results[i, 2] = port_sharpe
            results[i, 3:] = w

        # 결과 DataFrame
        cols = ['return', 'volatility', 'sharpe'] + [f'w_{t}' for t in self.tickers]
        sim_df = pd.DataFrame(results, columns=cols)

        # 최대 Sharpe 포트폴리오
        max_sharpe_idx = sim_df['sharpe'].idxmax()
        max_sharpe = sim_df.loc[max_sharpe_idx]

        # 최소 변동성 포트폴리오
        min_vol_idx = sim_df['volatility'].idxmin()
        min_vol = sim_df.loc[min_vol_idx]

        # 공격형: 수익률이 높으면서 Sharpe > 0.5인 것 중 최대 수익률
        aggressive_candidates = sim_df[sim_df['sharpe'] > 0.5]
        if len(aggressive_candidates) == 0:
            # Sharpe 제한이 너무 높으면 상위 10%에서 최대 수익률
            top_sharpe = sim_df.nlargest(int(n_simulations * 0.1), 'sharpe')
            aggressive = top_sharpe.loc[top_sharpe['return'].idxmax()]
        else:
            aggressive = aggressive_candidates.loc[aggressive_candidates['return'].idxmax()]

        print(f"  |  시뮬레이션 완료")

        return {
            'simulations': sim_df,
            'max_sharpe': {
                'weights': np.array([max_sharpe[f'w_{t}'] for t in self.tickers]),
                'return': max_sharpe['return'],
                'volatility': max_sharpe['volatility'],
                'sharpe': max_sharpe['sharpe'],
            },
            'min_volatility': {
                'weights': np.array([min_vol[f'w_{t}'] for t in self.tickers]),
                'return': min_vol['return'],
                'volatility': min_vol['volatility'],
                'sharpe': min_vol['sharpe'],
            },
            'aggressive': {
                'weights': np.array([aggressive[f'w_{t}'] for t in self.tickers]),
                'return': aggressive['return'],
                'volatility': aggressive['volatility'],
                'sharpe': aggressive['sharpe'],
            },
        }

    # ----------------------------------------------------------------
    # e) 리스크 분석
    # ----------------------------------------------------------------
    def analyze_risk(self, benchmark_ticker: str = None) -> dict:
        """VaR, CVaR, Beta 분석"""
        port_returns = (self.returns * self.weights).sum(axis=1)

        # VaR (Historical)
        var_95 = np.percentile(port_returns, 5)
        var_99 = np.percentile(port_returns, 1)

        # CVaR (Expected Shortfall)
        cvar_95 = port_returns[port_returns <= var_95].mean()
        cvar_99 = port_returns[port_returns <= var_99].mean()

        # 포트폴리오 Beta vs 벤치마크
        beta = None
        benchmark_name = None
        if benchmark_ticker:
            try:
                bm = yf.Ticker(benchmark_ticker)
                bm_data = bm.history(period=self.returns.index[0].strftime('%Y-%m-%d'))
                # 기간 맞추기 - 직접 period 사용
                bm_data2 = bm.history(start=self.returns.index[0], end=self.returns.index[-1])
                if len(bm_data2) > 30:
                    if bm_data2.index.tz is not None:
                        bm_data2.index = bm_data2.index.tz_localize(None)
                    bm_returns = bm_data2['Close'].pct_change().dropna()
                    # 공통 날짜
                    common_idx = port_returns.index.intersection(bm_returns.index)
                    if len(common_idx) > 30:
                        pr = port_returns.loc[common_idx]
                        br = bm_returns.loc[common_idx]
                        cov_pb = np.cov(pr, br)[0, 1]
                        var_b = np.var(br)
                        beta = cov_pb / var_b if var_b > 0 else None
                        benchmark_name = '^KS11' if benchmark_ticker == '^KS11' else '^GSPC'
            except Exception:
                pass

        # 스트레스 시나리오: 시장 10% 하락시 예상 손실
        stress_loss = None
        if beta is not None:
            stress_loss = beta * (-0.10)  # 시장 -10% 시 포트폴리오 예상 변동

        return {
            'var_95': var_95,
            'var_99': var_99,
            'cvar_95': cvar_95,
            'cvar_99': cvar_99,
            'beta': beta,
            'benchmark': benchmark_name,
            'stress_loss_10pct': stress_loss,
        }

    # ----------------------------------------------------------------
    # f) 벤치마크 비교 (Buy & Hold vs SMA200)
    # ----------------------------------------------------------------
    def calculate_benchmarks(self) -> dict:
        """Buy & Hold 및 SMA200 전략과의 벤치마크 비교를 수행합니다."""
        benchmarks = {}

        if self.returns.empty or len(self.returns) < 200:
            return benchmarks

        # 가중치 기반 포트폴리오 일간 수익률
        portfolio_returns = (self.returns * self.weights).sum(axis=1)

        # 1. Buy & Hold: cumulative returns
        bh_cumulative = (1 + portfolio_returns).cumprod()
        bh_total_return = (bh_cumulative.iloc[-1] - 1) * 100
        bh_annual_return = ((bh_cumulative.iloc[-1]) ** (252 / len(portfolio_returns)) - 1) * 100
        bh_volatility = portfolio_returns.std() * np.sqrt(252) * 100
        bh_sharpe = bh_annual_return / bh_volatility if bh_volatility > 0 else 0
        bh_max_dd = ((bh_cumulative / bh_cumulative.cummax()) - 1).min() * 100

        benchmarks['buy_hold'] = {
            'total_return': round(bh_total_return, 1),
            'annual_return': round(bh_annual_return, 1),
            'volatility': round(bh_volatility, 1),
            'sharpe': round(bh_sharpe, 2),
            'max_drawdown': round(bh_max_dd, 1),
        }

        # 2. SMA200 Strategy
        prices = bh_cumulative  # proxy for portfolio price
        sma200 = prices.rolling(200).mean()

        # Signal: 1 when price > SMA200, 0 otherwise (cash); shift(1) to avoid look-ahead bias
        signal = (prices > sma200).astype(float)
        signal = signal.shift(1).fillna(0)

        sma_returns = portfolio_returns * signal
        sma_cumulative = (1 + sma_returns).cumprod()
        sma_total_return = (sma_cumulative.iloc[-1] - 1) * 100
        sma_annual_return = ((sma_cumulative.iloc[-1]) ** (252 / len(sma_returns)) - 1) * 100
        active_returns = sma_returns[sma_returns != 0]
        sma_volatility = active_returns.std() * np.sqrt(252) * 100 if len(active_returns) > 0 else 0
        sma_sharpe = sma_annual_return / sma_volatility if sma_volatility > 0 else 0
        sma_max_dd = ((sma_cumulative / sma_cumulative.cummax()) - 1).min() * 100

        benchmarks['sma200'] = {
            'total_return': round(sma_total_return, 1),
            'annual_return': round(sma_annual_return, 1),
            'volatility': round(sma_volatility, 1),
            'sharpe': round(sma_sharpe, 2),
            'max_drawdown': round(sma_max_dd, 1),
        }

        return benchmarks

    # ----------------------------------------------------------------
    # g) 자산배분 제안
    # ----------------------------------------------------------------
    def suggest_allocations(self, mc_result: dict) -> list:
        """3가지 포트폴리오 제안: 공격형, 균형형, 안정형"""
        suggestions = []

        # 공격형: MC 결과의 aggressive
        agg = mc_result['aggressive']
        agg_metrics = self._calc_portfolio_metrics(agg['weights'])
        suggestions.append({
            'name': '공격형',
            'description': '높은 수익률 추구, Sharpe > 0.5 이상 유지',
            'weights': agg['weights'],
            'ann_return': agg['return'],
            'ann_volatility': agg['volatility'],
            'sharpe': agg['sharpe'],
            'var_95': np.percentile(agg_metrics['port_returns'], 5),
        })

        # 균형형: 최대 Sharpe
        bal = mc_result['max_sharpe']
        bal_metrics = self._calc_portfolio_metrics(bal['weights'])
        suggestions.append({
            'name': '균형형',
            'description': '위험 대비 수익 최적화 (최대 Sharpe)',
            'weights': bal['weights'],
            'ann_return': bal['return'],
            'ann_volatility': bal['volatility'],
            'sharpe': bal['sharpe'],
            'var_95': np.percentile(bal_metrics['port_returns'], 5),
        })

        # 안정형: 최소 변동성
        con = mc_result['min_volatility']
        con_metrics = self._calc_portfolio_metrics(con['weights'])
        suggestions.append({
            'name': '안정형',
            'description': '변동성 최소화, 자산 보전 우선',
            'weights': con['weights'],
            'ann_return': con['return'],
            'ann_volatility': con['volatility'],
            'sharpe': con['sharpe'],
            'var_95': np.percentile(con_metrics['port_returns'], 5),
        })

        return suggestions


# ====================================================================
# 3. Excel 보고서 생성기
# ====================================================================
class PortfolioExcelBuilder:

    # 색상 팔레트 (financial_analyzer.py 동일)
    COLORS = {
        'header_dark': '0D1B2A',
        'header_blue': '1B4F72',
        'accent': '2E86C1',
        'light_blue': 'D6EAF8',
        'white': 'FFFFFF',
        'light_gray': 'F2F3F4',
        'positive': 'EAFAF1',
        'negative': 'FDEDEC',
        'gold': 'F39C12',
    }

    # 상관관계 색상 그라데이션
    CORR_COLORS = [
        (-1.0, 'FDEDEC'),   # 음의 상관
        (-0.3, 'FADBD8'),
        (0.0,  'FFFFFF'),
        (0.3,  'D6EAF8'),
        (0.6,  '85C1E9'),
        (0.8,  '2E86C1'),
        (1.0,  '1B4F72'),
    ]

    def __init__(self, ticker_names: dict):
        self.wb = Workbook()
        self.ticker_names = ticker_names
        self.wb.remove(self.wb.active)

    def _style_header(self, ws, row, col, value, color='header_dark', font_color='FFFFFF', bold=True, size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=self.COLORS.get(color, color))
        cell.font = Font(bold=bold, color=font_color, size=size, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def _style_data(self, ws, row, col, value, bg_color=None, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if number_format:
            cell.number_format = number_format
        return cell

    def _corr_color(self, val: float) -> str:
        """상관계수에 따른 배경색"""
        for threshold, color in reversed(self.CORR_COLORS):
            if val >= threshold:
                return color
        return 'FFFFFF'

    def _corr_font_color(self, val: float) -> str:
        """상관계수에 따른 글자색"""
        if val >= 0.6 or val <= -0.6:
            return 'FFFFFF'
        return '0D1B2A'

    def _pct_str(self, val: float) -> str:
        """퍼센트 문자열 (부호 포함)"""
        if val >= 0:
            return f"+{val*100:.1f}%"
        return f"{val*100:.1f}%"

    # ----------------------------------------------------------------
    # 시트 1: 포트폴리오 요약
    # ----------------------------------------------------------------
    def add_summary_sheet(self, portfolio_perf: dict, comparison: dict,
                          risk: dict, suggestions: list, tickers: list,
                          weights: np.ndarray, start_date: str, end_date: str):
        ws = self.wb.create_sheet("포트폴리오 요약")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = "복합 자산 포트폴리오 분석 보고서"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 날짜
        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  분석 기간: {start_date} ~ {end_date}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 자산 구성
        row = 5
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "자산 구성", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        asset_headers = ['자산', '티커', '비중(%)']
        for col, h in enumerate(asset_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        for i, ticker in enumerate(tickers):
            bg = 'light_gray' if i % 2 == 0 else 'white'
            name = self.ticker_names.get(ticker, ticker)
            self._style_data(ws, row, 2, name, bg_color=bg)
            self._style_data(ws, row, 3, ticker, bg_color=bg)
            self._style_data(ws, row, 4, f"{weights[i]*100:.1f}%", bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 포트폴리오 성과
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "포트폴리오 성과", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        perf = portfolio_perf
        perf_items = [
            ('연간 수익률', f"{perf['ann_return']*100:.1f}%"),
            ('연간 변동성', f"{perf['ann_volatility']*100:.1f}%"),
            ('Sharpe Ratio', f"{perf['sharpe']:.2f}"),
            ('Sortino Ratio', f"{perf.get('sortino', 0):.2f}"),
            ('누적 수익률', f"{perf['total_return']*100:.1f}%"),
            ('최대 낙폭(MDD)', f"{perf['max_drawdown']*100:.1f}%"),
        ]

        for key, val in perf_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            cell_k = ws.cell(row=row, column=2, value=key)
            cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
            cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(f'C{row}:D{row}')
            self._style_data(ws, row, 3, val, bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 비교 테이블: 현재 vs 최적(균형형) vs 동일비중
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "포트폴리오 비교", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        comp_headers = ['', '현재 포트폴리오', '최적(균형형)', '동일비중']
        for col, h in enumerate(comp_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        # 균형형 데이터 (suggestions[1] = 균형형)
        balanced = suggestions[1] if len(suggestions) > 1 else None
        eq = comparison['equal_weight']

        comp_rows = [
            ('연간 수익률', perf['ann_return'], balanced['ann_return'] if balanced else 0, eq['ann_return']),
            ('연간 변동성', perf['ann_volatility'], balanced['ann_volatility'] if balanced else 0, eq['ann_volatility']),
            ('Sharpe', perf['sharpe'], balanced['sharpe'] if balanced else 0, eq['sharpe']),
            ('최대 낙폭', perf['max_drawdown'], None, eq['max_drawdown']),
        ]

        for label, v1, v2, v3 in comp_rows:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            cell_k = ws.cell(row=row, column=2, value=label)
            cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
            cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            fmt = lambda v: f"{v*100:.1f}%" if v is not None else 'N/A'
            if label == 'Sharpe':
                fmt = lambda v: f"{v:.2f}" if v is not None else 'N/A'

            self._style_data(ws, row, 3, fmt(v1), bg_color=bg)
            self._style_data(ws, row, 4, fmt(v2), bg_color=bg)
            self._style_data(ws, row, 5, fmt(v3), bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 리스크 요약
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "핵심 리스크 지표", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        risk_items = [
            ('VaR(95%)', f"{risk['var_95']*100:.2f}%"),
            ('VaR(99%)', f"{risk['var_99']*100:.2f}%"),
            ('CVaR(95%)', f"{risk['cvar_95']*100:.2f}%"),
            ('최대 낙폭', f"{perf['max_drawdown']*100:.1f}%"),
            ('Beta', f"{risk['beta']:.2f}" if risk['beta'] is not None else 'N/A'),
        ]

        for key, val in risk_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            cell_k = ws.cell(row=row, column=2, value=key)
            cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
            cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(f'C{row}:D{row}')
            self._style_data(ws, row, 3, val, bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        for c in 'FGH':
            ws.column_dimensions[c].width = 15

    # ----------------------------------------------------------------
    # 시트 2: 개별 자산 분석
    # ----------------------------------------------------------------
    def add_individual_sheet(self, asset_results: list):
        ws = self.wb.create_sheet("개별 자산 분석")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:K2')
        title = ws['B2']
        title.value = "개별 자산 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 헤더
        headers = ['자산', '연간수익률', '연간변동성', 'Sharpe', 'MDD',
                   'Calmar', '최고일', '최고일 날짜', '최악일', '최악일 날짜']
        row = 4
        for col, h in enumerate(headers, 2):
            self._style_header(ws, row, col, h, 'header_dark', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        for asset in asset_results:
            # 수익률 양/음에 따라 색상
            ret_bg = 'positive' if asset['ann_return'] >= 0 else 'negative'

            self._style_data(ws, row, 2, asset['name'], bg_color=ret_bg, bold=True)
            self._style_data(ws, row, 3, f"{asset['ann_return']*100:.1f}%", bg_color=ret_bg)
            self._style_data(ws, row, 4, f"{asset['ann_volatility']*100:.1f}%", bg_color=ret_bg)
            self._style_data(ws, row, 5, f"{asset['sharpe']:.2f}", bg_color=ret_bg)
            self._style_data(ws, row, 6, f"{asset['max_drawdown']*100:.1f}%", bg_color=ret_bg)
            self._style_data(ws, row, 7, f"{asset['calmar']:.2f}", bg_color=ret_bg)
            self._style_data(ws, row, 8, f"{asset['best_day']*100:.2f}%", bg_color='positive')
            self._style_data(ws, row, 9, asset['best_day_date'], bg_color='positive')
            self._style_data(ws, row, 10, f"{asset['worst_day']*100:.2f}%", bg_color='negative')
            self._style_data(ws, row, 11, asset['worst_day_date'], bg_color='negative')
            ws.row_dimensions[row].height = 24
            row += 1

        # 월간 수익률 최고/최악
        row += 1
        ws.merge_cells(f'B{row}:K{row}')
        self._style_header(ws, row, 2, "월간 수익률 극값", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        m_headers = ['자산', '최고 월간 수익률', '최악 월간 수익률']
        for col, h in enumerate(m_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        for asset in asset_results:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_data(ws, row, 2, asset['name'], bg_color=bg, bold=True)
            self._style_data(ws, row, 3, f"{asset['best_month']*100:.1f}%", bg_color='positive')
            self._style_data(ws, row, 4, f"{asset['worst_month']*100:.1f}%", bg_color='negative')
            ws.row_dimensions[row].height = 22
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        for c_idx in range(3, 12):
            ws.column_dimensions[get_column_letter(c_idx)].width = 16

    # ----------------------------------------------------------------
    # 시트 3: 상관관계
    # ----------------------------------------------------------------
    def add_correlation_sheet(self, corr_data: dict, tickers: list):
        ws = self.wb.create_sheet("상관관계")
        ws.sheet_view.showGridLines = False

        # 제목
        n = len(tickers)
        merge_end = get_column_letter(3 + n)
        ws.merge_cells(f'B2:{merge_end}2')
        title = ws['B2']
        title.value = "자산 간 상관관계 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 상관관계 매트릭스
        matrix = corr_data['matrix']
        row = 4
        ws.merge_cells(f'B{row}:{merge_end}{row}')
        self._style_header(ws, row, 2, "일간 수익률 상관관계 매트릭스", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        # 좌상단
        self._style_header(ws, row, 2, "", 'header_blue', size=10)
        # 열 헤더
        for j, t in enumerate(tickers):
            name = self.ticker_names.get(t, t)
            if len(name) > 12:
                name = name[:10] + '..'
            self._style_header(ws, row, 3 + j, name, 'header_blue', size=9)
        ws.row_dimensions[row].height = 28
        row += 1

        # 매트릭스 데이터
        for i, t1 in enumerate(tickers):
            name = self.ticker_names.get(t1, t1)
            if len(name) > 12:
                name = name[:10] + '..'
            self._style_header(ws, row, 2, name, 'header_blue', size=9)

            for j, t2 in enumerate(tickers):
                val = matrix.loc[t1, t2]
                bg = self._corr_color(val)
                fc = self._corr_font_color(val)
                cell = ws.cell(row=row, column=3 + j, value=round(val, 3))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(size=10, name='맑은 고딕', color=fc,
                                 bold=(i == j))
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.000'

            ws.row_dimensions[row].height = 24
            row += 1

        # 해석
        row += 1
        ws.merge_cells(f'B{row}:{merge_end}{row}')
        self._style_header(ws, row, 2, "상관관계 해석", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        highest = corr_data['highest_pair']
        lowest = corr_data['lowest_pair']

        if highest:
            n1 = self.ticker_names.get(highest[0], highest[0])
            n2 = self.ticker_names.get(highest[1], highest[1])
            ws.merge_cells(f'B{row}:{merge_end}{row}')
            cell = ws.cell(row=row, column=2,
                           value=f"가장 높은 상관관계: {n1} / {n2} ({highest[2]:.3f}) - 유사한 움직임, 분산 효과 낮음")
            cell.font = Font(size=10, name='맑은 고딕', color='922B21')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            row += 1

        if lowest:
            n1 = self.ticker_names.get(lowest[0], lowest[0])
            n2 = self.ticker_names.get(lowest[1], lowest[1])
            ws.merge_cells(f'B{row}:{merge_end}{row}')
            cell = ws.cell(row=row, column=2,
                           value=f"가장 낮은 상관관계: {n1} / {n2} ({lowest[2]:.3f}) - 좋은 분산 투자 조합")
            cell.font = Font(size=10, name='맑은 고딕', color='1A5276')
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            row += 1

        # 범례
        row += 1
        ws.merge_cells(f'B{row}:{merge_end}{row}')
        cell = ws.cell(row=row, column=2,
                       value="* 색상 진하면 높은 상관 (파랑), 연하면 낮은 상관/음의 상관 (빨강 계열)")
        cell.font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 롤링 상관관계
        rolling = corr_data.get('rolling', {})
        if rolling:
            row += 2
            ws.merge_cells(f'B{row}:{merge_end}{row}')
            self._style_header(ws, row, 2, "롤링 60일 상관관계 요약", 'header_dark', size=11)
            ws.row_dimensions[row].height = 28
            row += 1

            r_headers = ['자산 쌍', '현재', '평균', '최소', '최대']
            for col, h in enumerate(r_headers, 2):
                self._style_header(ws, row, col, h, 'header_blue', size=10)
            ws.row_dimensions[row].height = 24
            row += 1

            for pair_name, vals in rolling.items():
                bg = 'light_gray' if row % 2 == 0 else 'white'
                self._style_data(ws, row, 2, pair_name, bg_color=bg)
                self._style_data(ws, row, 3, f"{vals['current']:.3f}", bg_color=bg)
                self._style_data(ws, row, 4, f"{vals['mean']:.3f}", bg_color=bg)
                self._style_data(ws, row, 5, f"{vals['min']:.3f}", bg_color=bg)
                self._style_data(ws, row, 6, f"{vals['max']:.3f}", bg_color=bg)
                ws.row_dimensions[row].height = 22
                row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        for j in range(n + 3):
            col_letter = get_column_letter(3 + j)
            ws.column_dimensions[col_letter].width = 16

    # ----------------------------------------------------------------
    # 시트 4: 최적화 결과
    # ----------------------------------------------------------------
    def add_optimization_sheet(self, suggestions: list, tickers: list,
                               current_weights: np.ndarray, mc_result: dict):
        ws = self.wb.create_sheet("최적화 결과")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:J2')
        title = ws['B2']
        title.value = "포트폴리오 최적화 결과"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws['B3'] = "Monte Carlo 시뮬레이션 10,000회 기반"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 3가지 포트폴리오 제안
        row = 5
        ws.merge_cells(f'B{row}:J{row}')
        self._style_header(ws, row, 2, "자산배분 제안", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        # 헤더: 포트폴리오 유형, 각 자산 비중, 수익률, 변동성, Sharpe, VaR
        opt_headers = ['유형'] + [self.ticker_names.get(t, t)[:8] for t in tickers] + \
                      ['수익률', '변동성', 'Sharpe', 'VaR(95%)']
        for col, h in enumerate(opt_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=9)
        ws.row_dimensions[row].height = 28
        row += 1

        # 현재 포트폴리오
        self._style_data(ws, row, 2, '현재', bg_color='light_blue', bold=True)
        for i, t in enumerate(tickers):
            self._style_data(ws, row, 3 + i, f"{current_weights[i]*100:.1f}%", bg_color='light_blue')
        ws.row_dimensions[row].height = 24
        row += 1

        # 제안 포트폴리오
        type_colors = {'공격형': 'FDEDEC', '균형형': 'EAFAF1', '안정형': 'D6EAF8'}
        for sug in suggestions:
            bg = type_colors.get(sug['name'], 'white')
            self._style_data(ws, row, 2, sug['name'], bg_color=bg, bold=True)
            for i, t in enumerate(tickers):
                self._style_data(ws, row, 3 + i, f"{sug['weights'][i]*100:.1f}%", bg_color=bg)
            n_t = len(tickers)
            self._style_data(ws, row, 3 + n_t, f"{sug['ann_return']*100:.1f}%", bg_color=bg)
            self._style_data(ws, row, 4 + n_t, f"{sug['ann_volatility']*100:.1f}%", bg_color=bg)
            self._style_data(ws, row, 5 + n_t, f"{sug['sharpe']:.2f}", bg_color=bg)
            self._style_data(ws, row, 6 + n_t, f"{sug['var_95']*100:.2f}%", bg_color=bg)
            ws.row_dimensions[row].height = 24
            row += 1

        # 효율적 프론티어 데이터 (샘플)
        row += 1
        ws.merge_cells(f'B{row}:J{row}')
        self._style_header(ws, row, 2, "효율적 프론티어 데이터 (상위 100개 샘플)", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        ef_headers = ['순번', '수익률(%)', '변동성(%)', 'Sharpe']
        for col, h in enumerate(ef_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        # Sharpe 상위 100개 포트폴리오
        sim_df = mc_result['simulations']
        top100 = sim_df.nlargest(100, 'sharpe')

        for idx, (_, sim_row) in enumerate(top100.iterrows(), 1):
            bg = 'light_gray' if idx % 2 == 0 else 'white'
            self._style_data(ws, row, 2, idx, bg_color=bg)
            self._style_data(ws, row, 3, f"{sim_row['return']*100:.2f}", bg_color=bg)
            self._style_data(ws, row, 4, f"{sim_row['volatility']*100:.2f}", bg_color=bg)
            self._style_data(ws, row, 5, f"{sim_row['sharpe']:.3f}", bg_color=bg)
            ws.row_dimensions[row].height = 18
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 12
        for c_idx in range(3, 3 + len(tickers) + 5):
            ws.column_dimensions[get_column_letter(c_idx)].width = 14

    # ----------------------------------------------------------------
    # 시트 5: 리스크 분석
    # ----------------------------------------------------------------
    def add_risk_sheet(self, risk: dict, portfolio_perf: dict, suggestions: list):
        ws = self.wb.create_sheet("리스크 분석")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = "포트폴리오 리스크 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # VaR / CVaR 테이블
        row = 4
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "Value at Risk / Expected Shortfall", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        var_headers = ['지표', '신뢰수준', '일일 손실률', '투자금 2천만원 기준']
        for col, h in enumerate(var_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        invest_amount = 20000000  # 2천만원
        var_rows = [
            ('VaR', '95%', risk['var_95'], abs(risk['var_95']) * invest_amount),
            ('VaR', '99%', risk['var_99'], abs(risk['var_99']) * invest_amount),
            ('CVaR', '95%', risk['cvar_95'], abs(risk['cvar_95']) * invest_amount),
            ('CVaR', '99%', risk['cvar_99'], abs(risk['cvar_99']) * invest_amount),
        ]

        for label, conf, val, loss in var_rows:
            bg = 'negative' if abs(val) > 0.03 else ('light_gray' if row % 2 == 0 else 'white')
            self._style_data(ws, row, 2, label, bg_color=bg, bold=True)
            self._style_data(ws, row, 3, conf, bg_color=bg)
            self._style_data(ws, row, 4, f"{val*100:.2f}%", bg_color=bg)
            self._style_data(ws, row, 5, f"약 -{loss:,.0f}원", bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 최대 낙폭
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "최대 낙폭(MDD) 분석", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        dd_series = portfolio_perf.get('drawdown_series')
        if dd_series is not None:
            dd_items = [
                ('최대 낙폭', f"{dd_series.min()*100:.1f}%"),
                ('현재 낙폭', f"{dd_series.iloc[-1]*100:.1f}%"),
                ('평균 낙폭', f"{dd_series.mean()*100:.1f}%"),
            ]
            for key, val in dd_items:
                bg = 'light_gray' if row % 2 == 0 else 'white'
                cell_k = ws.cell(row=row, column=2, value=key)
                cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
                cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
                cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                ws.merge_cells(f'C{row}:D{row}')
                self._style_data(ws, row, 3, val, bg_color=bg)
                ws.row_dimensions[row].height = 22
                row += 1

        # 낙폭 시계열 (최근 60일 샘플)
        if dd_series is not None and len(dd_series) > 0:
            row += 1
            ws.merge_cells(f'B{row}:H{row}')
            self._style_header(ws, row, 2, "낙폭 시계열 (최근 60거래일)", 'header_blue', size=10)
            ws.row_dimensions[row].height = 24
            row += 1

            dd_headers = ['날짜', '낙폭(%)']
            for col, h in enumerate(dd_headers, 2):
                self._style_header(ws, row, col, h, 'header_blue', size=10)
            ws.row_dimensions[row].height = 22
            row += 1

            recent_dd = dd_series.tail(60)
            for date, dd_val in recent_dd.items():
                bg = 'negative' if dd_val < -0.05 else ('light_gray' if row % 2 == 0 else 'white')
                self._style_data(ws, row, 2, date.strftime('%Y-%m-%d'), bg_color=bg)
                self._style_data(ws, row, 3, f"{dd_val*100:.2f}%", bg_color=bg)
                ws.row_dimensions[row].height = 18
                row += 1

        # Beta 분석
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "베타 및 스트레스 분석", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        beta = risk.get('beta')
        benchmark = risk.get('benchmark', 'N/A')
        stress = risk.get('stress_loss_10pct')

        beta_items = [
            ('벤치마크', f"{benchmark}"),
            ('포트폴리오 Beta', f"{beta:.3f}" if beta is not None else 'N/A'),
        ]

        if beta is not None:
            if beta > 1.2:
                interp = "시장 대비 공격적 (시장 변동 대비 높은 민감도)"
            elif beta > 0.8:
                interp = "시장과 유사한 민감도"
            elif beta > 0.5:
                interp = "시장 대비 방어적 (시장 변동 대비 낮은 민감도)"
            else:
                interp = "시장과 낮은 상관 (대안 투자 성격)"
            beta_items.append(('해석', interp))

        if stress is not None:
            beta_items.append(('시장 10% 하락시', f"예상 포트폴리오 손실: {stress*100:.1f}%"))

        for key, val in beta_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            cell_k = ws.cell(row=row, column=2, value=key)
            cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
            cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(f'C{row}:F{row}')
            data_cell = ws.cell(row=row, column=3, value=val)
            data_cell.font = Font(size=10, name='맑은 고딕')
            data_cell.fill = PatternFill("solid", fgColor=self.COLORS[bg])
            data_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 22
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 22
        for c in 'FGH':
            ws.column_dimensions[c].width = 15

    # ----------------------------------------------------------------
    # 시트 6: 벤치마크 비교
    # ----------------------------------------------------------------
    def add_benchmark_sheet(self, benchmarks: dict, portfolio_perf: dict):
        """Buy & Hold 및 SMA200 벤치마크 비교 시트 추가"""
        ws = self.wb.create_sheet("벤치마크 비교")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "벤치마크 비교 분석 (Buy & Hold vs SMA200)"
        title.font = Font(bold=True, size=13, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 설명
        ws['B3'] = "포트폴리오 성과를 Buy & Hold 전략 및 SMA200 기반 시장 타이밍 전략과 비교합니다."
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        row = 5
        # 헤더
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "전략별 성과 비교", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        comp_headers = ['지표', '현재 포트폴리오', 'Buy & Hold', 'SMA200 전략']
        for col, h in enumerate(comp_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        bh = benchmarks.get('buy_hold', {})
        sma = benchmarks.get('sma200', {})
        port_annual = portfolio_perf.get('ann_return', 0) * 100
        port_vol = portfolio_perf.get('ann_volatility', 0) * 100
        port_sharpe = portfolio_perf.get('sharpe', 0)
        port_mdd = portfolio_perf.get('max_drawdown', 0) * 100

        def _color_compare(val, ref):
            """val > ref → positive (green), else negative (red)"""
            if val is None or ref is None:
                return 'white'
            return 'positive' if float(val) >= float(ref) else 'negative'

        metrics = [
            ('연간 수익률(%)', f"{port_annual:.1f}%",
             f"{bh.get('annual_return', 'N/A')}%" if bh else 'N/A',
             f"{sma.get('annual_return', 'N/A')}%" if sma else 'N/A',
             bh.get('annual_return'), sma.get('annual_return'), port_annual),
            ('연간 변동성(%)', f"{port_vol:.1f}%",
             f"{bh.get('volatility', 'N/A')}%" if bh else 'N/A',
             f"{sma.get('volatility', 'N/A')}%" if sma else 'N/A',
             None, None, None),
            ('Sharpe Ratio', f"{port_sharpe:.2f}",
             f"{bh.get('sharpe', 'N/A')}" if bh else 'N/A',
             f"{sma.get('sharpe', 'N/A')}" if sma else 'N/A',
             bh.get('sharpe'), sma.get('sharpe'), port_sharpe),
            ('최대 낙폭(MDD%)', f"{port_mdd:.1f}%",
             f"{bh.get('max_drawdown', 'N/A')}%" if bh else 'N/A',
             f"{sma.get('max_drawdown', 'N/A')}%" if sma else 'N/A',
             None, None, None),
        ]

        for label, v_port, v_bh, v_sma, bh_val, sma_val, port_val in metrics:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            cell_k = ws.cell(row=row, column=2, value=label)
            cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
            cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            self._style_data(ws, row, 3, v_port, bg_color=bg)

            # Color-code benchmark cells: green if portfolio beats them
            bh_bg = _color_compare(port_val, bh_val) if (bh_val is not None and port_val is not None) else bg
            sma_bg = _color_compare(port_val, sma_val) if (sma_val is not None and port_val is not None) else bg

            self._style_data(ws, row, 4, v_bh, bg_color=bh_bg)
            self._style_data(ws, row, 5, v_sma, bg_color=sma_bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 설명 섹션
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "전략 설명", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        descriptions = [
            ('Buy & Hold', '기간 내 포트폴리오를 매수 후 리밸런싱 없이 보유. 시장 평균 수익률의 기준선.'),
            ('SMA200 전략', '200일 이동평균선 위에 있을 때만 투자, 그 외에는 현금 보유. 시장 하락을 일부 회피.'),
            ('현재 포트폴리오', '몬테카를로 최적화 가중치 적용. 초과 성과 여부를 위 두 전략과 비교.'),
        ]

        for name, desc in descriptions:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            cell_k = ws.cell(row=row, column=2, value=name)
            cell_k.font = Font(bold=True, size=10, name='맑은 고딕')
            cell_k.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            cell_k.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(f'C{row}:G{row}')
            desc_cell = ws.cell(row=row, column=3, value=desc)
            desc_cell.font = Font(size=10, name='맑은 고딕')
            desc_cell.fill = PatternFill("solid", fgColor=self.COLORS[bg])
            desc_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

    def save(self, filepath: str):
        self.wb.save(filepath)
        print(f"\n  [저장 완료] {filepath}")


# ====================================================================
# 4. 콘솔 리포트
# ====================================================================
def print_console_report(tickers: list, weights: np.ndarray, ticker_names: dict,
                         portfolio_perf: dict, asset_results: list,
                         suggestions: list, risk: dict,
                         start_date: str, end_date: str):
    """콘솔 결과 출력"""

    # 자산 구성 문자열
    asset_str = ', '.join([
        f"{ticker_names.get(t, t)}({w*100:.0f}%)"
        for t, w in zip(tickers, weights)
    ])

    print("\n" + "=" * 60)
    print("  복합 자산 포트폴리오 분석")
    print("=" * 60)

    print(f"\n  자산 구성: {asset_str}")
    print(f"  분석 기간: {start_date} ~ {end_date}")

    # 포트폴리오 성과
    perf = portfolio_perf
    print(f"\n  --- 포트폴리오 성과 ---")
    print(f"  연간 수익률:  {perf['ann_return']*100:.1f}%")
    print(f"  연간 변동성:  {perf['ann_volatility']*100:.1f}%")
    print(f"  Sharpe Ratio: {perf['sharpe']:.2f}")
    print(f"  Sortino Ratio: {perf.get('sortino', 0):.2f}")
    print(f"  Max Drawdown: {perf['max_drawdown']*100:.1f}%")

    # 개별 자산
    print(f"\n  --- 개별 자산 ---")
    for asset in asset_results:
        name_padded = asset['name']
        if len(name_padded) < 10:
            name_padded = name_padded + ' ' * (10 - len(name_padded))
        print(f"  {name_padded} 수익률: {asset['ann_return']*100:.1f}%"
              f"  변동성: {asset['ann_volatility']*100:.1f}%"
              f"  Sharpe: {asset['sharpe']:.2f}")

    # 최적 자산배분 제안
    if suggestions:
        print(f"\n  --- 최적 자산배분 제안 ---")
        for sug in suggestions:
            weight_str = ' / '.join([
                f"{ticker_names.get(t, t)} {sug['weights'][i]*100:.0f}%"
                for i, t in enumerate(tickers)
            ])
            print(f"  {sug['name']}: {weight_str} (Sharpe {sug['sharpe']:.2f})")

    # 리스크
    print(f"\n  --- 리스크 ---")
    invest_amount = 20000000
    var_95_loss = abs(risk['var_95']) * invest_amount
    print(f"  VaR(95%):  일일 {risk['var_95']*100:.1f}% (약 -{var_95_loss:,.0f}원/2천만원 투자시)")
    print(f"  VaR(99%):  일일 {risk['var_99']*100:.1f}%")

    if risk.get('stress_loss_10pct') is not None:
        print(f"  시장 10% 하락시 예상 손실: {risk['stress_loss_10pct']*100:.1f}%")

    print("\n" + "=" * 60)


# ====================================================================
# 5. 메인 실행
# ====================================================================
def main():
    parser = argparse.ArgumentParser(description="복합 자산 포트폴리오 분석 시스템")
    parser.add_argument('--tickers', required=True,
                        help='종목 코드 (쉼표 구분, 예: 005930,AAPL,BTC-USD)')
    parser.add_argument('--weights', default=None,
                        help='자산 비중 (쉼표 구분, 예: 0.5,0.3,0.2). 생략시 동일비중')
    parser.add_argument('--period', default='2y',
                        help='분석 기간 (기본: 2y)')
    args = parser.parse_args()

    # 티커 파싱
    tickers = [t.strip() for t in args.tickers.split(',') if t.strip()]
    if len(tickers) == 0:
        print("[오류] 최소 1개 이상의 티커를 입력하세요.")
        sys.exit(1)

    # 가중치 파싱
    weights = None
    if args.weights:
        try:
            weights = [float(w.strip()) for w in args.weights.split(',')]
            if len(weights) != len(tickers):
                print(f"[경고] 가중치({len(weights)}개)와 티커({len(tickers)}개) 수 불일치. 동일비중으로 진행합니다.")
                weights = None
        except ValueError:
            print("[경고] 가중치 파싱 실패. 동일비중으로 진행합니다.")
            weights = None

    print("\n" + "=" * 60)
    print("  복합 자산 포트폴리오 분석 시스템")
    print("=" * 60)

    # 1. 데이터 수집
    collector = PortfolioDataCollector(tickers, period=args.period)
    if not collector.fetch_all():
        print("[오류] 데이터 수집 실패. 프로그램을 종료합니다.")
        sys.exit(1)

    valid_tickers = collector.tickers
    if len(valid_tickers) == 0:
        print("[오류] 유효한 자산이 없습니다.")
        sys.exit(1)

    # 가중치 재조정 (실패한 티커 제외)
    if weights:
        # 유효 티커에 맞게 재매핑
        raw_to_yf = {}
        for raw_t in tickers:
            for yf_t in valid_tickers:
                if raw_t.strip().upper() in yf_t.upper() or raw_t.strip() in yf_t:
                    raw_to_yf[raw_t] = yf_t
                    break
        # 가중치 필터링
        new_weights = []
        for i, raw_t in enumerate(tickers):
            if raw_t in raw_to_yf and i < len(weights):
                new_weights.append(weights[i])
        if len(new_weights) == len(valid_tickers):
            weights = new_weights
        else:
            weights = None

    # 2. 분석
    analyzer = PortfolioAnalyzer(
        collector.aligned_prices,
        collector.daily_returns,
        weights=weights,
        ticker_names=collector.ticker_names,
    )

    print(f"\n[분석] 포트폴리오 분석 시작...")

    # a) 개별 자산 분석
    print(f"  |- 개별 자산 분석 중...")
    asset_results = analyzer.analyze_individual_assets()

    # b) 상관관계
    print(f"  |- 상관관계 분석 중...")
    corr_data = analyzer.analyze_correlations()

    # c) 포트폴리오 성과
    print(f"  |- 포트폴리오 성과 계산 중...")
    portfolio_perf = analyzer.analyze_portfolio_performance()
    comparison = analyzer.compare_portfolios()

    # d) 최적화 (2개 이상 자산인 경우만)
    mc_result = None
    suggestions = []
    if len(valid_tickers) >= 2:
        mc_result = analyzer.optimize_monte_carlo()
        suggestions = analyzer.suggest_allocations(mc_result)
    else:
        print(f"  |- [참고] 자산 1개 - 최적화 건너뜁니다.")

    # e) 리스크 분석
    print(f"  |- 리스크 분석 중...")
    majority_market = collector.get_majority_market()
    benchmark = '^KS11' if majority_market == 'KR' else '^GSPC'
    risk = analyzer.analyze_risk(benchmark_ticker=benchmark)

    # f) 벤치마크 비교
    print(f"  |- 벤치마크 비교 (Buy & Hold / SMA200) 계산 중...")
    benchmarks = analyzer.calculate_benchmarks()

    # 기간 정보
    start_date = collector.aligned_prices.index[0].strftime('%Y-%m-%d')
    end_date = collector.aligned_prices.index[-1].strftime('%Y-%m-%d')

    # 3. 콘솔 출력
    print_console_report(
        valid_tickers, analyzer.weights, collector.ticker_names,
        portfolio_perf, asset_results, suggestions, risk,
        start_date, end_date,
    )

    # 4. Excel 저장
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"output/포트폴리오분석_{date_str}.xlsx"

    print(f"\n  Excel 보고서 생성 중...")
    builder = PortfolioExcelBuilder(collector.ticker_names)
    builder.add_summary_sheet(
        portfolio_perf, comparison, risk, suggestions,
        valid_tickers, analyzer.weights, start_date, end_date,
    )
    builder.add_individual_sheet(asset_results)

    if len(valid_tickers) >= 2:
        builder.add_correlation_sheet(corr_data, valid_tickers)

    if mc_result and suggestions:
        builder.add_optimization_sheet(suggestions, valid_tickers, analyzer.weights, mc_result)

    builder.add_risk_sheet(risk, portfolio_perf, suggestions)

    if benchmarks:
        builder.add_benchmark_sheet(benchmarks, portfolio_perf)

    builder.save(filename)

    print(f"\n  다음 단계: output 폴더의 Excel 파일을 확인하세요.")
    print(f"  파일 경로: {filename}\n")


if __name__ == "__main__":
    main()
