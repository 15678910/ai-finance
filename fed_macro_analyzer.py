"""
연준/매크로 경제 분석 시스템 - FRED 데이터 기반
================================================
사용법:
  1. pip install requests pandas openpyxl numpy
  2. python fed_macro_analyzer.py                    (전체 분석)
  3. python fed_macro_analyzer.py --period 2y        (기간 지정)
  4. python fed_macro_analyzer.py --ticker 005930    (종목 상관관계 포함)

생성 파일:
  - output/매크로분석_[날짜].xlsx
"""

import os
import sys
import argparse
import warnings
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')

# -- 필수 라이브러리 설치 확인 ----------------------------------------
def check_and_install():
    required = {
        'requests': 'requests',
        'pandas': 'pandas',
        'openpyxl': 'openpyxl',
        'numpy': 'numpy',
    }
    missing = []
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[설치 필요] pip install {' '.join(missing)}")
        sys.exit(1)

check_and_install()

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- 출력 폴더 생성 ---------------------------------------------------
Path("output").mkdir(exist_ok=True)


# ====================================================================
# 1. FRED 매크로 데이터 수집기
# ====================================================================
class MacroDataCollector:
    """FRED CSV URL을 통한 매크로 경제 데이터 수집 (API 키 불필요)"""

    # FRED 시리즈 정의: (시리즈ID, 한글명, 단위, 설명)
    SERIES_MAP = {
        'FEDFUNDS':  ('기준금리(FFR)',      '%',     '연방기금금리'),
        'CPIAUCSL':  ('소비자물가지수(CPI)', 'Index', '도시 소비자 물가지수'),
        'UNRATE':    ('실업률',              '%',     '미국 실업률'),
        'GDP':       ('실질GDP',             '십억$', '미국 실질 GDP (분기)'),
        'T10Y2Y':    ('장단기금리차(10Y-2Y)','%',     '10년-2년 국채 스프레드'),
        'T10YIE':    ('기대인플레이션(10Y)', '%',     '10년 손익분기 인플레이션율'),
        'VIXCLS':    ('VIX지수',             'Index', '시장 변동성 지수'),
        'DGS10':     ('10년국채금리',        '%',     '10년 만기 국채 수익률'),
        'DGS2':      ('2년국채금리',         '%',     '2년 만기 국채 수익률'),
        'DTWEXBGS':  ('달러인덱스(TWI)',     'Index', '무역가중 달러 지수'),
        'DFEDTARU':  ('FF목표금리(상단)',    '%',     '연방기금금리 목표 상단'),
        'DFEDTARL':  ('FF목표금리(하단)',    '%',     '연방기금금리 목표 하단'),
    }

    FRED_CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv"

    def __init__(self, period: str = '2y'):
        self.period = period
        self.start_date = self._parse_period(period)
        self.data = {}       # series_id -> DataFrame
        self.trends = {}     # series_id -> trend dict
        self.failed = []     # 수집 실패 시리즈

    @staticmethod
    def _parse_period(period: str) -> str:
        """기간 문자열을 시작 날짜로 변환"""
        now = datetime.now()
        period = period.lower().strip()
        if period.endswith('y'):
            years = int(period.replace('y', ''))
            start = now - timedelta(days=365 * years)
        elif period.endswith('m'):
            months = int(period.replace('m', ''))
            start = now - timedelta(days=30 * months)
        else:
            # 기본 2년
            start = now - timedelta(days=730)
        return start.strftime('%Y-%m-%d')

    def fetch_series(self, series_id: str) -> pd.DataFrame:
        """단일 FRED 시리즈 CSV 다운로드"""
        try:
            params = {
                'id': series_id,
                'cosd': self.start_date,
                'coed': datetime.now().strftime('%Y-%m-%d'),
            }
            resp = requests.get(self.FRED_CSV_URL, params=params, timeout=10)
            resp.raise_for_status()

            # CSV 파싱
            from io import StringIO
            df = pd.read_csv(StringIO(resp.text))

            # 컬럼명 정리
            df.columns = ['DATE', 'VALUE']
            df['DATE'] = pd.to_datetime(df['DATE'], errors='coerce')
            df['VALUE'] = pd.to_numeric(df['VALUE'], errors='coerce')
            df = df.dropna(subset=['DATE'])
            df = df.set_index('DATE').sort_index()

            # '.' 또는 결측치 제거
            df = df.dropna()

            return df

        except Exception as e:
            print(f"  [!] {series_id} 수집 실패: {e}")
            self.failed.append(series_id)
            return pd.DataFrame()

    def fetch_all(self):
        """모든 FRED 시리즈 수집"""
        print("\n[수집] FRED 매크로 데이터 수집 시작...")
        total = len(self.SERIES_MAP)

        for idx, (series_id, (name, unit, desc)) in enumerate(self.SERIES_MAP.items(), 1):
            print(f"  ({idx}/{total}) {name} [{series_id}] 수집 중...")
            df = self.fetch_series(series_id)
            if not df.empty:
                self.data[series_id] = df
                print(f"         -> {len(df)}개 데이터포인트 수집 완료")
            else:
                print(f"         -> 수집 실패 (건너뜀)")

        success = len(self.data)
        fail = len(self.failed)
        print(f"\n  [결과] 성공: {success}개, 실패: {fail}개")

        if success == 0:
            print("\n  [오류] 모든 지표 수집에 실패했습니다. 네트워크 연결을 확인하세요.")
            sys.exit(1)

        return self.data

    def calculate_trends(self) -> dict:
        """각 시리즈별 추세 계산: 현재값, 3개월전, 6개월전, 1년전, 방향"""
        print("\n[분석] 추세 계산 중...")

        for series_id, df in self.data.items():
            name, unit, desc = self.SERIES_MAP[series_id]

            current_val = df['VALUE'].iloc[-1] if len(df) > 0 else None
            current_date = df.index[-1] if len(df) > 0 else None

            # N개월 전 값 찾기 (가장 가까운 날짜)
            val_3m = self._get_value_months_ago(df, 3)
            val_6m = self._get_value_months_ago(df, 6)
            val_1y = self._get_value_months_ago(df, 12)

            # 변화량 계산
            chg_3m = (current_val - val_3m) if (current_val is not None and val_3m is not None) else None
            chg_6m = (current_val - val_6m) if (current_val is not None and val_6m is not None) else None
            chg_1y = (current_val - val_1y) if (current_val is not None and val_1y is not None) else None

            # 방향 결정 (3개월 변화 기준)
            if chg_3m is not None:
                if abs(chg_3m) < 0.01 * abs(current_val) if current_val != 0 else abs(chg_3m) < 0.01:
                    direction = "보합"
                    arrow = "-> "
                elif chg_3m > 0:
                    direction = "상승"
                    arrow = "^ "
                else:
                    direction = "하락"
                    arrow = "v "
            else:
                direction = "N/A"
                arrow = "- "

            self.trends[series_id] = {
                'name': name,
                'unit': unit,
                'desc': desc,
                'current': current_val,
                'current_date': current_date,
                'val_3m': val_3m,
                'val_6m': val_6m,
                'val_1y': val_1y,
                'chg_3m': chg_3m,
                'chg_6m': chg_6m,
                'chg_1y': chg_1y,
                'direction': direction,
                'arrow': arrow,
            }

        return self.trends

    @staticmethod
    def _get_value_months_ago(df: pd.DataFrame, months: int):
        """N개월 전 가장 가까운 데이터 값 반환"""
        if df.empty:
            return None
        target_date = df.index[-1] - pd.DateOffset(months=months)
        # 가장 가까운 날짜 찾기
        idx = df.index.searchsorted(target_date)
        if idx >= len(df):
            idx = len(df) - 1
        if idx < 0:
            idx = 0
        val = df['VALUE'].iloc[idx]
        return val if pd.notna(val) else None

    def get_monthly_series(self, series_id: str) -> pd.DataFrame:
        """월간 데이터로 리샘플링 (마지막 값 기준)"""
        if series_id not in self.data:
            return pd.DataFrame()
        df = self.data[series_id].copy()
        # 월말 리샘플링
        monthly = df.resample('ME').last().dropna()
        return monthly


# ====================================================================
# 2. 연준/매크로 분석기
# ====================================================================
class FedAnalyzer:
    """연준 정책 및 매크로 경제 환경 종합 분석"""

    # 경기 사이클 단계 정의
    CYCLE_STAGES = {
        'expansion':  ('확장기', '경제 성장세 지속. 고용 증가, 소비/투자 활발.'),
        'peak':       ('정점기', '성장 둔화 조짐. 인플레이션 상승, 금리 고점 근접.'),
        'contraction':('수축기', '경기 하강 국면. 실업 증가, 소비 위축 시작.'),
        'trough':     ('저점기', '경기 바닥 근접. 정책 완화 기대, 회복 신호 탐색.'),
    }

    # 자산별 분석 매트릭스
    ASSET_CLASSES = [
        '미국 성장주',
        '미국 가치주',
        '장기 채권(10Y+)',
        '단기 채권(2Y)',
        '미국 달러',
        '금/귀금속',
        '원자재/에너지',
        '암호화폐/비트코인',
    ]

    def __init__(self, collector: MacroDataCollector):
        self.collector = collector
        self.trends = collector.trends
        self.data = collector.data
        self.rate_analysis = {}
        self.inflation_analysis = {}
        self.cycle_analysis = {}
        self.asset_outlook = {}

    def analyze_rates(self) -> dict:
        """금리 환경 분석"""
        print("\n[분석] 금리 환경 분석 중...")

        ffr = self.trends.get('FEDFUNDS', {})
        dgs10 = self.trends.get('DGS10', {})
        dgs2 = self.trends.get('DGS2', {})

        current_rate = ffr.get('current')
        rate_3m_ago = ffr.get('val_3m')
        rate_6m_ago = ffr.get('val_6m')
        rate_1y_ago = ffr.get('val_1y')

        # 금리 사이클 단계 판단
        if current_rate is not None and rate_3m_ago is not None and rate_6m_ago is not None:
            recent_chg = current_rate - rate_3m_ago
            medium_chg = current_rate - rate_6m_ago

            if recent_chg > 0.1 and medium_chg > 0.1:
                cycle_stage = '인상기'
                cycle_desc = '연준이 금리를 인상하는 긴축 국면입니다.'
            elif recent_chg < -0.1 and medium_chg < -0.1:
                cycle_stage = '인하기'
                cycle_desc = '연준이 금리를 인하하는 완화 국면입니다.'
            elif abs(recent_chg) <= 0.1:
                cycle_stage = '동결기'
                cycle_desc = '연준이 금리를 동결하며 관망하는 국면입니다.'
            else:
                cycle_stage = '전환기'
                cycle_desc = '금리 방향성이 전환되는 과도기입니다.'
        else:
            cycle_stage = 'N/A'
            cycle_desc = '데이터 부족으로 판단 불가.'

        # 장단기 스프레드
        spread = self.trends.get('T10Y2Y', {})
        spread_val = spread.get('current')
        if spread_val is not None:
            if spread_val < 0:
                yield_curve_status = '역전 (경기침체 경고)'
            elif spread_val < 0.5:
                yield_curve_status = '평탄화 (경기 둔화 신호)'
            else:
                yield_curve_status = '정상 (경기 확장 신호)'
        else:
            yield_curve_status = 'N/A'

        self.rate_analysis = {
            'current_rate': current_rate,
            'rate_3m_ago': rate_3m_ago,
            'rate_6m_ago': rate_6m_ago,
            'rate_1y_ago': rate_1y_ago,
            'cycle_stage': cycle_stage,
            'cycle_desc': cycle_desc,
            'dgs10': dgs10.get('current'),
            'dgs2': dgs2.get('current'),
            'spread': spread_val,
            'yield_curve_status': yield_curve_status,
        }
        return self.rate_analysis

    def analyze_inflation(self) -> dict:
        """인플레이션 분석"""
        print("[분석] 인플레이션 분석 중...")

        cpi = self.trends.get('CPIAUCSL', {})
        breakeven = self.trends.get('T10YIE', {})
        ffr = self.trends.get('FEDFUNDS', {})

        # CPI 전년비 변화율 계산 (YoY)
        cpi_yoy = None
        if 'CPIAUCSL' in self.data:
            cpi_df = self.data['CPIAUCSL']
            if len(cpi_df) > 12:
                current_cpi = cpi_df['VALUE'].iloc[-1]
                year_ago_idx = cpi_df.index.searchsorted(cpi_df.index[-1] - pd.DateOffset(months=12))
                if 0 <= year_ago_idx < len(cpi_df):
                    year_ago_cpi = cpi_df['VALUE'].iloc[year_ago_idx]
                    if year_ago_cpi > 0:
                        cpi_yoy = ((current_cpi - year_ago_cpi) / year_ago_cpi) * 100

        # 실질금리 = 기준금리 - 기대인플레이션
        real_rate = None
        if ffr.get('current') is not None and breakeven.get('current') is not None:
            real_rate = ffr['current'] - breakeven['current']

        # 인플레이션 상태 판단
        if cpi_yoy is not None:
            if cpi_yoy > 5:
                inflation_status = '고인플레이션 (5% 초과)'
                inflation_risk = '매우 높음'
            elif cpi_yoy > 3:
                inflation_status = '인플레이션 상승 (3-5%)'
                inflation_risk = '높음'
            elif cpi_yoy > 2:
                inflation_status = '적정 수준 (2-3%)'
                inflation_risk = '보통'
            elif cpi_yoy > 0:
                inflation_status = '저인플레이션 (0-2%)'
                inflation_risk = '낮음'
            else:
                inflation_status = '디플레이션 위험'
                inflation_risk = '특수 (디플레이션)'
        else:
            inflation_status = 'N/A'
            inflation_risk = 'N/A'

        self.inflation_analysis = {
            'cpi_current': cpi.get('current'),
            'cpi_yoy': cpi_yoy,
            'breakeven_10y': breakeven.get('current'),
            'real_rate': real_rate,
            'inflation_status': inflation_status,
            'inflation_risk': inflation_risk,
        }
        return self.inflation_analysis

    def analyze_cycle(self) -> dict:
        """경기 사이클 진단"""
        print("[분석] 경기 사이클 진단 중...")

        unrate = self.trends.get('UNRATE', {})
        gdp = self.trends.get('GDP', {})
        spread = self.trends.get('T10Y2Y', {})
        vix = self.trends.get('VIXCLS', {})

        # 신호 수집
        signals = []
        score = 0  # 양수 = 확장, 음수 = 수축

        # 1) 장단기 금리차
        spread_val = spread.get('current')
        if spread_val is not None:
            if spread_val < 0:
                signals.append(('장단기금리차 역전', '수축 신호', 'negative'))
                score -= 2
            elif spread_val < 0.5:
                signals.append(('장단기금리차 축소', '둔화 신호', 'negative'))
                score -= 1
            else:
                signals.append(('장단기금리차 정상', '확장 신호', 'positive'))
                score += 1

        # 2) 실업률
        unrate_val = unrate.get('current')
        unrate_chg = unrate.get('chg_3m')
        if unrate_val is not None and unrate_chg is not None:
            if unrate_chg > 0.3:
                signals.append(('실업률 상승 중', '수축 신호', 'negative'))
                score -= 2
            elif unrate_chg > 0:
                signals.append(('실업률 소폭 상승', '둔화 신호', 'negative'))
                score -= 1
            elif unrate_val < 4.0:
                signals.append(('실업률 낮은 수준', '확장 신호', 'positive'))
                score += 1
            else:
                signals.append(('실업률 안정', '중립', 'neutral'))

        # 3) GDP 성장
        gdp_chg = gdp.get('chg_3m')
        if gdp_chg is not None:
            if gdp_chg > 0:
                signals.append(('GDP 성장 지속', '확장 신호', 'positive'))
                score += 1
            else:
                signals.append(('GDP 감소', '수축 신호', 'negative'))
                score -= 1

        # 4) VIX 변동성
        vix_val = vix.get('current')
        if vix_val is not None:
            if vix_val > 30:
                signals.append(('VIX 30 초과 (공포)', '불안정', 'negative'))
                score -= 1
            elif vix_val > 20:
                signals.append(('VIX 20-30 (경계)', '주의', 'negative'))
            elif vix_val < 15:
                signals.append(('VIX 15 미만 (안정)', '낙관', 'positive'))
                score += 1
            else:
                signals.append(('VIX 정상 범위', '중립', 'neutral'))

        # 종합 판단
        if score >= 2:
            cycle = 'expansion'
        elif score >= 0:
            cycle = 'peak'
        elif score >= -2:
            cycle = 'contraction'
        else:
            cycle = 'trough'

        stage_name, stage_desc = self.CYCLE_STAGES[cycle]

        self.cycle_analysis = {
            'cycle': cycle,
            'stage_name': stage_name,
            'stage_desc': stage_desc,
            'score': score,
            'signals': signals,
            'unemployment': unrate_val,
            'unemployment_chg': unrate_chg,
            'gdp_current': gdp.get('current'),
            'vix': vix_val,
            'spread': spread_val,
        }
        return self.cycle_analysis

    def analyze_asset_outlook(self) -> dict:
        """자산별 전망 분석"""
        print("[분석] 자산별 전망 분석 중...")

        cycle = self.cycle_analysis.get('cycle', 'peak')
        rate_stage = self.rate_analysis.get('cycle_stage', 'N/A')
        inflation = self.inflation_analysis.get('cpi_yoy')
        real_rate = self.inflation_analysis.get('real_rate')
        vix = self.cycle_analysis.get('vix')
        spread = self.cycle_analysis.get('spread')

        outlook = {}

        # 미국 성장주
        if rate_stage == '인하기' and cycle in ('trough', 'expansion'):
            outlook['미국 성장주'] = {
                'impact': '긍정',
                'strategy': '금리 인하기에 성장주 밸류에이션 상승 기대. 기술/AI 섹터 비중 확대.',
                'risk': '경기 침체 심화 시 실적 하향 가능성',
            }
        elif rate_stage == '인상기':
            outlook['미국 성장주'] = {
                'impact': '부정',
                'strategy': '고금리 환경에서 성장주 밸류에이션 압박. 비중 축소 권장.',
                'risk': '금리 추가 인상 시 추가 하락 가능성',
            }
        else:
            outlook['미국 성장주'] = {
                'impact': '중립',
                'strategy': '금리 동결기 선별적 접근. 실적 성장 확실한 기업 중심.',
                'risk': '금리 방향 전환 시 변동성 확대',
            }

        # 미국 가치주
        if cycle == 'expansion':
            outlook['미국 가치주'] = {
                'impact': '긍정',
                'strategy': '경기 확장기 가치주 수혜. 금융/에너지/산업재 비중 확대.',
                'risk': '경기 정점 도달 시 모멘텀 약화',
            }
        elif cycle in ('contraction', 'trough'):
            outlook['미국 가치주'] = {
                'impact': '부정',
                'strategy': '경기 수축기 실적 하향 우려. 방어적 가치주(배당주)만 선별.',
                'risk': '경기 침체 장기화 시 배당 삭감 가능성',
            }
        else:
            outlook['미국 가치주'] = {
                'impact': '중립',
                'strategy': '경기 둔화 국면에서 선별적 접근. 현금흐름 우량 기업 선호.',
                'risk': '섹터별 차별화 심화',
            }

        # 장기 채권(10Y+)
        if rate_stage == '인하기':
            outlook['장기 채권(10Y+)'] = {
                'impact': '긍정',
                'strategy': '금리 인하 사이클에서 장기채 가격 상승 기대. 듀레이션 확대.',
                'risk': '인플레이션 재발 시 금리 반등 가능성',
            }
        elif rate_stage == '인상기':
            outlook['장기 채권(10Y+)'] = {
                'impact': '부정',
                'strategy': '금리 인상기 장기채 손실 확대. 단기채 또는 현금 선호.',
                'risk': '금리 추가 인상 시 추가 손실',
            }
        else:
            outlook['장기 채권(10Y+)'] = {
                'impact': '중립',
                'strategy': '금리 동결기 캐리 수익 확보. 듀레이션 중립 유지.',
                'risk': '금리 방향 전환 시 가격 변동',
            }

        # 단기 채권(2Y)
        ffr_current = self.rate_analysis.get('current_rate')
        if ffr_current is not None and ffr_current > 4:
            outlook['단기 채권(2Y)'] = {
                'impact': '긍정',
                'strategy': '고금리 환경에서 단기채 매력적. 안정적 이자수익 확보.',
                'risk': '금리 인하 시 재투자 수익률 하락',
            }
        elif rate_stage == '인하기':
            outlook['단기 채권(2Y)'] = {
                'impact': '중립',
                'strategy': '금리 인하 예상 시 단기채 수익률 하락. 점진적 장기채 전환 검토.',
                'risk': '금리 인하 속도에 따른 수익률 변동',
            }
        else:
            outlook['단기 채권(2Y)'] = {
                'impact': '중립',
                'strategy': '안전자산으로서 포트폴리오 안정화 역할.',
                'risk': '인플레이션 대비 실질 수익률 저조 가능성',
            }

        # 미국 달러
        if rate_stage == '인상기':
            outlook['미국 달러'] = {
                'impact': '긍정',
                'strategy': '금리 인상기 달러 강세 기대. 달러 자산 비중 유지/확대.',
                'risk': '글로벌 경기 회복 시 달러 약세 전환',
            }
        elif rate_stage == '인하기':
            outlook['미국 달러'] = {
                'impact': '부정',
                'strategy': '금리 인하기 달러 약세 가능성. 비달러 자산/이머징 분산.',
                'risk': '안전자산 수요 급증 시 달러 강세 반전',
            }
        else:
            outlook['미국 달러'] = {
                'impact': '중립',
                'strategy': '달러 방향성 제한적. 통화 헤지 유지.',
                'risk': '지정학적 리스크에 따른 급변동',
            }

        # 금/귀금속
        if real_rate is not None and real_rate < 0:
            outlook['금/귀금속'] = {
                'impact': '긍정',
                'strategy': '실질금리 마이너스 환경에서 금 매력 상승. 포트폴리오 5-15% 배분.',
                'risk': '실질금리 반등 시 금 가격 하락',
            }
        elif rate_stage == '인하기':
            outlook['금/귀금속'] = {
                'impact': '긍정',
                'strategy': '금리 인하 + 달러 약세 시 금 상승 기대. 인플레이션 헤지.',
                'risk': '디플레이션 환경에서는 금도 약세 가능',
            }
        elif rate_stage == '인상기':
            outlook['금/귀금속'] = {
                'impact': '부정',
                'strategy': '금리 인상기 금 기회비용 증가. 비중 축소 권장.',
                'risk': '지정학적 리스크 시 급등 가능성',
            }
        else:
            outlook['금/귀금속'] = {
                'impact': '중립',
                'strategy': '포트폴리오 헤지 목적으로 5-10% 유지.',
                'risk': '실질금리 변동에 민감',
            }

        # 원자재/에너지
        if cycle == 'expansion':
            outlook['원자재/에너지'] = {
                'impact': '긍정',
                'strategy': '경기 확장기 원자재 수요 증가. 에너지/광물 관련 투자 확대.',
                'risk': '공급 과잉 또는 경기 둔화 전환 시 급락',
            }
        elif cycle in ('contraction', 'trough'):
            outlook['원자재/에너지'] = {
                'impact': '부정',
                'strategy': '경기 수축기 원자재 수요 감소. 비중 축소.',
                'risk': '지정학적 공급 차질 시 급등 가능성',
            }
        else:
            outlook['원자재/에너지'] = {
                'impact': '중립',
                'strategy': '공급/수요 균형 탐색 구간. 선별적 접근.',
                'risk': 'OPEC+ 정책 및 글로벌 수요 변화',
            }

        # 암호화폐/비트코인
        if rate_stage == '인하기' and cycle in ('trough', 'expansion'):
            outlook['암호화폐/비트코인'] = {
                'impact': '긍정',
                'strategy': '유동성 확대기 위험자산 선호. 소규모 포지션(1-5%) 고려.',
                'risk': '규제 변화, 높은 변동성',
            }
        elif rate_stage == '인상기':
            outlook['암호화폐/비트코인'] = {
                'impact': '부정',
                'strategy': '유동성 축소기 위험자산 회피. 비중 최소화 또는 제로.',
                'risk': '추가 하락 및 유동성 위기',
            }
        elif vix is not None and vix > 25:
            outlook['암호화폐/비트코인'] = {
                'impact': '부정',
                'strategy': '시장 변동성 확대기 암호화폐 리스크 높음. 관망 권장.',
                'risk': '전통 자산 대비 과도한 변동성',
            }
        else:
            outlook['암호화폐/비트코인'] = {
                'impact': '중립',
                'strategy': '투기적 소규모 배분(1-3%) 가능. 장기 관점 접근.',
                'risk': '높은 변동성, 규제 불확실성',
            }

        self.asset_outlook = outlook
        return outlook

    # FOMC 회의 일정 (결정일 기준 - 마지막 날)
    FOMC_SCHEDULE_2025 = [
        '2025-01-29',  # Jan 28-29
        '2025-03-19',  # Mar 18-19
        '2025-05-07',  # May 6-7
        '2025-06-18',  # Jun 17-18
        '2025-07-30',  # Jul 29-30
        '2025-09-17',  # Sep 16-17
        '2025-10-29',  # Oct 28-29
        '2025-12-17',  # Dec 16-17
    ]

    FOMC_SCHEDULE_2026 = [
        '2026-01-28',  # Jan 27-28
        '2026-03-18',  # Mar 17-18
        '2026-05-06',  # May 5-6
        '2026-06-17',  # Jun 16-17
        '2026-07-29',  # Jul 28-29
        '2026-09-16',  # Sep 15-16
        '2026-10-28',  # Oct 27-28
        '2026-12-16',  # Dec 15-16
    ]

    def analyze_fomc(self) -> dict:
        """FOMC 회의 일정 및 시장 금리 기대 분석"""
        print("[분석] FOMC 일정 및 금리 기대 분석 중...")

        today = datetime.now().date()

        # 전체 일정 (2025 + 2026) 파싱
        all_dates = []
        for d in self.FOMC_SCHEDULE_2025:
            all_dates.append(datetime.strptime(d, '%Y-%m-%d').date())
        for d in self.FOMC_SCHEDULE_2026:
            all_dates.append(datetime.strptime(d, '%Y-%m-%d').date())
        all_dates.sort()

        # 다음 FOMC 날짜 탐색
        next_meeting = None
        days_until = None
        for d in all_dates:
            if d >= today:
                next_meeting = d
                days_until = (d - today).days
                break

        # 현재 목표금리 상/하단 (DFEDTARU / DFEDTARL)
        taru_trend = self.trends.get('DFEDTARU', {})
        tarl_trend = self.trends.get('DFEDTARL', {})
        ffr_trend  = self.trends.get('FEDFUNDS', {})

        current_target_upper = taru_trend.get('current')
        current_target_lower = tarl_trend.get('current')
        current_effective    = ffr_trend.get('current')

        # 목표금리 중간값
        target_mid = None
        if current_target_upper is not None and current_target_lower is not None:
            target_mid = (current_target_upper + current_target_lower) / 2

        # 최근 결정 이력: DFEDTARU 월간 값 기준으로 변화 감지
        recent_decisions = []
        if 'DFEDTARU' in self.data:
            taru_df = self.data['DFEDTARU']
            # 월간 리샘플 후 변화가 있는 지점만 추출
            monthly = taru_df.resample('ME').last().dropna()
            prev_val = None
            for date, row in monthly.iterrows():
                val = row['VALUE']
                if prev_val is not None and abs(val - prev_val) >= 0.01:
                    direction = '인상' if val > prev_val else '인하'
                    recent_decisions.append({
                        'date': date.strftime('%Y-%m'),
                        'rate': round(val, 2),
                        'change': round(val - prev_val, 2),
                        'action': direction,
                    })
                prev_val = val
            # 최근 3건만 유지
            recent_decisions = recent_decisions[-3:]

        # 시장 기대 판단 (실효금리 vs 목표 중간값, 추세)
        market_expectation = 'hold'
        if current_effective is not None and target_mid is not None:
            diff = current_effective - target_mid

        # 실효금리 6개월 추세로 기대 방향 판단
        ffr_chg_6m = ffr_trend.get('chg_6m')
        ffr_chg_3m = ffr_trend.get('chg_3m')

        # DFEDTARU 추세가 더 직접적인 신호
        taru_chg_6m = taru_trend.get('chg_6m')
        taru_chg_3m = taru_trend.get('chg_3m')

        if taru_chg_3m is not None:
            if taru_chg_3m < -0.05:
                market_expectation = 'cut'
            elif taru_chg_3m > 0.05:
                market_expectation = 'hike'
            else:
                # 6개월 추세도 확인
                if taru_chg_6m is not None:
                    if taru_chg_6m < -0.1:
                        market_expectation = 'cut'
                    elif taru_chg_6m > 0.1:
                        market_expectation = 'hike'
                    else:
                        market_expectation = 'hold'
                else:
                    market_expectation = 'hold'
        elif ffr_chg_3m is not None:
            # DFEDTARU 없으면 실효금리 추세로 대체
            if ffr_chg_3m < -0.1:
                market_expectation = 'cut'
            elif ffr_chg_3m > 0.1:
                market_expectation = 'hike'
            else:
                market_expectation = 'hold'

        # 2026 일정 문자열 목록
        fomc_2026_list = list(self.FOMC_SCHEDULE_2026)

        self.fomc_analysis = {
            'next_meeting': next_meeting.strftime('%Y-%m-%d') if next_meeting else 'N/A',
            'days_until': days_until if days_until is not None else 'N/A',
            'current_target_upper': current_target_upper,
            'current_target_lower': current_target_lower,
            'current_effective': current_effective,
            'target_mid': target_mid,
            'recent_decisions': recent_decisions,
            'market_expectation': market_expectation,
            'fomc_schedule_2026': fomc_2026_list,
        }
        return self.fomc_analysis

    def run_full_analysis(self) -> dict:
        """전체 분석 실행"""
        self.analyze_rates()
        self.analyze_inflation()
        self.analyze_cycle()
        self.analyze_asset_outlook()
        self.analyze_fomc()

        return {
            'rates': self.rate_analysis,
            'inflation': self.inflation_analysis,
            'cycle': self.cycle_analysis,
            'assets': self.asset_outlook,
            'fomc': self.fomc_analysis,
        }


# ====================================================================
# 3. 종목 상관관계 분석 (옵션)
# ====================================================================
class StockMacroCorrelation:
    """특정 종목과 매크로 지표 간 상관관계 분석"""

    def __init__(self, ticker: str, collector: MacroDataCollector):
        self.ticker_raw = ticker
        self.collector = collector
        self.correlation_df = None
        self.stock_data = None

    def fetch_stock_data(self) -> pd.DataFrame:
        """yfinance를 사용하지 않고 간단히 주가 데이터 수집 시도"""
        # yfinance가 설치되어 있으면 사용
        try:
            import yfinance as yf

            if self.ticker_raw.isdigit():
                ticker = self._detect_market(self.ticker_raw)
            else:
                ticker = self.ticker_raw

            stock = yf.Ticker(ticker)
            df = stock.history(period=self.collector.period)
            if df.empty:
                return pd.DataFrame()
            df.index = df.index.tz_localize(None)
            # 월간 수익률 계산
            monthly = df['Close'].resample('ME').last()
            monthly_ret = monthly.pct_change().dropna()
            self.stock_data = monthly_ret
            self.company_name = stock.info.get('longName') or stock.info.get('shortName', self.ticker_raw)
            return monthly_ret
        except ImportError:
            print("  [!] yfinance 미설치. 종목 상관관계 분석을 건너뜁니다.")
            print("      pip install yfinance 로 설치 후 다시 시도하세요.")
            return pd.DataFrame()
        except Exception as e:
            print(f"  [!] 주가 데이터 수집 실패: {e}")
            return pd.DataFrame()

    @staticmethod
    def _detect_market(ticker_code):
        """코스피(.KS) / 코스닥(.KQ) 자동 감지"""
        import yfinance as yf
        for suffix in [".KS", ".KQ"]:
            try:
                t = yf.Ticker(ticker_code + suffix)
                hist = t.history(period="5d")
                if not hist.empty and len(hist) > 0:
                    info = t.info or {}
                    name = info.get('longName') or info.get('shortName', '')
                    if name and not name.startswith(ticker_code):
                        return ticker_code + suffix
            except Exception:
                pass
        print(f"  [!] 시장 자동 감지 실패, 코스피(.KS)로 시도합니다.")
        return ticker_code + ".KS"

    def calculate_correlations(self) -> pd.DataFrame:
        """매크로 지표와 주가 수익률 간 상관관계 계산"""
        if self.stock_data is None or self.stock_data.empty:
            return pd.DataFrame()

        results = []
        for series_id, df in self.collector.data.items():
            name = self.collector.SERIES_MAP[series_id][0]
            # 월간 변화율
            monthly = df.resample('ME').last().dropna()
            monthly_chg = monthly['VALUE'].pct_change().dropna()

            # 공통 날짜 기준 정렬
            common_idx = self.stock_data.index.intersection(monthly_chg.index)
            if len(common_idx) < 6:
                results.append({
                    'series_id': series_id,
                    'name': name,
                    'correlation': None,
                    'data_points': len(common_idx),
                    'interpretation': '데이터 부족',
                })
                continue

            stock_aligned = self.stock_data.loc[common_idx]
            macro_aligned = monthly_chg.loc[common_idx]

            corr = stock_aligned.corr(macro_aligned)

            if abs(corr) > 0.6:
                strength = '강한'
            elif abs(corr) > 0.3:
                strength = '보통'
            else:
                strength = '약한'

            direction = '양(+)의' if corr > 0 else '음(-)의'
            interpretation = f"{strength} {direction} 상관관계"

            results.append({
                'series_id': series_id,
                'name': name,
                'correlation': round(corr, 4),
                'data_points': len(common_idx),
                'interpretation': interpretation,
            })

        self.correlation_df = pd.DataFrame(results)
        return self.correlation_df


# ====================================================================
# 4. Excel 보고서 생성
# ====================================================================
class MacroExcelBuilder:
    """매크로 분석 Excel 보고서 빌더"""

    # 색상 팔레트 (financial_analyzer.py 동일)
    COLORS = {
        'header_dark': '0D1B2A',
        'header_blue': '1B4F72',
        'accent': '2E86C1',
        'light_blue': 'D6EAF8',
        'white': 'FFFFFF',
        'light_gray': 'F2F3F4',
        'positive': 'EAFAF1',
        'negative': 'FDEDEC',
        'gold': 'F39C12',
    }

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 기본 시트 제거

    def _style_header(self, ws, row, col, value, color='header_dark', font_color='FFFFFF', bold=True, size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=self.COLORS.get(color, color))
        cell.font = Font(bold=bold, color=font_color, size=size, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def _style_data(self, ws, row, col, value, bg_color=None, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if number_format:
            cell.number_format = number_format
        return cell

    def _style_text(self, ws, row, col, value, bg_color=None, bold=False, indent=1):
        """좌측 정렬 텍스트 셀"""
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=indent)
        return cell

    def add_dashboard_sheet(self, trends: dict, series_map: dict):
        """시트 1: 매크로 대시보드"""
        ws = self.wb.create_sheet("매크로 대시보드")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = "연준/매크로 경제 대시보드"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 날짜
        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 헤더
        headers = ['지표명', '현재값', '3개월전', '6개월전', '1년전', '3M 변화', '추세']
        ws.merge_cells('B5:H5')
        self._style_header(ws, 5, 2, "주요 매크로 경제 지표 현황", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        row = 6
        for col, h in enumerate(headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        # 데이터
        for series_id in series_map.keys():
            if series_id not in trends:
                continue
            t = trends[series_id]
            bg = 'light_gray' if row % 2 == 0 else 'white'

            # 지표명
            self._style_text(ws, row, 2, t['name'], bg_color=bg, bold=True)

            # 현재값
            fmt = '#,##0.00' if t['unit'] == '%' else '#,##0.0'
            self._style_data(ws, row, 3, t['current'], bg_color=bg, number_format=fmt)

            # 3개월전
            self._style_data(ws, row, 4, t['val_3m'], bg_color=bg, number_format=fmt)

            # 6개월전
            self._style_data(ws, row, 5, t['val_6m'], bg_color=bg, number_format=fmt)

            # 1년전
            self._style_data(ws, row, 6, t['val_1y'], bg_color=bg, number_format=fmt)

            # 3M 변화
            chg = t['chg_3m']
            if chg is not None:
                chg_bg = 'positive' if chg > 0 else ('negative' if chg < 0 else bg)
                chg_str = f"{chg:+.2f}"
                self._style_data(ws, row, 7, chg_str, bg_color=chg_bg)
            else:
                self._style_data(ws, row, 7, 'N/A', bg_color=bg)

            # 추세
            arrow = t['arrow'] + t['direction']
            direction_bg = bg
            if t['direction'] == '상승':
                direction_bg = 'positive'
            elif t['direction'] == '하락':
                direction_bg = 'negative'
            self._style_data(ws, row, 8, arrow, bg_color=direction_bg, bold=True)

            ws.row_dimensions[row].height = 24
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 24
        for c_letter in 'CDEFGH':
            ws.column_dimensions[c_letter].width = 16

    def add_rates_inflation_sheet(self, rate_analysis: dict, inflation_analysis: dict,
                                   collector: MacroDataCollector):
        """시트 2: 금리-인플레이션"""
        ws = self.wb.create_sheet("금리-인플레이션")
        ws.sheet_view.showGridLines = False

        # --- 금리 환경 분석 ---
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "금리 환경 및 인플레이션 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 금리 분석 요약
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "금리 환경 분석", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        rate_items = [
            ('현재 기준금리 (FFR)', self._fmt_val(rate_analysis.get('current_rate'), '%')),
            ('3개월전 기준금리', self._fmt_val(rate_analysis.get('rate_3m_ago'), '%')),
            ('6개월전 기준금리', self._fmt_val(rate_analysis.get('rate_6m_ago'), '%')),
            ('1년전 기준금리', self._fmt_val(rate_analysis.get('rate_1y_ago'), '%')),
            ('금리 사이클 단계', rate_analysis.get('cycle_stage', 'N/A')),
            ('사이클 설명', rate_analysis.get('cycle_desc', '')),
            ('10년 국채금리', self._fmt_val(rate_analysis.get('dgs10'), '%')),
            ('2년 국채금리', self._fmt_val(rate_analysis.get('dgs2'), '%')),
            ('장단기 금리차 (10Y-2Y)', self._fmt_val(rate_analysis.get('spread'), '%')),
            ('수익률 곡선 상태', rate_analysis.get('yield_curve_status', 'N/A')),
        ]

        row = 6
        for key, value in rate_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, key, bg_color='light_blue', bold=True)
            ws.merge_cells(f'C{row}:G{row}')
            self._style_text(ws, row, 3, str(value), bg_color=bg)
            ws.row_dimensions[row].height = 24
            row += 1

        # --- 인플레이션 분석 ---
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "인플레이션 분석", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        inflation_items = [
            ('현재 CPI 지수', self._fmt_val(inflation_analysis.get('cpi_current'), '')),
            ('CPI 전년비 변화율 (YoY)', self._fmt_val(inflation_analysis.get('cpi_yoy'), '%')),
            ('10년 기대인플레이션', self._fmt_val(inflation_analysis.get('breakeven_10y'), '%')),
            ('실질금리 (FFR - 기대인플레)', self._fmt_val(inflation_analysis.get('real_rate'), '%')),
            ('인플레이션 상태', inflation_analysis.get('inflation_status', 'N/A')),
            ('인플레이션 리스크', inflation_analysis.get('inflation_risk', 'N/A')),
        ]

        for key, value in inflation_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, key, bg_color='light_blue', bold=True)
            ws.merge_cells(f'C{row}:G{row}')
            self._style_text(ws, row, 3, str(value), bg_color=bg)
            ws.row_dimensions[row].height = 24
            row += 1

        # --- 기준금리 히스토리 (월간) ---
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "기준금리 월간 히스토리", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        hist_headers = ['날짜', '기준금리(%)', 'CPI(YoY 추정)', '실질금리 추정']
        for col, h in enumerate(hist_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        # 기준금리 월간 데이터
        ffr_monthly = collector.get_monthly_series('FEDFUNDS')
        cpi_monthly = collector.get_monthly_series('CPIAUCSL')

        if not ffr_monthly.empty:
            for date, ffr_row in ffr_monthly.iterrows():
                bg = 'light_gray' if row % 2 == 0 else 'white'
                self._style_data(ws, row, 2, date.strftime('%Y-%m'), bg_color=bg)
                self._style_data(ws, row, 3, round(ffr_row['VALUE'], 2), bg_color=bg, number_format='0.00')

                # CPI YoY 추정
                if not cpi_monthly.empty:
                    target_1y = date - pd.DateOffset(months=12)
                    idx = cpi_monthly.index.searchsorted(target_1y)
                    if 0 <= idx < len(cpi_monthly):
                        current_cpi_val = None
                        cpi_idx = cpi_monthly.index.searchsorted(date)
                        if 0 <= cpi_idx < len(cpi_monthly):
                            current_cpi_val = cpi_monthly['VALUE'].iloc[cpi_idx]
                        past_cpi_val = cpi_monthly['VALUE'].iloc[idx]
                        if current_cpi_val is not None and past_cpi_val > 0:
                            yoy = ((current_cpi_val - past_cpi_val) / past_cpi_val) * 100
                            self._style_data(ws, row, 4, round(yoy, 2), bg_color=bg, number_format='0.00')
                            real = ffr_row['VALUE'] - yoy
                            real_bg = 'positive' if real > 0 else 'negative'
                            self._style_data(ws, row, 5, round(real, 2), bg_color=real_bg, number_format='0.00')
                        else:
                            self._style_data(ws, row, 4, 'N/A', bg_color=bg)
                            self._style_data(ws, row, 5, 'N/A', bg_color=bg)
                    else:
                        self._style_data(ws, row, 4, 'N/A', bg_color=bg)
                        self._style_data(ws, row, 5, 'N/A', bg_color=bg)
                else:
                    self._style_data(ws, row, 4, 'N/A', bg_color=bg)
                    self._style_data(ws, row, 5, 'N/A', bg_color=bg)

                ws.row_dimensions[row].height = 20
                row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 28
        for c_letter in 'CDEFG':
            ws.column_dimensions[c_letter].width = 18

    def add_cycle_sheet(self, cycle_analysis: dict, collector: MacroDataCollector):
        """시트 3: 경기 사이클"""
        ws = self.wb.create_sheet("경기 사이클")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "경기 사이클 진단"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 현재 사이클 단계
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "현재 경기 사이클 진단", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        stage_name = cycle_analysis.get('stage_name', 'N/A')
        stage_desc = cycle_analysis.get('stage_desc', '')
        score = cycle_analysis.get('score', 0)

        # 사이클 단계별 색상
        cycle_colors = {
            '확장기': ('EAFAF1', '1B7A43'),
            '정점기': ('FEF9E7', '7D6608'),
            '수축기': ('FDEDEC', 'C0392B'),
            '저점기': ('EBF5FB', '2E86C1'),
        }
        bg_color, font_color = cycle_colors.get(stage_name, ('F2F3F4', '0D1B2A'))

        cycle_items = [
            ('현재 경기 단계', stage_name),
            ('종합 점수', f"{score:+d} (양수: 확장, 음수: 수축)"),
            ('단계 설명', stage_desc),
            ('실업률', self._fmt_val(cycle_analysis.get('unemployment'), '%')),
            ('실업률 변화 (3M)', self._fmt_val(cycle_analysis.get('unemployment_chg'), '%p')),
            ('실질 GDP', self._fmt_val(cycle_analysis.get('gdp_current'), '십억$')),
            ('VIX 지수', self._fmt_val(cycle_analysis.get('vix'), '')),
            ('장단기 금리차', self._fmt_val(cycle_analysis.get('spread'), '%')),
        ]

        row = 6
        for key, value in cycle_items:
            this_bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, key, bg_color='light_blue', bold=True)
            ws.merge_cells(f'C{row}:G{row}')

            if key == '현재 경기 단계':
                cell = ws.cell(row=row, column=3, value=str(value))
                cell.fill = PatternFill("solid", fgColor=bg_color)
                cell.font = Font(size=12, name='맑은 고딕', bold=True, color=font_color)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                self._style_text(ws, row, 3, str(value), bg_color=this_bg)

            ws.row_dimensions[row].height = 26
            row += 1

        # 핵심 신호 목록
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "핵심 경기 신호", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        signal_headers = ['신호', '해석', '방향']
        for col, h in enumerate(signal_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        signals = cycle_analysis.get('signals', [])
        for signal_name, interpretation, direction in signals:
            if direction == 'positive':
                sig_bg = 'positive'
            elif direction == 'negative':
                sig_bg = 'negative'
            else:
                sig_bg = 'light_gray'

            self._style_text(ws, row, 2, signal_name, bg_color=sig_bg, bold=True)
            self._style_text(ws, row, 3, interpretation, bg_color=sig_bg)
            self._style_data(ws, row, 4, direction, bg_color=sig_bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # --- 히스토리 타임라인 (월간) ---
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "월간 주요지표 히스토리", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        hist_headers = ['날짜', '실업률(%)', 'VIX', '장단기차(%)', '달러인덱스']
        for col, h in enumerate(hist_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        # 월간 데이터 결합
        unrate_m = collector.get_monthly_series('UNRATE')
        vix_m = collector.get_monthly_series('VIXCLS')
        spread_m = collector.get_monthly_series('T10Y2Y')
        dollar_m = collector.get_monthly_series('DTWEXBGS')

        # 공통 날짜 기준 (실업률 기준)
        if not unrate_m.empty:
            for date in unrate_m.index:
                bg = 'light_gray' if row % 2 == 0 else 'white'
                self._style_data(ws, row, 2, date.strftime('%Y-%m'), bg_color=bg)
                self._style_data(ws, row, 3, round(unrate_m.loc[date, 'VALUE'], 1),
                                 bg_color=bg, number_format='0.0')

                # VIX
                vix_val = self._find_nearest(vix_m, date)
                self._style_data(ws, row, 4, round(vix_val, 1) if vix_val else 'N/A',
                                 bg_color=bg, number_format='0.0')

                # 장단기차
                spread_val = self._find_nearest(spread_m, date)
                if spread_val is not None:
                    sp_bg = 'negative' if spread_val < 0 else bg
                    self._style_data(ws, row, 5, round(spread_val, 2),
                                     bg_color=sp_bg, number_format='0.00')
                else:
                    self._style_data(ws, row, 5, 'N/A', bg_color=bg)

                # 달러인덱스
                dollar_val = self._find_nearest(dollar_m, date)
                self._style_data(ws, row, 6, round(dollar_val, 1) if dollar_val else 'N/A',
                                 bg_color=bg, number_format='0.0')

                ws.row_dimensions[row].height = 20
                row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 24
        for c_letter in 'CDEFG':
            ws.column_dimensions[c_letter].width = 16

    def add_asset_outlook_sheet(self, asset_outlook: dict, cycle_analysis: dict,
                                 rate_analysis: dict):
        """시트 4: 자산별 전망"""
        ws = self.wb.create_sheet("자산별 전망")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "자산별 투자 전망"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 현재 환경 요약
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "현재 매크로 환경 요약", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        env_items = [
            ('경기 단계', cycle_analysis.get('stage_name', 'N/A')),
            ('금리 사이클', rate_analysis.get('cycle_stage', 'N/A')),
            ('수익률 곡선', rate_analysis.get('yield_curve_status', 'N/A')),
        ]
        row = 6
        for key, value in env_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, key, bg_color='light_blue', bold=True)
            ws.merge_cells(f'C{row}:G{row}')
            self._style_text(ws, row, 3, str(value), bg_color=bg, bold=True)
            ws.row_dimensions[row].height = 24
            row += 1

        # 자산별 전망 매트릭스
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "자산 클래스별 전망", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        asset_headers = ['자산 클래스', '환경 영향', '투자 전략', '리스크 요인']
        for col, h in enumerate(asset_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        impact_colors = {
            '긍정': ('EAFAF1', '1B7A43'),
            '부정': ('FDEDEC', 'C0392B'),
            '중립': ('FEF9E7', '7D6608'),
        }

        for asset_class in FedAnalyzer.ASSET_CLASSES:
            info = asset_outlook.get(asset_class, {
                'impact': 'N/A', 'strategy': 'N/A', 'risk': 'N/A'
            })

            impact = info.get('impact', 'N/A')
            imp_bg, imp_font = impact_colors.get(impact, ('F2F3F4', '0D1B2A'))

            # 자산 클래스
            self._style_text(ws, row, 2, asset_class, bg_color='light_blue', bold=True)

            # 영향
            cell = ws.cell(row=row, column=3, value=impact)
            cell.fill = PatternFill("solid", fgColor=imp_bg)
            cell.font = Font(size=11, name='맑은 고딕', bold=True, color=imp_font)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 전략
            self._style_text(ws, row, 4, info.get('strategy', 'N/A'))

            # 리스크
            self._style_text(ws, row, 5, info.get('risk', 'N/A'), bg_color='FDEDEC')

            ws.row_dimensions[row].height = 50
            row += 1

        # 면책 조항
        row += 2
        ws.merge_cells(f'B{row}:G{row}')
        disclaimer = ws.cell(row=row, column=2,
            value="* 본 분석은 참고용이며 투자 권유가 아닙니다. 투자 결정은 본인의 판단과 책임하에 이루어져야 합니다.")
        disclaimer.font = Font(size=9, color='7F8C8D', name='맑은 고딕', italic=True)
        disclaimer.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 35

    def add_correlation_sheet(self, corr_df: pd.DataFrame, ticker: str, company_name: str):
        """시트 5 (옵션): 종목-매크로 상관관계"""
        if corr_df is None or corr_df.empty:
            return

        ws = self.wb.create_sheet("종목 상관관계")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = f"{company_name} ({ticker}) - 매크로 지표 상관관계"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 헤더
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "매크로 지표와 월간 수익률 상관관계", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        headers = ['매크로 지표', '상관계수', '데이터 포인트', '해석']
        row = 6
        for col, h in enumerate(headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        for _, r in corr_df.iterrows():
            bg = 'light_gray' if row % 2 == 0 else 'white'

            self._style_text(ws, row, 2, r['name'], bg_color=bg, bold=True)

            corr_val = r['correlation']
            if corr_val is not None:
                if abs(corr_val) > 0.6:
                    corr_bg = 'positive' if corr_val > 0 else 'negative'
                elif abs(corr_val) > 0.3:
                    corr_bg = 'light_blue'
                else:
                    corr_bg = bg
                self._style_data(ws, row, 3, corr_val, bg_color=corr_bg, number_format='0.0000')
            else:
                self._style_data(ws, row, 3, 'N/A', bg_color=bg)

            self._style_data(ws, row, 4, r['data_points'], bg_color=bg)
            self._style_text(ws, row, 5, r['interpretation'], bg_color=bg)

            ws.row_dimensions[row].height = 24
            row += 1

        # 설명
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        note = ws.cell(row=row, column=2,
            value="* 상관계수: +1(완전 양의 상관) ~ -1(완전 음의 상관). 월간 변화율 기준 계산.")
        note.font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 30

    def add_fomc_sheet(self, fomc_analysis: dict):
        """시트 5: FOMC 일정 및 금리 기대"""
        ws = self.wb.create_sheet("FOMC 분석")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "FOMC 회의 일정 및 금리 정책 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # --- 현재 금리 상태 ---
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "현재 연준 금리 현황", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        # 시장 기대 번역
        expectation_map = {
            'hold': '동결 (Hold)',
            'cut':  '인하 기대 (Cut Expected)',
            'hike': '인상 기대 (Hike Expected)',
        }
        expectation_colors = {
            'hold': 'FEF9E7',
            'cut':  'EAFAF1',
            'hike': 'FDEDEC',
        }
        market_exp = fomc_analysis.get('market_expectation', 'hold')
        exp_label = expectation_map.get(market_exp, market_exp)
        exp_bg = expectation_colors.get(market_exp, 'light_gray')

        rate_items = [
            ('FF목표금리 상단 (DFEDTARU)', self._fmt_val(fomc_analysis.get('current_target_upper'), '%')),
            ('FF목표금리 하단 (DFEDTARL)', self._fmt_val(fomc_analysis.get('current_target_lower'), '%')),
            ('목표금리 중간값',             self._fmt_val(fomc_analysis.get('target_mid'), '%')),
            ('실효 연방기금금리 (FEDFUNDS)', self._fmt_val(fomc_analysis.get('current_effective'), '%')),
            ('시장 금리 기대',              exp_label),
        ]

        row = 6
        for key, value in rate_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, key, bg_color='light_blue', bold=True)
            ws.merge_cells(f'C{row}:G{row}')
            # 시장 기대 행에는 색상 강조
            if key == '시장 금리 기대':
                cell = ws.cell(row=row, column=3, value=str(value))
                cell.fill = PatternFill("solid", fgColor=exp_bg)
                cell.font = Font(size=11, name='맑은 고딕', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                self._style_text(ws, row, 3, str(value), bg_color=bg)
            ws.row_dimensions[row].height = 26
            row += 1

        # --- 다음 FOMC 회의 ---
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "다음 FOMC 회의", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        next_meeting = fomc_analysis.get('next_meeting', 'N/A')
        days_until = fomc_analysis.get('days_until', 'N/A')

        meeting_items = [
            ('다음 FOMC 결정일', str(next_meeting)),
            ('D-day (남은 일수)', f"D-{days_until}" if isinstance(days_until, int) else str(days_until)),
        ]
        for key, value in meeting_items:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, key, bg_color='light_blue', bold=True)
            ws.merge_cells(f'C{row}:G{row}')
            self._style_text(ws, row, 3, value, bg_color=bg, bold=True)
            ws.row_dimensions[row].height = 26
            row += 1

        # --- 최근 금리 결정 이력 ---
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "최근 금리 결정 이력 (목표금리 변동 기준)", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        dec_headers = ['결정월', '목표금리(상단,%)', '변화폭(%p)', '결정']
        for col, h in enumerate(dec_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        recent_decisions = fomc_analysis.get('recent_decisions', [])
        if recent_decisions:
            for dec in recent_decisions:
                bg = 'light_gray' if row % 2 == 0 else 'white'
                action_bg = 'positive' if dec.get('action') == '인하' else ('negative' if dec.get('action') == '인상' else bg)
                self._style_data(ws, row, 2, dec.get('date', 'N/A'), bg_color=bg)
                self._style_data(ws, row, 3, dec.get('rate', 'N/A'), bg_color=bg, number_format='0.00')
                chg = dec.get('change', 0)
                self._style_data(ws, row, 4, f"{chg:+.2f}" if isinstance(chg, float) else chg, bg_color=action_bg)
                self._style_data(ws, row, 5, dec.get('action', 'N/A'), bg_color=action_bg, bold=True)
                ws.row_dimensions[row].height = 22
                row += 1
        else:
            ws.merge_cells(f'C{row}:F{row}')
            self._style_text(ws, row, 2, '(분석 기간 내 금리 변동 없음 또는 데이터 없음)', bg_color='light_gray')
            ws.row_dimensions[row].height = 22
            row += 1

        # --- 2026 FOMC 일정표 ---
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "2026년 FOMC 회의 일정", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        sched_headers = ['회차', '결정일', '비고']
        for col, h in enumerate(sched_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        today_str = datetime.now().strftime('%Y-%m-%d')
        schedule_2026 = fomc_analysis.get('fomc_schedule_2026', [])
        for idx, meeting_date in enumerate(schedule_2026, 1):
            bg = 'light_gray' if row % 2 == 0 else 'white'
            # 다음 회의는 강조
            if meeting_date == fomc_analysis.get('next_meeting'):
                bg = 'light_blue'
                note = '← 다음 회의'
            elif meeting_date < today_str:
                note = '완료'
            else:
                note = ''
            self._style_data(ws, row, 2, f"{idx}차", bg_color=bg)
            self._style_data(ws, row, 3, meeting_date, bg_color=bg, bold=(bg == 'light_blue'))
            self._style_text(ws, row, 4, note, bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 출처 주석
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        note_cell = ws.cell(row=row, column=2,
            value="* 출처: FRED (DFEDTARU/DFEDTARL/FEDFUNDS). FOMC 일정은 연준 공식 발표 기준.")
        note_cell.font = Font(size=9, color='7F8C8D', name='맑은 고딕', italic=True)
        note_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

    def save(self, filepath: str):
        self.wb.save(filepath)
        print(f"\n  [저장 완료] {filepath}")

    @staticmethod
    def _fmt_val(val, unit: str) -> str:
        """값 포맷팅"""
        if val is None:
            return 'N/A'
        if isinstance(val, float):
            return f"{val:.2f}{unit}"
        return f"{val}{unit}"

    @staticmethod
    def _find_nearest(monthly_df: pd.DataFrame, target_date) -> float:
        """월간 데이터에서 가장 가까운 값 찾기"""
        if monthly_df is None or monthly_df.empty:
            return None
        idx = monthly_df.index.searchsorted(target_date)
        if idx >= len(monthly_df):
            idx = len(monthly_df) - 1
        if idx < 0:
            return None
        return monthly_df['VALUE'].iloc[idx]


# ====================================================================
# 5. 콘솔 출력
# ====================================================================
def print_console_report(trends: dict, analysis: dict, series_map: dict):
    """콘솔에 매크로 분석 결과 출력"""

    print("\n" + "=" * 65)
    print("  연준/매크로 경제 분석 결과")
    print("=" * 65)

    # 1. 매크로 대시보드
    print("\n  --- 주요 매크로 지표 ---")
    print(f"  {'지표':<22s} {'현재값':>10s} {'3M변화':>10s} {'추세':>8s}")
    print("  " + "-" * 52)

    for series_id in series_map.keys():
        if series_id not in trends:
            continue
        t = trends[series_id]
        current = f"{t['current']:.2f}" if t['current'] is not None else 'N/A'
        chg = f"{t['chg_3m']:+.2f}" if t['chg_3m'] is not None else 'N/A'
        print(f"  {t['name']:<22s} {current:>10s} {chg:>10s} {t['arrow'] + t['direction']:>8s}")

    # 2. 금리 환경
    rates = analysis.get('rates', {})
    print(f"\n  --- 금리 환경 ---")
    print(f"  금리 사이클: [{rates.get('cycle_stage', 'N/A')}]")
    print(f"  {rates.get('cycle_desc', '')}")
    print(f"  수익률 곡선: {rates.get('yield_curve_status', 'N/A')}")

    # 3. 인플레이션
    inf = analysis.get('inflation', {})
    print(f"\n  --- 인플레이션 ---")
    print(f"  CPI YoY: {inf.get('cpi_yoy', 'N/A')}")
    if inf.get('cpi_yoy') is not None:
        print(f"           {inf['cpi_yoy']:.2f}%")
    print(f"  상태: {inf.get('inflation_status', 'N/A')}")
    print(f"  실질금리: {inf.get('real_rate', 'N/A')}")
    if inf.get('real_rate') is not None:
        print(f"            {inf['real_rate']:.2f}%")

    # 4. 경기 사이클
    cyc = analysis.get('cycle', {})
    print(f"\n  --- 경기 사이클 진단 ---")
    print(f"  현재 단계: [{cyc.get('stage_name', 'N/A')}] (점수: {cyc.get('score', 'N/A'):+d})")
    print(f"  {cyc.get('stage_desc', '')}")
    print(f"\n  핵심 신호:")
    for signal_name, interpretation, direction in cyc.get('signals', []):
        marker = '[+]' if direction == 'positive' else ('[-]' if direction == 'negative' else '[=]')
        print(f"    {marker} {signal_name} - {interpretation}")

    # 5. 자산별 전망
    assets = analysis.get('assets', {})
    print(f"\n  --- 자산별 전망 ---")
    print(f"  {'자산 클래스':<22s} {'영향':>6s}  전략 요약")
    print("  " + "-" * 60)
    for asset_class in FedAnalyzer.ASSET_CLASSES:
        info = assets.get(asset_class, {})
        impact = info.get('impact', 'N/A')
        strategy = info.get('strategy', 'N/A')
        # 전략 요약 (50자 이내)
        if len(strategy) > 50:
            strategy_short = strategy[:47] + '...'
        else:
            strategy_short = strategy
        print(f"  {asset_class:<22s} [{impact:>2s}]  {strategy_short}")

    # 6. FOMC 분석
    fomc = analysis.get('fomc', {})
    print(f"\n  --- FOMC 분석 ---")
    print(f"  다음 회의: {fomc.get('next_meeting', 'N/A')}  "
          f"(D-{fomc.get('days_until', 'N/A')})")
    upper = fomc.get('current_target_upper')
    lower = fomc.get('current_target_lower')
    eff   = fomc.get('current_effective')
    if upper is not None and lower is not None:
        eff_str = f"(실효: {eff:.2f}%)" if eff is not None else "(실효: N/A)"
        print(f"  목표금리: {lower:.2f}% ~ {upper:.2f}%  {eff_str}")
    expectation_label = {'hold': '동결', 'cut': '인하 기대', 'hike': '인상 기대'}
    print(f"  시장 기대: [{expectation_label.get(fomc.get('market_expectation', 'hold'), 'N/A')}]")
    recent = fomc.get('recent_decisions', [])
    if recent:
        print(f"  최근 결정:")
        for dec in recent:
            print(f"    {dec.get('date')}  {dec.get('action')}  "
                  f"({dec.get('change', 0):+.2f}%p → {dec.get('rate', 'N/A'):.2f}%)")

    print("\n" + "=" * 65)


# ====================================================================
# 6. 메인 실행
# ====================================================================
def main():
    parser = argparse.ArgumentParser(description="연준/매크로 경제 분석 시스템")
    parser.add_argument('--period', default='2y', help='분석 기간 (기본: 2y, 예: 1y, 6m, 3y)')
    parser.add_argument('--ticker', default=None, help='종목코드 (선택, 예: 005930, AAPL) - 매크로 상관관계 분석 추가')
    args = parser.parse_args()

    print("\n" + "=" * 65)
    print("  연준/매크로 경제 분석 시스템 - FRED 데이터 기반")
    print("=" * 65)
    print(f"  분석 기간: {args.period}")
    print(f"  실행 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    if args.ticker:
        print(f"  종목 상관관계: {args.ticker}")

    # 1. 데이터 수집
    collector = MacroDataCollector(period=args.period)
    collector.fetch_all()
    collector.calculate_trends()

    # 2. 분석
    analyzer = FedAnalyzer(collector)
    analysis = analyzer.run_full_analysis()

    # 3. 콘솔 출력
    print_console_report(collector.trends, analysis, collector.SERIES_MAP)

    # 4. 종목 상관관계 (옵션)
    corr_df = None
    corr_company = None
    if args.ticker:
        print(f"\n[추가분석] {args.ticker} 종목-매크로 상관관계 분석 중...")
        stock_corr = StockMacroCorrelation(args.ticker, collector)
        stock_ret = stock_corr.fetch_stock_data()
        if not stock_ret.empty:
            corr_df = stock_corr.calculate_correlations()
            corr_company = getattr(stock_corr, 'company_name', args.ticker)
            if corr_df is not None and not corr_df.empty:
                print(f"\n  --- {corr_company} 매크로 상관관계 ---")
                for _, r in corr_df.iterrows():
                    corr_val = f"{r['correlation']:.4f}" if r['correlation'] is not None else 'N/A'
                    print(f"  {r['name']:<22s} : {corr_val:>8s}  ({r['interpretation']})")

    # 5. Excel 생성
    print("\n[저장] Excel 보고서 생성 중...")
    builder = MacroExcelBuilder()
    builder.add_dashboard_sheet(collector.trends, collector.SERIES_MAP)
    builder.add_rates_inflation_sheet(analysis['rates'], analysis['inflation'], collector)
    builder.add_cycle_sheet(analysis['cycle'], collector)
    builder.add_asset_outlook_sheet(analysis['assets'], analysis['cycle'], analysis['rates'])
    builder.add_fomc_sheet(analysis['fomc'])

    if corr_df is not None and not corr_df.empty:
        builder.add_correlation_sheet(corr_df, args.ticker, corr_company)

    date_str = datetime.now().strftime('%Y%m%d')
    output_path = f"output/매크로분석_{date_str}.xlsx"
    builder.save(output_path)

    # 6. 결과 안내
    print("\n" + "=" * 65)
    print("  완료!")
    print("=" * 65)
    print(f"\n  생성 파일: {output_path}")
    if collector.failed:
        print(f"\n  [참고] 수집 실패 지표: {', '.join(collector.failed)}")
    print(f"\n  다음 단계:")
    print(f"  - Excel 파일을 열어 매크로 환경을 확인하세요.")
    print(f"  - financial_analyzer.py, hmm_regime_detector.py와 함께 종합 분석에 활용하세요.")
    print()


if __name__ == "__main__":
    main()
