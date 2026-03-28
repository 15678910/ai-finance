"""
AI 금융 분석 시스템 - Python 자동화 스크립트
==============================================
사용법:
  1. pip install -r requirements.txt
  2. python financial_analyzer.py --ticker 005930  (삼성전자)
  3. python financial_analyzer.py --ticker AAPL    (애플, 미국주식)

생성 파일:
  - output/[종목코드]_재무데이터.xlsx   → Cowork에 업로드해서 분석
  - output/[종목코드]_공시목록.pdf      → Cowork에 업로드해서 분석
"""

import os
import sys
import json
import argparse
import warnings
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')

# ── 필수 라이브러리 설치 확인 ──────────────────────────────────
def check_and_install():
    required = {
        'yfinance': 'yfinance',
        'pandas': 'pandas',
        'openpyxl': 'openpyxl',
        'requests': 'requests',
        'dart_fss': 'dart-fss',
    }
    missing = []
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[설치 필요] pip install {' '.join(missing)}")
        sys.exit(1)

check_and_install()

import yfinance as yf
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 출력 폴더 생성 ──────────────────────────────────────────────
Path("output").mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════
# 1. Yahoo Finance 재무 데이터 수집
# ══════════════════════════════════════════════════════════════════
class FinancialDataCollector:

    def __init__(self, ticker: str):
        # 국내 종목은 .KS 추가 (코스피) 또는 .KQ (코스닥)
        if ticker.isdigit():
            self.ticker_raw = ticker
            # 코스닥/코스피 자동 감지
            self.ticker = self._detect_market(ticker)
            self.is_korean = True
        else:
            self.ticker_raw = ticker
            self.ticker = ticker
            self.is_korean = False

        print(f"\n📊 [{self.ticker}] 데이터 수집 시작...")
        self.stock = yf.Ticker(self.ticker)
        self.info = self.stock.info
        self.company_name = self.info.get('longName') or self.info.get('shortName', ticker)

    @staticmethod
    def _detect_market(ticker_code):
        """코스피(.KS) / 코스닥(.KQ) 자동 감지"""
        for suffix in [".KS", ".KQ"]:
            try:
                t = yf.Ticker(ticker_code + suffix)
                hist = t.history(period="5d")
                if not hist.empty and len(hist) > 0:
                    info = t.info or {}
                    name = info.get('longName') or info.get('shortName', '')
                    if name and not name.startswith(ticker_code):
                        return ticker_code + suffix
            except Exception:
                pass
        print(f"  [!] 시장 자동 감지 실패, 코스피(.KS)로 시도합니다.")
        return ticker_code + ".KS"

    def get_price_data(self, period="2y") -> pd.DataFrame:
        """주가 데이터 수집 (기본 2년)"""
        print("  ├─ 주가 데이터 수집 중...")
        df = self.stock.history(period=period)
        if df.empty:
            print(f"  [!] {self.ticker} 주가 데이터를 찾을 수 없습니다.")
            return pd.DataFrame()
        if df.index.tz is not None:
            df.index = df.index.tz_localize(None)
        return df

    def get_financials(self) -> dict:
        """재무제표 수집"""
        print("  ├─ 재무제표 수집 중...")
        return {
            'income_stmt': self.stock.income_stmt,
            'balance_sheet': self.stock.balance_sheet,
            'cash_flow': self.stock.cashflow,
            'quarterly_income': self.stock.quarterly_income_stmt,
        }

    def get_key_metrics(self) -> dict:
        """핵심 투자지표"""
        print("  ├─ 투자지표 수집 중...")
        info = self.info
        metrics = {
            '기업명': self.company_name,
            '티커': self.ticker_raw,
            '섹터': info.get('sector', 'N/A'),
            '산업': info.get('industry', 'N/A'),
            '시가총액(억)': round(info.get('marketCap', 0) / 1e8, 0) if info.get('marketCap') else 'N/A',
            '현재주가': info.get('currentPrice') or info.get('regularMarketPrice', 'N/A'),
            '52주최고': info.get('fiftyTwoWeekHigh', 'N/A'),
            '52주최저': info.get('fiftyTwoWeekLow', 'N/A'),
            'PER': round(info.get('trailingPE', 0), 2) if info.get('trailingPE') else 'N/A',
            'Forward PER': round(info.get('forwardPE', 0), 2) if info.get('forwardPE') else 'N/A',
            'PBR': round(info.get('priceToBook', 0), 2) if info.get('priceToBook') else 'N/A',
            'EV/EBITDA': round(info.get('enterpriseToEbitda', 0), 2) if info.get('enterpriseToEbitda') else 'N/A',
            'ROE(%)': round(info.get('returnOnEquity', 0) * 100, 2) if info.get('returnOnEquity') else 'N/A',
            'ROA(%)': round(info.get('returnOnAssets', 0) * 100, 2) if info.get('returnOnAssets') else 'N/A',
            '배당수익률(%)': round(info.get('dividendYield', 0) * 100, 2) if info.get('dividendYield') else 'N/A',
            '부채비율(%)': round(info.get('debtToEquity', 0), 2) if info.get('debtToEquity') else 'N/A',
            '매출성장률(%)': round(info.get('revenueGrowth', 0) * 100, 2) if info.get('revenueGrowth') else 'N/A',
            '영업이익률(%)': round(info.get('operatingMargins', 0) * 100, 2) if info.get('operatingMargins') else 'N/A',
            '순이익률(%)': round(info.get('profitMargins', 0) * 100, 2) if info.get('profitMargins') else 'N/A',
            '베타': round(info.get('beta', 0), 3) if info.get('beta') else 'N/A',
            '직원수': info.get('fullTimeEmployees', 'N/A'),
            '기업설명': info.get('longBusinessSummary', 'N/A')[:200] + '...' if info.get('longBusinessSummary') else 'N/A',
        }
        return metrics


# ══════════════════════════════════════════════════════════════════
# 2. Excel 보고서 생성
# ══════════════════════════════════════════════════════════════════
class ExcelReportBuilder:

    # 색상 팔레트
    COLORS = {
        'header_dark': '0D1B2A',
        'header_blue': '1B4F72',
        'accent': '2E86C1',
        'light_blue': 'D6EAF8',
        'white': 'FFFFFF',
        'light_gray': 'F2F3F4',
        'positive': 'EAFAF1',
        'negative': 'FDEDEC',
        'gold': 'F39C12',
    }

    def __init__(self, ticker: str, company_name: str):
        self.wb = Workbook()
        self.ticker = ticker
        self.company_name = company_name
        self.wb.remove(self.wb.active)  # 기본 시트 제거

    def _style_header(self, ws, row, col, value, color='header_dark', font_color='FFFFFF', bold=True, size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=self.COLORS.get(color, color))
        cell.font = Font(bold=bold, color=font_color, size=size, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def _style_data(self, ws, row, col, value, bg_color=None, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if number_format:
            cell.number_format = number_format
        return cell

    def add_summary_sheet(self, metrics: dict):
        """시트 1: 기업 요약"""
        ws = self.wb.create_sheet("📋 기업요약")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = f"{self.company_name} ({self.ticker}) — 기업 분석 요약"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 날짜
        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 핵심 지표
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "핵심 투자지표", 'header_dark', size=11)
        ws.row_dimensions[5].height = 28

        row = 6
        skip_keys = ['기업설명']
        for key, value in metrics.items():
            if key in skip_keys:
                continue
            bg = 'light_gray' if row % 2 == 0 else 'white'
            ws.cell(row=row, column=2, value=key).font = Font(bold=True, size=10, name='맑은 고딕')
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            self._style_data(ws, row, 3, value, bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 기업 설명
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "기업 설명", 'header_blue', size=10)
        row += 1
        ws.merge_cells(f'B{row}:G{row+3}')
        desc_cell = ws.cell(row=row, column=2, value=metrics.get('기업설명', ''))
        desc_cell.font = Font(size=10, name='맑은 고딕')
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[row].height = 80

        # Cowork 안내 메시지
        row += 6
        ws.merge_cells(f'B{row}:G{row}')
        guide = ws.cell(row=row, column=2,
            value="💡 [다음 단계] 이 파일을 Claude Cowork에 업로드 → /dcf-model 또는 /comps-analysis 명령어로 심층 분석하세요.")
        guide.font = Font(size=10, color='1B4F72', name='맑은 고딕', bold=True)
        guide.fill = PatternFill("solid", fgColor='D6EAF8')
        guide.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 30

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        for col in 'DEFG':
            ws.column_dimensions[col].width = 15

    def add_price_sheet(self, price_df: pd.DataFrame):
        """시트 2: 주가 데이터"""
        ws = self.wb.create_sheet("📈 주가데이터")
        ws.sheet_view.showGridLines = False

        headers = ['날짜', '시가', '고가', '저가', '종가', '거래량', '등락률(%)']
        for col, h in enumerate(headers, 1):
            self._style_header(ws, 1, col, h, 'header_dark')
            ws.column_dimensions[get_column_letter(col)].width = 16

        # 최근 252거래일 (1년)
        recent = price_df.tail(252).copy()
        recent['등락률'] = recent['Close'].pct_change() * 100

        for row_idx, (date, row_data) in enumerate(recent.iterrows(), 2):
            change = row_data.get('등락률', 0)
            bg = 'positive' if change > 0 else ('negative' if change < 0 else 'white')
            bg = 'light_gray' if row_idx % 2 == 0 and bg == 'white' else bg

            ws.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
            for col, field in enumerate(['Open', 'High', 'Low', 'Close'], 2):
                val = round(row_data.get(field, 0), 0)
                self._style_data(ws, row_idx, col, val, bg_color=bg, number_format='#,##0')
            vol = row_data.get('Volume', 0)
            vol = 0 if pd.isna(vol) else int(vol)
            self._style_data(ws, row_idx, 6, vol, bg_color=bg, number_format='#,##0')
            self._style_data(ws, row_idx, 7, round(change, 2), bg_color=bg, number_format='0.00')
            ws.row_dimensions[row_idx].height = 18

        ws.row_dimensions[1].height = 28

    def add_financials_sheet(self, financials: dict):
        """시트 3: 재무제표"""
        ws = self.wb.create_sheet("📊 재무제표")
        ws.sheet_view.showGridLines = False

        # 손익계산서
        ws.merge_cells('A1:E1')
        self._style_header(ws, 1, 1, "손익계산서 (Income Statement)", 'header_dark', size=11)
        ws.row_dimensions[1].height = 30

        income = financials.get('income_stmt')
        if income is not None and not income.empty:
            # 헤더 (연도)
            for col, year in enumerate(income.columns, 2):
                year_str = str(year)[:10] if hasattr(year, '__str__') else str(year)
                self._style_header(ws, 2, col, year_str, 'header_blue', size=10)
                ws.column_dimensions[get_column_letter(col)].width = 18

            ws.column_dimensions['A'].width = 35

            key_items = {
                'Total Revenue': '매출액',
                'Gross Profit': '매출총이익',
                'Operating Income': '영업이익',
                'EBITDA': 'EBITDA',
                'Net Income': '순이익',
                'Basic EPS': 'EPS(기본)',
            }

            row = 3
            for eng_key, kor_label in key_items.items():
                ws.cell(row=row, column=1, value=kor_label)
                ws.cell(row=row, column=1).font = Font(bold=True, size=10, name='맑은 고딕')
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
                bg = 'light_gray' if row % 2 == 0 else 'white'

                for col, year in enumerate(income.columns, 2):
                    try:
                        val = income.loc[eng_key, year] if eng_key in income.index else 'N/A'
                        if val != 'N/A' and pd.notna(val):
                            if eng_key == 'Basic EPS':
                                # EPS는 주당 단위이므로 억 단위 변환하지 않음
                                self._style_data(ws, row, col, round(val, 0), bg_color=bg, number_format='#,##0')
                            else:
                                val = round(val / 1e8, 0)  # 억 단위
                                self._style_data(ws, row, col, val, bg_color=bg, number_format='#,##0')
                        else:
                            self._style_data(ws, row, col, 'N/A', bg_color=bg)
                    except Exception:
                        self._style_data(ws, row, col, 'N/A', bg_color=bg)
                ws.row_dimensions[row].height = 22
                row += 1

            # 단위 안내
            row += 1
            ws.cell(row=row, column=1, value="* 단위: 억원 (EPS 제외)").font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # Cowork 프롬프트 안내
        row = ws.max_row + 3
        ws.merge_cells(f'A{row}:E{row}')
        guide = ws.cell(row=row, column=1,
            value="💡 [Cowork 프롬프트] 이 파일 첨부 후: '/3-statement-model 이 재무 데이터로 통합 3-statement 모델 구축하고 향후 3년 예측 추가해줘'")
        guide.font = Font(size=10, color='1B4F72', name='맑은 고딕', bold=True)
        guide.fill = PatternFill("solid", fgColor='D6EAF8')
        guide.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
        ws.row_dimensions[row].height = 40

    def add_prompts_sheet(self):
        """시트 4: Cowork 프롬프트 가이드"""
        ws = self.wb.create_sheet("💬 Cowork프롬프트")
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:D1')
        self._style_header(ws, 1, 1, "Claude Cowork 프롬프트 가이드", 'header_dark', size=12)
        ws.row_dimensions[1].height = 35

        prompts = [
            ("DCF 가치평가", "/dcf-model",
             f"{self.company_name} DCF 모델 구축. WACC 8.5%, 5년 성장률 12%, 터미널 3%. 민감도 분석 포함. Excel로 생성.",
             "재무제표 시트 포함한 이 파일"),
            ("동종사 비교", "/comps-analysis",
             f"{self.company_name}와 동종 섹터 10개 기업 비교. EV/EBITDA, P/E, P/B, EV/Revenue 배수 포함. 프리미엄/디스카운트 근거 포함.",
             "기업요약 시트 포함한 이 파일"),
            ("3-Statement 모델", "/3-statement-model",
             f"이 재무 데이터로 손익계산서/대차대조표/현금흐름표 통합 모델 구축. 향후 3년 예측 포함. Excel 완성본 생성.",
             "재무제표 시트 포함한 이 파일"),
            ("경쟁 환경 분석", "/competitive-analysis",
             f"{self.company_name}의 경쟁 환경 분석. Porter's 5 Forces, 시장점유율, 핵심 경쟁우위/위협 요인 포함. 전략적 권고사항 포함.",
             "기업요약 시트 포함한 이 파일"),
            ("투자 메모 작성", "/pitch-deck",
             f"{self.company_name} 투자 메모 작성. Executive Summary, 비즈니스 모델, 성장 동력, 리스크, 밸류에이션, 투자 결론 포함. PPT 형식.",
             "모든 시트 포함한 이 파일"),
        ]

        headers = ['분석 유형', '슬래시 명령어', '프롬프트 내용', '첨부 파일']
        for col, h in enumerate(headers, 1):
            self._style_header(ws, 2, col, h, 'header_blue', size=10)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 28

        for row_idx, (ptype, cmd, prompt, attach) in enumerate(prompts, 3):
            bg = 'light_gray' if row_idx % 2 == 0 else 'white'
            ws.cell(row=row_idx, column=1, value=ptype).font = Font(bold=True, size=10, color='1B4F72', name='맑은 고딕')
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=2, value=cmd).font = Font(size=10, name='맑은 고딕', color='E74C3C')
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')
            prompt_cell = ws.cell(row=row_idx, column=3, value=prompt)
            prompt_cell.font = Font(size=10, name='맑은 고딕')
            prompt_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_idx, column=4, value=attach).font = Font(size=10, name='맑은 고딕', color='7F8C8D')
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor=self.COLORS.get(bg, bg))
            ws.row_dimensions[row_idx].height = 55

        ws.row_dimensions[2].height = 28

    def save(self, output_path: str):
        self.wb.save(output_path)
        print(f"  └─ Excel 저장 완료: {output_path}")


# ══════════════════════════════════════════════════════════════════
# 3. DART 공시 수집 (한국 주식)
# ══════════════════════════════════════════════════════════════════
class DartCollector:
    """
    DART Open API 활용
    API 키 발급: https://opendart.fss.or.kr/
    """

    BASE_URL = "https://opendart.fss.or.kr/api"

    def __init__(self, api_key: str = None):
        self.api_key = api_key or os.getenv('DART_API_KEY', '')
        if not self.api_key:
            print("  ⚠️  DART API 키 없음. 환경변수 DART_API_KEY 설정 필요.")
            print("     발급: https://opendart.fss.or.kr/")

    def search_company(self, corp_name: str) -> dict:
        """기업명으로 corp_code 검색"""
        if not self.api_key:
            return {}
        url = f"{self.BASE_URL}/company.json"
        params = {'crtfc_key': self.api_key, 'corp_name': corp_name, 'page_count': 5}
        try:
            resp = requests.get(url, params=params, timeout=10)
            data = resp.json()
            if data.get('status') == '000':
                return data.get('list', [])
        except Exception as e:
            print(f"  ⚠️  DART 검색 오류: {e}")
        return []

    def get_disclosures(self, corp_code: str, days: int = 180) -> list:
        """최근 공시 목록"""
        if not self.api_key:
            return []
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=days)
        url = f"{self.BASE_URL}/list.json"
        params = {
            'crtfc_key': self.api_key,
            'corp_code': corp_code,
            'bgn_de': start_dt.strftime('%Y%m%d'),
            'end_de': end_dt.strftime('%Y%m%d'),
            'page_count': 40,
        }
        try:
            resp = requests.get(url, params=params, timeout=10)
            data = resp.json()
            if data.get('status') == '000':
                return data.get('list', [])
        except Exception as e:
            print(f"  ⚠️  공시 수집 오류: {e}")
        return []


# ══════════════════════════════════════════════════════════════════
# 4. 메인 실행
# ══════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description='AI 금융 분석 시스템 — 데이터 수집기')
    parser.add_argument('--ticker', required=True, help='종목코드 (예: 005930, AAPL)')
    parser.add_argument('--period', default='2y', help='주가 수집 기간 (기본: 2y)')
    args = parser.parse_args()

    ticker = args.ticker.upper()
    print("\n" + "="*60)
    print("  AI 금융 분석 시스템 — 데이터 수집 및 Excel 생성")
    print("="*60)

    # 1. 데이터 수집
    collector = FinancialDataCollector(ticker)
    metrics = collector.get_key_metrics()
    price_data = collector.get_price_data(args.period)
    if price_data.empty:
        print(f"\n  [!] 주가 데이터가 없습니다. 티커를 확인하세요: {ticker}")
        sys.exit(1)
    financials = collector.get_financials()

    print(f"\n  ✅ 기업명: {collector.company_name}")
    print(f"  ✅ 섹터: {metrics.get('섹터', 'N/A')}")
    print(f"  ✅ 시가총액: {metrics.get('시가총액(억)', 'N/A'):,}억원" if isinstance(metrics.get('시가총액(억)'), (int, float)) else f"  ✅ 시가총액: {metrics.get('시가총액(억)', 'N/A')}")

    # 2. Excel 생성
    print("\n📝 Excel 보고서 생성 중...")
    builder = ExcelReportBuilder(ticker, collector.company_name)
    builder.add_summary_sheet(metrics)
    builder.add_price_sheet(price_data)
    builder.add_financials_sheet(financials)
    builder.add_prompts_sheet()

    output_path = f"output/{ticker}_금융분석_{datetime.now().strftime('%Y%m%d')}.xlsx"
    builder.save(output_path)

    # 3. 결과 안내
    print("\n" + "="*60)
    print("  🎉 완료!")
    print("="*60)
    print(f"\n  📁 생성 파일: {output_path}")
    print("\n  📌 다음 단계 (Claude Cowork):")
    print("  ┌─────────────────────────────────────────────────────")
    print("  │ 1. Claude Desktop → Cowork 탭 → 새 작업")
    print(f"  │ 2. '{output_path}' 파일 첨부")
    print("  │ 3. 아래 명령어 중 선택:")
    print(f"  │    /dcf-model {collector.company_name} DCF 가치평가 모델 구축")
    print(f"  │    /comps-analysis {collector.company_name} 동종사 비교 분석")
    print(f"  │    /3-statement-model 통합 재무모델 구축 및 3년 예측")
    print("  └─────────────────────────────────────────────────────\n")

    # 4. JSON 저장 (Claude API 연동용)
    json_path = f"output/{ticker}_metrics.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(metrics, f, ensure_ascii=False, indent=2, default=str)
    print(f"  📄 JSON 저장 (API 연동용): {json_path}")


if __name__ == "__main__":
    main()
