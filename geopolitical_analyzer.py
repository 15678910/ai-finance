"""
지정학 리스크 분석 시스템 - 무료 데이터 기반
================================================
사용법:
  1. pip install requests pandas numpy openpyxl yfinance
  2. python geopolitical_analyzer.py                    (전체 분석)
  3. python geopolitical_analyzer.py --category 관세     (특정 카테고리만)
  4. python geopolitical_analyzer.py --days 7            (최근 7일 뉴스)

생성 파일:
  - output/지정학리스크_[날짜].xlsx
"""

import os
import sys
import argparse
import warnings
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')

# -- 필수 라이브러리 설치 확인 ----------------------------------------
def check_and_install():
    required = {
        'requests': 'requests',
        'pandas': 'pandas',
        'numpy': 'numpy',
        'openpyxl': 'openpyxl',
        'yfinance': 'yfinance',
    }
    missing = []
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[설치 필요] pip install {' '.join(missing)}")
        sys.exit(1)

check_and_install()

import numpy as np
import pandas as pd
import requests
import yfinance as yf
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- 출력 폴더 생성 ---------------------------------------------------
Path("output").mkdir(exist_ok=True)


# ====================================================================
# 0. 헤드라인 번역 (키워드 기반)
# ====================================================================

HEADLINE_KR_MAP = {
    "war": "전쟁",
    "Iran": "이란",
    "Russia": "러시아",
    "Ukraine": "우크라이나",
    "Trump": "트럼프",
    "ceasefire": "휴전",
    "attack": "공격",
    "drone": "드론",
    "drones": "드론",
    "missile": "미사일",
    "sanctions": "제재",
    "tariff": "관세",
    "China": "중국",
    "North Korea": "북한",
    "nuclear": "핵",
    "oil": "석유/유가",
    "trade": "무역",
    "military": "군사",
    "conflict": "분쟁",
    "peace": "평화",
    "NATO": "나토",
    "U.S.": "미국",
    "Japan": "일본",
    "Israel": "이스라엘",
    "Gaza": "가자",
    "Hamas": "하마스",
    "Hezbollah": "헤즈볼라",
    "Syria": "시리아",
    "Taiwan": "대만",
    "Korea": "한국",
    "Asia": "아시아",
    "Europe": "유럽",
    "launches": "발사",
    "threatens": "위협",
    "bomb": "폭탄",
    "invasion": "침공",
    "defense": "방어",
    "weapons": "무기",
    "Biden": "바이든",
    "plan": "계획",
    "says": "발언",
    "report": "보도",
    "President": "대통령",
    "largest": "최대규모",
    "effort": "노력",
    "end": "종결",
    "accept": "수용",
    "received": "수령",
    "shows": "보여줌",
    "norms": "규범",
    "overturned": "전복",
    "accelerate": "가속화",
    "shift": "전환",
    "renewable": "재생에너지",
    "moment": "순간",
    "international": "국제",
    "period": "기간",
}


def translate_headline(title: str) -> str:
    """영어 헤드라인 키워드에 한글 주석을 괄호로 추가합니다.
    예: 'Iran war' → 'Iran(이란) war(전쟁)'
    긴 키워드(다중 단어)를 먼저 처리합니다.
    """
    result = title
    # 다중 단어 키워드 우선 처리 (긴 것 먼저)
    sorted_keys = sorted(HEADLINE_KR_MAP.keys(), key=lambda k: len(k), reverse=True)
    replaced = set()
    for kw in sorted_keys:
        if kw in replaced:
            continue
        kr = HEADLINE_KR_MAP[kw]
        # 이미 주석이 달린 부분은 건너뜀
        if f"{kw}({kr})" in result:
            continue
        if kw in result:
            result = result.replace(kw, f"{kw}({kr})", 1)
            replaced.add(kw)
    return result


# ====================================================================
# 1. 지정학 뉴스 수집기
# ====================================================================
class GeopoliticalNewsCollector:
    """RSS 피드 기반 지정학 뉴스 수집 및 리스크 분류"""

    # RSS 피드 소스
    RSS_FEEDS = {
        'Reuters': 'https://feeds.reuters.com/Reuters/worldNews',
        'CNBC': 'https://search.cnbc.com/rs/search/combinedcms/view.xml?partnerId=wrss01&id=100727362',
        'BBC': 'https://feeds.bbci.co.uk/news/world/rss.xml',
        'Google(EN)': 'https://news.google.com/rss/search?q=geopolitics+tariff+war+sanctions&hl=en',
        'Google(KR)': 'https://news.google.com/rss/search?q=%EA%B4%80%EC%84%B8+%EC%A0%84%EC%9F%81+%EC%A0%9C%EC%9E%AC+%EB%AC%B4%EC%97%AD%EB%B6%84%EC%9F%81&hl=ko&gl=KR',
    }

    # 리스크 카테고리 정의
    RISK_CATEGORIES = {
        "전쟁/군사충돌": {
            "keywords_en": [
                "war", "military", "attack", "missile", "strike", "troops",
                "invasion", "conflict", "combat", "airstrike", "drone strike",
                "ceasefire", "NATO", "nuclear",
            ],
            "keywords_kr": [
                "전쟁", "군사", "공격", "미사일", "폭격", "침공",
                "충돌", "교전", "공습", "드론", "핵",
            ],
            "affected_sectors": ["방산", "에너지", "항공"],
            "risk_weight": 3,
        },
        "관세/무역분쟁": {
            "keywords_en": [
                "tariff", "trade war", "trade dispute", "import duty",
                "export ban", "trade restriction", "trade deal", "WTO",
                "dumping", "quota",
            ],
            "keywords_kr": [
                "관세", "무역전쟁", "무역분쟁", "수입규제",
                "수출규제", "반덤핑", "통상",
            ],
            "affected_sectors": ["반도체", "자동차", "철강", "배터리"],
            "risk_weight": 2,
        },
        "경제제재": {
            "keywords_en": [
                "sanction", "embargo", "blacklist", "asset freeze",
                "export control", "entity list", "OFAC",
            ],
            "keywords_kr": [
                "제재", "금수조치", "블랙리스트", "수출통제", "엔티티리스트",
            ],
            "affected_sectors": ["반도체", "에너지", "금융"],
            "risk_weight": 2,
        },
        "지역갈등": {
            "keywords_en": [
                "Taiwan", "South China Sea", "Ukraine", "Russia", "Iran",
                "Israel", "Gaza", "North Korea", "Middle East", "Red Sea", "Houthi",
            ],
            "keywords_kr": [
                "대만", "남중국해", "우크라이나", "러시아", "이란",
                "이스라엘", "가자", "북한", "중동", "홍해", "후티",
            ],
            "affected_sectors": ["방산", "에너지", "해운"],
            "risk_weight": 2,
        },
        "공급망위기": {
            "keywords_en": [
                "supply chain", "chip shortage", "rare earth", "lithium",
                "semiconductor ban", "TSMC", "decoupling",
            ],
            "keywords_kr": [
                "공급망", "반도체부족", "희토류", "리튬", "디커플링", "칩",
            ],
            "affected_sectors": ["반도체", "배터리", "자동차", "IT"],
            "risk_weight": 2,
        },
        "정치불안": {
            "keywords_en": [
                "coup", "protest", "election crisis", "impeach",
                "regime change", "martial law", "political crisis",
            ],
            "keywords_kr": [
                "쿠데타", "시위", "탄핵", "계엄", "정치위기", "정권교체",
            ],
            "affected_sectors": ["금융", "부동산", "내수"],
            "risk_weight": 1,
        },
    }

    def __init__(self, days: int = 3, category_filter: str = None):
        self.days = days
        self.category_filter = category_filter
        self.raw_news = []        # 전체 수집 뉴스
        self.scored_news = []     # 리스크 점수가 부여된 뉴스
        self.category_scores = {} # 카테고리별 리스크 점수
        self.feed_counts = {}     # 소스별 수집 건수

    def fetch_all_feeds(self) -> list:
        """모든 RSS 피드에서 뉴스 수집"""
        print("\n[수집] RSS 뉴스 피드 수집 중...")

        cutoff_date = datetime.now() - timedelta(days=self.days)

        for source_name, url in self.RSS_FEEDS.items():
            count = self._fetch_single_feed(source_name, url, cutoff_date)
            self.feed_counts[source_name] = count

        # 소스별 수집 결과 출력
        parts = []
        for src, cnt in self.feed_counts.items():
            parts.append(f"{src}: {cnt}건")
        print(f"  {', '.join(parts)}")
        print(f"  총 {len(self.raw_news)}건 수집")

        return self.raw_news

    def _fetch_single_feed(self, source_name: str, url: str, cutoff_date: datetime) -> int:
        """단일 RSS 피드 파싱"""
        count = 0
        try:
            resp = requests.get(url, timeout=10, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) Finance-Analyzer/1.0'
            })
            resp.raise_for_status()

            root = ET.fromstring(resp.content)

            # RSS 2.0 형식: channel/item
            items = root.findall('.//item')
            if not items:
                # Atom 형식 시도: entry
                items = root.findall('.//{http://www.w3.org/2005/Atom}entry')

            for item in items:
                title = self._get_text(item, 'title')
                link = self._get_text(item, 'link')
                pub_date_str = self._get_text(item, 'pubDate')
                if not pub_date_str:
                    pub_date_str = self._get_text(item, 'published')

                if not title:
                    continue

                pub_date = self._parse_date(pub_date_str)

                # 날짜 필터링 (파싱 실패 시 포함)
                if pub_date and pub_date < cutoff_date:
                    continue

                self.raw_news.append({
                    'title': title,
                    'link': link or '',
                    'date': pub_date.strftime('%Y-%m-%d %H:%M') if pub_date else 'N/A',
                    'date_obj': pub_date,
                    'source': source_name,
                })
                count += 1

        except Exception as e:
            print(f"  [!] {source_name} 수집 실패: {e}")

        return count

    @staticmethod
    def _get_text(element, tag_name: str) -> str:
        """XML 요소에서 텍스트 추출"""
        el = element.find(tag_name)
        if el is not None and el.text:
            return el.text.strip()
        # Atom 네임스페이스 시도
        el = element.find(f'{{http://www.w3.org/2005/Atom}}{tag_name}')
        if el is not None and el.text:
            return el.text.strip()
        return ''

    @staticmethod
    def _parse_date(date_str: str) -> datetime:
        """다양한 날짜 형식 파싱"""
        if not date_str:
            return None

        # RFC 822 형식 (RSS 표준)
        formats = [
            '%a, %d %b %Y %H:%M:%S %z',
            '%a, %d %b %Y %H:%M:%S GMT',
            '%Y-%m-%dT%H:%M:%SZ',
            '%Y-%m-%dT%H:%M:%S%z',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
        ]
        for fmt in formats:
            try:
                dt = datetime.strptime(date_str.strip(), fmt)
                # timezone-aware를 naive로 변환
                if dt.tzinfo is not None:
                    dt = dt.replace(tzinfo=None)
                return dt
            except ValueError:
                continue

        return None

    def score_news(self) -> list:
        """뉴스 리스크 점수 계산 및 카테고리 분류"""
        print("\n[분석] 뉴스 리스크 점수 산정 중...")

        # 카테고리별 점수 초기화
        for cat_name in self.RISK_CATEGORIES:
            self.category_scores[cat_name] = {
                'score': 0,
                'news_count': 0,
                'top_news': [],
            }

        geo_count = 0

        for news in self.raw_news:
            title_lower = news['title'].lower()
            news['categories'] = []
            news['risk_score'] = 0
            news['matched_keywords'] = []

            total_score = 0

            for cat_name, cat_info in self.RISK_CATEGORIES.items():
                # 카테고리 필터 적용
                if self.category_filter:
                    if self.category_filter not in cat_name:
                        continue

                matched = []
                # 영어 키워드 매칭
                for kw in cat_info['keywords_en']:
                    if kw.lower() in title_lower:
                        matched.append(kw)
                # 한국어 키워드 매칭
                for kw in cat_info['keywords_kr']:
                    if kw in news['title']:
                        matched.append(kw)

                if matched:
                    cat_score = len(matched) * cat_info['risk_weight']
                    total_score += cat_score
                    news['categories'].append(cat_name)
                    news['matched_keywords'].extend(matched)

                    self.category_scores[cat_name]['score'] += cat_score
                    self.category_scores[cat_name]['news_count'] += 1
                    self.category_scores[cat_name]['top_news'].append({
                        'title': news['title'],
                        'title_kr': translate_headline(news['title']),
                        'link': news.get('link', ''),
                        'score': cat_score,
                        'source': news['source'],
                        'date': news['date'],
                    })

            if total_score > 0:
                news['risk_score'] = total_score
                self.scored_news.append(news)
                geo_count += 1

        # 각 카테고리의 top_news를 점수 내림차순 정렬, 상위 3건만 유지
        for cat_name in self.category_scores:
            tops = self.category_scores[cat_name]['top_news']
            tops.sort(key=lambda x: x['score'], reverse=True)
            self.category_scores[cat_name]['top_news'] = tops[:3]

        # 전체 scored_news를 점수 내림차순 정렬
        self.scored_news.sort(key=lambda x: x['risk_score'], reverse=True)

        print(f"  지정학 관련 {geo_count}건 감지")

        return self.scored_news

    def get_normalized_news_risk(self) -> float:
        """뉴스 리스크 점수를 0-100 범위로 정규화"""
        if not self.scored_news:
            return 0.0

        total_score = sum(n['risk_score'] for n in self.scored_news)
        news_count = len(self.scored_news)

        # 평균 리스크 점수 기반, 최대 예상값 대비 정규화
        avg_score = total_score / max(news_count, 1)
        # 평균 점수 6 이상이면 100점 만점으로 간주
        normalized = min(100, (avg_score / 6.0) * 100)

        # 뉴스 건수 보정: 관련 뉴스가 많을수록 리스크 증가
        count_factor = min(1.5, 1.0 + (news_count / 50.0))
        normalized = min(100, normalized * count_factor)

        return round(normalized, 1)


# ====================================================================
# 2. 안전자산 모니터
# ====================================================================
class SafeHavenMonitor:
    """안전자산 가격 추이를 통한 리스크 수준 평가"""

    # 모니터링 자산 목록: (티커, 한글명, 유형)
    ASSETS = {
        'GC=F':  ('금(Gold)', 'commodity'),
        'CL=F':  ('유가(WTI)', 'commodity'),
        '^VIX':  ('VIX', 'volatility'),
        'KRW=X': ('USD/KRW', 'currency'),
        '^TNX':  ('미국10년국채', 'bond'),
        'ITA':   ('방산ETF(ITA)', 'equity'),
        'EEM':   ('신흥국ETF(EEM)', 'equity'),
    }

    def __init__(self):
        self.asset_data = {}   # 티커 -> DataFrame
        self.asset_stats = {}  # 티커 -> 통계 dict
        self.risk_score = 0.0  # 종합 리스크 온도계 (0-100)
        self.risk_level = ''   # 레벨 텍스트

    def fetch_all(self):
        """모든 안전자산 30일 데이터 수집"""
        print("\n[수집] 안전자산 가격 데이터 수집 중...")

        for ticker, (name, asset_type) in self.ASSETS.items():
            try:
                data = yf.download(ticker, period='1mo', progress=False)
                if data is not None and not data.empty:
                    # MultiIndex 컬럼 처리
                    if isinstance(data.columns, pd.MultiIndex):
                        data.columns = data.columns.get_level_values(0)
                    self.asset_data[ticker] = data
                    print(f"  {name}: {len(data)}일 데이터 수집")
                else:
                    print(f"  [!] {name}: 데이터 없음 (건너뜀)")
            except Exception as e:
                print(f"  [!] {name}: 수집 실패 - {e}")

    def calculate_stats(self) -> dict:
        """각 자산의 변화율 및 통계 계산"""
        print("\n[분석] 안전자산 통계 계산 중...")

        for ticker, (name, asset_type) in self.ASSETS.items():
            if ticker not in self.asset_data:
                self.asset_stats[ticker] = {
                    'name': name,
                    'type': asset_type,
                    'current': None,
                    'chg_1w': None,
                    'chg_1m': None,
                    'available': False,
                }
                continue

            df = self.asset_data[ticker]
            close = df['Close']

            if len(close) < 2:
                self.asset_stats[ticker] = {
                    'name': name,
                    'type': asset_type,
                    'current': None,
                    'chg_1w': None,
                    'chg_1m': None,
                    'available': False,
                }
                continue

            current = float(close.iloc[-1])

            # 1주 변화율
            if len(close) >= 5:
                week_ago = float(close.iloc[-5])
                chg_1w = ((current - week_ago) / week_ago) * 100 if week_ago != 0 else 0
            else:
                chg_1w = None

            # 1월 변화율
            month_ago = float(close.iloc[0])
            chg_1m = ((current - month_ago) / month_ago) * 100 if month_ago != 0 else 0

            self.asset_stats[ticker] = {
                'name': name,
                'type': asset_type,
                'current': round(current, 2),
                'chg_1w': round(chg_1w, 2) if chg_1w is not None else None,
                'chg_1m': round(chg_1m, 2),
                'available': True,
            }

        return self.asset_stats

    def calculate_risk_thermometer(self, news_risk_score: float) -> tuple:
        """
        종합 리스크 온도계 (0-100) 계산
        - VIX 기여: 30%
        - 금 추세: 20%
        - 유가 추세: 15%
        - USD/KRW 추세: 15%
        - 뉴스 리스크: 20%
        """
        scores = {}

        # VIX 기여 (30%) - 현재 수준 기반
        vix_stat = self.asset_stats.get('^VIX', {})
        if vix_stat.get('available') and vix_stat.get('current') is not None:
            vix_val = vix_stat['current']
            # VIX 12 이하 = 0점, VIX 40 이상 = 100점
            vix_score = min(100, max(0, (vix_val - 12) / 28 * 100))
            scores['vix'] = vix_score
        else:
            scores['vix'] = 50  # 기본값

        # 금 추세 (20%) - 상승 = 리스크 상승
        gold_stat = self.asset_stats.get('GC=F', {})
        if gold_stat.get('available') and gold_stat.get('chg_1w') is not None:
            gold_chg = gold_stat['chg_1w']
            # 주간 +5% 이상 = 100점, -5% 이하 = 0점
            gold_score = min(100, max(0, (gold_chg + 5) / 10 * 100))
            scores['gold'] = gold_score
        else:
            scores['gold'] = 50

        # 유가 추세 (15%) - 급등 = 지정학 프리미엄
        oil_stat = self.asset_stats.get('CL=F', {})
        if oil_stat.get('available') and oil_stat.get('chg_1w') is not None:
            oil_chg = oil_stat['chg_1w']
            oil_score = min(100, max(0, (oil_chg + 5) / 10 * 100))
            scores['oil'] = oil_score
        else:
            scores['oil'] = 50

        # USD/KRW 추세 (15%) - 원화 약세(상승) = 리스크 상승
        krw_stat = self.asset_stats.get('KRW=X', {})
        if krw_stat.get('available') and krw_stat.get('chg_1w') is not None:
            krw_chg = krw_stat['chg_1w']
            krw_score = min(100, max(0, (krw_chg + 3) / 6 * 100))
            scores['krw'] = krw_score
        else:
            scores['krw'] = 50

        # 뉴스 리스크 (20%)
        scores['news'] = news_risk_score

        # 가중 평균
        composite = (
            scores['vix'] * 0.30 +
            scores['gold'] * 0.20 +
            scores['oil'] * 0.15 +
            scores['krw'] * 0.15 +
            scores['news'] * 0.20
        )

        self.risk_score = round(min(100, max(0, composite)), 1)

        # 레벨 판정
        if self.risk_score <= 20:
            self.risk_level = '안정'
        elif self.risk_score <= 40:
            self.risk_level = '관심'
        elif self.risk_score <= 60:
            self.risk_level = '주의'
        elif self.risk_score <= 80:
            self.risk_level = '경계'
        else:
            self.risk_level = '위험'

        return self.risk_score, self.risk_level


# ====================================================================
# 3. 섹터 영향 분석기
# ====================================================================
class SectorImpactAnalyzer:
    """지정학 리스크가 한국 시장 섹터에 미치는 영향 분석"""

    # 섹터별 관련 종목
    SECTOR_STOCKS = {
        '방산': [
            ('한화에어로스페이스', '012450'),
            ('LIG넥스원', '079550'),
            ('현대로템', '064350'),
        ],
        '에너지': [
            ('한국전력', '015760'),
            ('SK이노베이션', '096770'),
            ('S-Oil', '010950'),
        ],
        '반도체': [
            ('삼성전자', '005930'),
            ('SK하이닉스', '000660'),
        ],
        '배터리': [
            ('LG에너지솔루션', '373220'),
            ('삼성SDI', '006400'),
            ('에코프로비엠', '247540'),
        ],
        '자동차': [
            ('현대자동차', '005380'),
            ('기아', '000270'),
        ],
        '해운': [
            ('HMM', '011200'),
            ('팬오션', '028670'),
        ],
        '항공': [
            ('대한항공', '003490'),
            ('아시아나항공', '020560'),
        ],
        '철강': [
            ('포스코홀딩스', '005490'),
            ('현대제철', '004020'),
        ],
        '금융': [
            ('KB금융', '105560'),
            ('신한지주', '055550'),
        ],
        'IT': [
            ('네이버', '035420'),
            ('카카오', '035720'),
        ],
    }

    # 리스크별 섹터 영향 매트릭스 (수혜/피해/중립/혼조)
    IMPACT_MATRIX = {
        '전쟁/군사충돌': {
            '방산': ('수혜', '군사 긴장 고조로 방산 수요 증가 전망'),
            '에너지': ('혼조', '유가 상승 수혜 vs 공급 불안'),
            '반도체': ('중립', '직접 영향 제한적'),
            '배터리': ('중립', '직접 영향 제한적'),
            '자동차': ('피해', '글로벌 수요 위축 우려'),
            '해운': ('피해', '항로 리스크 증가, 운임 변동'),
            '항공': ('피해', '항공 노선 제한, 유가 부담'),
            '철강': ('혼조', '방산 수요 vs 글로벌 경기 둔화'),
            '금융': ('피해', '리스크 프리미엄 확대'),
            'IT': ('중립', '간접 영향'),
        },
        '관세/무역분쟁': {
            '방산': ('중립', '직접 영향 제한적'),
            '에너지': ('피해', '무역 제한으로 원자재 조달 비용 증가'),
            '반도체': ('피해', '대중국 수출 규제 리스크'),
            '배터리': ('피해', '공급망 비용 상승, 수출 제한'),
            '자동차': ('피해', '관세 직접 타격, 수출 경쟁력 약화'),
            '해운': ('혼조', '물동량 감소 vs 운임 변동'),
            '항공': ('중립', '간접 영향'),
            '철강': ('피해', '반덤핑 관세 직접 타격'),
            '금융': ('피해', '수출 기업 실적 악화 파급'),
            'IT': ('중립', '내수 비중 높아 제한적'),
        },
        '경제제재': {
            '방산': ('수혜', '안보 위기의식 강화'),
            '에너지': ('혼조', '에너지 수급 불안 vs 가격 상승'),
            '반도체': ('피해', '기술 수출 통제 리스크'),
            '배터리': ('피해', '원자재 수급 불안'),
            '자동차': ('중립', '간접 영향'),
            '해운': ('피해', '교역량 감소 우려'),
            '항공': ('중립', '간접 영향'),
            '철강': ('중립', '간접 영향'),
            '금융': ('피해', '제재 대상 관련 자산 리스크'),
            'IT': ('중립', '간접 영향'),
        },
        '지역갈등': {
            '방산': ('수혜', '지역 군비 경쟁 수혜'),
            '에너지': ('혼조', '중동 갈등 시 유가 프리미엄'),
            '반도체': ('피해', '대만 리스크 직접 영향'),
            '배터리': ('중립', '간접 영향'),
            '자동차': ('중립', '간접 영향'),
            '해운': ('피해', '홍해/남중국해 항로 리스크'),
            '항공': ('피해', '갈등 지역 운항 제한'),
            '철강': ('중립', '간접 영향'),
            '금융': ('피해', '지정학 불확실성 할인'),
            'IT': ('중립', '간접 영향'),
        },
        '공급망위기': {
            '방산': ('중립', '간접 영향'),
            '에너지': ('혼조', '에너지 전환 수요 vs 비용'),
            '반도체': ('피해', '칩 부족 장기화, 공급 차질'),
            '배터리': ('피해', '희토류/리튬 수급 불안'),
            '자동차': ('피해', '부품 수급 차질'),
            '해운': ('수혜', '공급망 재편에 따른 물류 수요'),
            '항공': ('중립', '간접 영향'),
            '철강': ('중립', '간접 영향'),
            '금융': ('중립', '간접 영향'),
            'IT': ('피해', '부품 조달 비용 증가'),
        },
        '정치불안': {
            '방산': ('중립', '간접 영향'),
            '에너지': ('중립', '간접 영향'),
            '반도체': ('중립', '간접 영향'),
            '배터리': ('중립', '간접 영향'),
            '자동차': ('피해', '내수 소비 위축'),
            '해운': ('중립', '간접 영향'),
            '항공': ('피해', '여행 수요 위축'),
            '철강': ('중립', '간접 영향'),
            '금융': ('피해', '시장 불확실성 확대'),
            'IT': ('피해', '내수 소비 위축'),
        },
    }

    def __init__(self, category_scores: dict):
        self.category_scores = category_scores
        self.sector_impacts = {}      # 섹터 -> 영향 분석 결과
        self.sector_exposure = {}     # 섹터 -> 종합 리스크 노출도
        self.sector_strategies = {}   # 섹터 -> 전략 권고

    def analyze(self) -> dict:
        """활성 리스크 기반 섹터별 영향 분석"""
        print("\n[분석] 섹터 영향 분석 중...")

        # 활성 카테고리 식별 (뉴스 1건 이상)
        active_categories = {}
        for cat_name, cat_data in self.category_scores.items():
            if cat_data['news_count'] > 0:
                active_categories[cat_name] = cat_data

        # 섹터별 영향 집계
        for sector_name in self.SECTOR_STOCKS:
            impacts = []
            total_risk_weight = 0
            positive_count = 0
            negative_count = 0

            for cat_name, cat_data in active_categories.items():
                if cat_name in self.IMPACT_MATRIX:
                    sector_impact = self.IMPACT_MATRIX[cat_name].get(sector_name)
                    if sector_impact:
                        impact_type, description = sector_impact
                        cat_score = cat_data['score']

                        impacts.append({
                            'category': cat_name,
                            'impact': impact_type,
                            'description': description,
                            'score': cat_score,
                        })

                        total_risk_weight += cat_score
                        if impact_type == '수혜':
                            positive_count += 1
                        elif impact_type == '피해':
                            negative_count += 1

            # 종합 리스크 노출도 결정
            if total_risk_weight == 0:
                exposure = '낮음'
            elif negative_count > positive_count and total_risk_weight > 10:
                exposure = '높음'
            elif total_risk_weight > 5:
                exposure = '보통'
            else:
                exposure = '낮음'

            # 전략 권고 생성
            strategy = self._generate_strategy(sector_name, impacts, exposure)

            self.sector_impacts[sector_name] = impacts
            self.sector_exposure[sector_name] = exposure
            self.sector_strategies[sector_name] = strategy

        return self.sector_impacts

    def _generate_strategy(self, sector_name: str, impacts: list, exposure: str) -> str:
        """섹터별 전략 권고 생성"""
        if not impacts:
            return f"{sector_name} 섹터: 현재 지정학 리스크 영향 제한적. 기존 투자 전략 유지 권고."

        # 주요 영향 분석
        benefit_cats = [i['category'] for i in impacts if i['impact'] == '수혜']
        damage_cats = [i['category'] for i in impacts if i['impact'] == '피해']

        parts = []
        if benefit_cats:
            parts.append(f"수혜 요인: {', '.join(benefit_cats)}")
        if damage_cats:
            parts.append(f"리스크 요인: {', '.join(damage_cats)}")

        if exposure == '높음':
            action = "비중 축소 또는 헤지 전략 검토 필요"
        elif exposure == '보통':
            action = "선별적 접근, 모니터링 강화 권고"
        else:
            action = "기존 전략 유지, 기회 포착 가능"

        return f"{'; '.join(parts)}. {action}."

    def get_impact_matrix_data(self) -> list:
        """섹터 x 카테고리 영향 매트릭스 데이터"""
        matrix_rows = []
        for sector_name in self.SECTOR_STOCKS:
            row = {'sector': sector_name}
            for cat_name in self.category_scores:
                if cat_name in self.IMPACT_MATRIX:
                    sector_info = self.IMPACT_MATRIX[cat_name].get(sector_name, ('중립', ''))
                    row[cat_name] = sector_info[0]
                else:
                    row[cat_name] = '중립'
            row['exposure'] = self.sector_exposure.get(sector_name, '낮음')
            matrix_rows.append(row)
        return matrix_rows


# ====================================================================
# 4. Excel 보고서 빌더
# ====================================================================
class GeopoliticalExcelBuilder:
    """지정학 리스크 분석 Excel 보고서 빌더"""

    # 색상 팔레트 (financial_analyzer.py 동일)
    COLORS = {
        'header_dark': '0D1B2A',
        'header_blue': '1B4F72',
        'accent': '2E86C1',
        'light_blue': 'D6EAF8',
        'white': 'FFFFFF',
        'light_gray': 'F2F3F4',
        'positive': 'EAFAF1',
        'negative': 'FDEDEC',
        'gold': 'F39C12',
    }

    # 리스크 레벨별 색상
    RISK_LEVEL_COLORS = {
        '안정': ('EAFAF1', '1E8449'),
        '관심': ('FEF9E7', '7D6608'),
        '주의': ('FDF2E9', 'CA6F1E'),
        '경계': ('FDEDEC', 'E74C3C'),
        '위험': ('922B21', 'FFFFFF'),
    }

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 기본 시트 제거

    def _style_header(self, ws, row, col, value, color='header_dark', font_color='FFFFFF', bold=True, size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=self.COLORS.get(color, color))
        cell.font = Font(bold=bold, color=font_color, size=size, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def _style_data(self, ws, row, col, value, bg_color=None, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if number_format:
            cell.number_format = number_format
        return cell

    def _style_text(self, ws, row, col, value, bg_color=None, bold=False, indent=1):
        """좌측 정렬 텍스트 셀"""
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=indent)
        return cell

    def add_dashboard_sheet(self, risk_score, risk_level, asset_stats, summary_text):
        """시트 1: 리스크 대시보드"""
        ws = self.wb.create_sheet("리스크 대시보드")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = "지정학 리스크 대시보드"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        ws.merge_cells('B3:H3')
        date_cell = ws['B3']
        date_cell.value = f"분석일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        date_cell.font = Font(size=9, color='666666', name='맑은 고딕')
        date_cell.alignment = Alignment(horizontal='center')

        # 리스크 온도계
        ws.merge_cells('B5:H5')
        self._style_header(ws, 5, 2, "리스크 온도계", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        # 점수 및 레벨
        level_bg, level_font = self.RISK_LEVEL_COLORS.get(risk_level, ('F2F3F4', '0D1B2A'))
        ws.merge_cells('B6:H6')
        score_cell = ws.cell(row=6, column=2, value=f"[{risk_level}]  {risk_score}/100")
        score_cell.fill = PatternFill("solid", fgColor=level_bg)
        score_cell.font = Font(bold=True, size=18, color=level_font, name='맑은 고딕')
        score_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[6].height = 50

        # 게이지 바
        ws.merge_cells('B7:H7')
        filled = int(risk_score / 10)
        empty = 10 - filled
        gauge = "[" + "=" * filled + " " * empty + f"] {risk_score}/100"
        gauge_cell = ws.cell(row=7, column=2, value=gauge)
        gauge_cell.font = Font(size=10, name='Consolas', color='0D1B2A')
        gauge_cell.alignment = Alignment(horizontal='center', vertical='center')
        gauge_cell.fill = PatternFill("solid", fgColor='F2F3F4')
        ws.row_dimensions[7].height = 25

        # Safe-Haven 자산 현황
        row = 9
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "Safe-Haven 자산 현황", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        asset_headers = ['자산', '현재가', '1주 변화(%)', '1월 변화(%)', '시그널']
        for col, h in enumerate(asset_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        for ticker, stat in asset_stats.items():
            if not stat.get('available'):
                continue

            bg = 'light_gray' if row % 2 == 0 else 'white'

            # 자산명
            self._style_text(ws, row, 2, stat['name'], bg_color=bg, bold=True)

            # 현재가
            current = stat.get('current')
            if current is not None:
                if current >= 1000:
                    fmt_str = '#,##0.0'
                else:
                    fmt_str = '#,##0.00'
                self._style_data(ws, row, 3, current, bg_color=bg, number_format=fmt_str)
            else:
                self._style_data(ws, row, 3, 'N/A', bg_color=bg)

            # 1주 변화
            chg_1w = stat.get('chg_1w')
            if chg_1w is not None:
                chg_bg = 'positive' if chg_1w >= 0 else 'negative'
                self._style_data(ws, row, 4, f"{chg_1w:+.2f}%", bg_color=chg_bg)
            else:
                self._style_data(ws, row, 4, 'N/A', bg_color=bg)

            # 1월 변화
            chg_1m = stat.get('chg_1m')
            if chg_1m is not None:
                chg_bg = 'positive' if chg_1m >= 0 else 'negative'
                self._style_data(ws, row, 5, f"{chg_1m:+.2f}%", bg_color=chg_bg)
            else:
                self._style_data(ws, row, 5, 'N/A', bg_color=bg)

            # 시그널
            signal = self._get_asset_signal(ticker, stat)
            self._style_text(ws, row, 6, signal, bg_color=bg)

            ws.row_dimensions[row].height = 24
            row += 1

        # 종합 판단
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "종합 판단", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f'B{row}:H{row+1}')
        summary_cell = ws.cell(row=row, column=2, value=summary_text)
        summary_cell.font = Font(size=10, name='맑은 고딕', color='1B4F72')
        summary_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
        summary_cell.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
        ws.row_dimensions[row].height = 50

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 5
        ws.column_dimensions['H'].width = 5

    @staticmethod
    def _get_asset_signal(ticker, stat):
        """자산별 리스크 시그널 텍스트"""
        chg = stat.get('chg_1w')
        if chg is None:
            return '데이터 부족'

        if ticker == 'GC=F':
            if chg > 2:
                return '<- 리스크 상승 신호 (금 강세)'
            elif chg < -2:
                return '<- 리스크 완화 신호 (금 약세)'
            return '금 가격 안정'
        elif ticker == 'CL=F':
            if chg > 3:
                return '<- 지정학 프리미엄 가능성'
            elif chg < -3:
                return '<- 수요 둔화 신호'
            return '유가 안정'
        elif ticker == '^VIX':
            current = stat.get('current', 0)
            if current > 30:
                return '<- 공포 수준 (높은 변동성)'
            elif current > 20:
                return '<- 경계 수준'
            return '정상 범위'
        elif ticker == 'KRW=X':
            if chg > 1:
                return '<- 원화 약세 (리스크오프)'
            elif chg < -1:
                return '<- 원화 강세 (리스크온)'
            return '환율 안정'
        elif ticker == 'ITA':
            if chg > 2:
                return '<- 방산주 강세 (군사 긴장)'
            return '방산ETF 안정'
        elif ticker == 'EEM':
            if chg < -2:
                return '<- 신흥국 위험 회피'
            return '신흥국 안정'
        elif ticker == '^TNX':
            if chg > 3:
                return '<- 금리 상승 (인플레 우려)'
            elif chg < -3:
                return '<- 안전자산 선호'
            return '국채 금리 안정'

        return ''

    def add_category_sheet(self, category_scores: dict, risk_categories: dict):
        """시트 2: 리스크 카테고리 분석"""
        ws = self.wb.create_sheet("리스크 카테고리 분석")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = "리스크 카테고리별 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 헤더
        row = 4
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "카테고리별 리스크 현황", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        headers = ['카테고리', '리스크 점수', '관련 뉴스', '핵심 뉴스', '영향 섹터']
        for col, h in enumerate(headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        # 카테고리별 데이터 (점수 내림차순)
        sorted_cats = sorted(category_scores.items(), key=lambda x: x[1]['score'], reverse=True)

        for cat_name, cat_data in sorted_cats:
            score = cat_data['score']
            news_count = cat_data['news_count']
            top_news = cat_data.get('top_news', [])

            # 리스크 수준에 따른 배경색
            if score >= 15:
                bg = 'negative'
            elif score >= 5:
                bg = 'FDF2E9'  # 노란/주황 경고
            else:
                bg = 'positive'

            # 카테고리명
            self._style_text(ws, row, 2, cat_name, bg_color=bg, bold=True)

            # 리스크 점수
            self._style_data(ws, row, 3, score, bg_color=bg, bold=True)

            # 관련 뉴스 수
            self._style_data(ws, row, 4, f"{news_count}건", bg_color=bg)

            # 핵심 뉴스 (상위 3건 제목)
            if top_news:
                news_titles = '; '.join([n['title'][:40] for n in top_news[:3]])
            else:
                news_titles = '관련 뉴스 없음'
            self._style_text(ws, row, 5, news_titles, bg_color=bg)

            # 영향 섹터
            cat_info = risk_categories.get(cat_name, {})
            sectors = ', '.join(cat_info.get('affected_sectors', []))
            self._style_text(ws, row, 6, sectors, bg_color=bg)

            ws.row_dimensions[row].height = 30
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 5
        ws.column_dimensions['H'].width = 5

    def add_news_sheet(self, scored_news: list):
        """시트 3: 주요 뉴스"""
        ws = self.wb.create_sheet("주요 뉴스")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:I2')
        title = ws['B2']
        title.value = "지정학 리스크 관련 주요 뉴스"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 헤더
        row = 4
        ws.merge_cells(f'B{row}:I{row}')
        self._style_header(ws, row, 2, f"최근 뉴스 목록 (상위 {min(50, len(scored_news))}건)", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        headers = ['날짜', '제목', '출처', '카테고리', '리스크 점수', '한글제목', '링크']
        for col, h in enumerate(headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        # 뉴스 목록 (최대 50건)
        for news in scored_news[:50]:
            risk = news['risk_score']

            # 리스크 점수별 배경색
            if risk >= 6:
                bg = 'negative'
            elif risk >= 3:
                bg = 'FDF2E9'
            else:
                bg = 'positive'

            # 날짜
            self._style_data(ws, row, 2, news.get('date', 'N/A'), bg_color=bg)

            # 제목
            title_text = news['title'][:80]
            self._style_text(ws, row, 3, title_text, bg_color=bg)

            # 출처
            self._style_data(ws, row, 4, news.get('source', 'N/A'), bg_color=bg)

            # 카테고리
            cats = ', '.join(news.get('categories', []))
            self._style_text(ws, row, 5, cats if cats else '분류없음', bg_color=bg)

            # 리스크 점수
            self._style_data(ws, row, 6, risk, bg_color=bg, bold=True)

            # 한글제목 (키워드 번역)
            title_kr = translate_headline(news['title'])
            self._style_text(ws, row, 7, title_kr[:100], bg_color=bg)

            # 링크 (URL)
            link_val = news.get('link', '')
            self._style_text(ws, row, 8, link_val[:200], bg_color=bg)

            ws.row_dimensions[row].height = 26
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 55
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 45
        ws.column_dimensions['H'].width = 50

    def add_sector_sheet(self, sector_analyzer: SectorImpactAnalyzer, category_scores: dict):
        """시트 4: 섹터 영향 분석"""
        ws = self.wb.create_sheet("섹터 영향 분석")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:I2')
        title = ws['B2']
        title.value = "섹터별 지정학 리스크 영향 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # --- 영향 매트릭스 ---
        row = 4
        cat_names = [c for c in category_scores.keys()]
        total_cols = 2 + len(cat_names) + 1  # 섹터 + 카테고리들 + 노출도
        end_col_letter = get_column_letter(total_cols)
        ws.merge_cells(f'B{row}:{end_col_letter}{row}')
        self._style_header(ws, row, 2, "섹터 x 리스크 카테고리 영향 매트릭스", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        # 매트릭스 헤더
        self._style_header(ws, row, 2, '섹터', 'header_blue', size=10)
        col = 3
        for cat_name in cat_names:
            # 카테고리명 줄여서 표시
            short_name = cat_name.split('/')[0] if '/' in cat_name else cat_name
            self._style_header(ws, row, col, short_name, 'header_blue', size=9)
            col += 1
        self._style_header(ws, row, col, '노출도', 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        # 매트릭스 데이터
        matrix_data = sector_analyzer.get_impact_matrix_data()
        for m_row in matrix_data:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, m_row['sector'], bg_color=bg, bold=True)

            col = 3
            for cat_name in cat_names:
                impact = m_row.get(cat_name, '중립')
                if impact == '수혜':
                    imp_bg = 'positive'
                elif impact == '피해':
                    imp_bg = 'negative'
                elif impact == '혼조':
                    imp_bg = 'FDF2E9'
                else:
                    imp_bg = bg
                self._style_data(ws, row, col, impact, bg_color=imp_bg)
                col += 1

            # 노출도
            exposure = m_row.get('exposure', '낮음')
            if exposure == '높음':
                exp_bg = 'negative'
            elif exposure == '보통':
                exp_bg = 'FDF2E9'
            else:
                exp_bg = 'positive'
            self._style_data(ws, row, col, exposure, bg_color=exp_bg, bold=True)

            ws.row_dimensions[row].height = 24
            row += 1

        # --- 섹터별 전략 권고 ---
        row += 1
        ws.merge_cells(f'B{row}:{end_col_letter}{row}')
        self._style_header(ws, row, 2, "섹터별 전략 권고", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        strategy_headers = ['섹터', '리스크 노출도', '전략 권고']
        for col_idx, h in enumerate(strategy_headers, 2):
            self._style_header(ws, row, col_idx, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        for sector_name in sector_analyzer.SECTOR_STOCKS:
            bg = 'light_gray' if row % 2 == 0 else 'white'
            self._style_text(ws, row, 2, sector_name, bg_color=bg, bold=True)

            exposure = sector_analyzer.sector_exposure.get(sector_name, '낮음')
            if exposure == '높음':
                exp_bg = 'negative'
            elif exposure == '보통':
                exp_bg = 'FDF2E9'
            else:
                exp_bg = 'positive'
            self._style_data(ws, row, 3, exposure, bg_color=exp_bg, bold=True)

            strategy = sector_analyzer.sector_strategies.get(sector_name, '')
            ws.merge_cells(f'D{row}:{end_col_letter}{row}')
            self._style_text(ws, row, 4, strategy, bg_color=bg)

            ws.row_dimensions[row].height = 30
            row += 1

        # --- 종목별 대응 전략 ---
        row += 1
        ws.merge_cells(f'B{row}:{end_col_letter}{row}')
        self._style_header(ws, row, 2, "종목별 대응 전략", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        stock_headers = ['섹터', '종목명', '종목코드', '섹터 노출도', '대응 방향']
        for col_idx, h in enumerate(stock_headers, 2):
            self._style_header(ws, row, col_idx, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 28
        row += 1

        for sector_name, stocks in sector_analyzer.SECTOR_STOCKS.items():
            exposure = sector_analyzer.sector_exposure.get(sector_name, '낮음')
            for stock_name, stock_code in stocks:
                bg = 'light_gray' if row % 2 == 0 else 'white'
                self._style_text(ws, row, 2, sector_name, bg_color=bg)
                self._style_text(ws, row, 3, stock_name, bg_color=bg, bold=True)
                self._style_data(ws, row, 4, stock_code, bg_color=bg)

                if exposure == '높음':
                    exp_bg = 'negative'
                    direction = '비중 축소 검토'
                elif exposure == '보통':
                    exp_bg = 'FDF2E9'
                    direction = '모니터링 강화'
                else:
                    exp_bg = 'positive'
                    direction = '기존 전략 유지'

                self._style_data(ws, row, 5, exposure, bg_color=exp_bg, bold=True)
                self._style_text(ws, row, 6, direction, bg_color=bg)

                ws.row_dimensions[row].height = 22
                row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        for i in range(3, total_cols + 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 14

    def save(self, filename: str):
        """Excel 파일 저장"""
        filepath = os.path.join('output', filename)
        self.wb.save(filepath)
        return filepath


# ====================================================================
# 5. 콘솔 출력기
# ====================================================================
class ConsoleReporter:
    """콘솔 분석 결과 출력"""

    @staticmethod
    def print_report(news_collector, safe_haven, sector_analyzer):
        """전체 분석 결과 콘솔 출력"""

        print("\n" + "=" * 60)
        print("  지정학 리스크 분석 시스템")
        print("=" * 60)

        # 수집 결과
        total = len(news_collector.raw_news)
        geo = len(news_collector.scored_news)
        feed_parts = []
        for src, cnt in news_collector.feed_counts.items():
            feed_parts.append(f"{src}: {cnt}건")
        print(f"\n[수집] {', '.join(feed_parts)}")
        print(f"  총 {total}건 수집, 지정학 관련 {geo}건 감지")

        # 리스크 온도계
        score = safe_haven.risk_score
        level = safe_haven.risk_level
        print(f"\n[분석] 리스크 온도계: [{level}] {score}/100")

        # 카테고리별 리스크
        print("\n  --- 카테고리별 리스크 ---")
        max_name_len = max(len(n) for n in news_collector.category_scores.keys())

        for cat_name, cat_data in news_collector.category_scores.items():
            cat_score = cat_data['score']
            news_count = cat_data['news_count']

            # 바 그래프 (10칸)
            if cat_score > 0:
                max_possible = 30  # 대략적 최대값
                bar_len = min(10, max(0, int(cat_score / max_possible * 10)))
            else:
                bar_len = 0
            bar = "=" * bar_len + " " * (10 - bar_len)

            # 레벨
            if cat_score >= 15:
                level_str = "높음"
            elif cat_score >= 5:
                level_str = "보통"
            elif cat_score >= 1:
                level_str = "낮음"
            else:
                level_str = "안정"

            padding = " " * (max_name_len - len(cat_name) + 2)
            print(f"  {cat_name}{padding}[{bar}]  {level_str} (뉴스 {news_count}건)")

        # Safe-Haven 자산
        print("\n  --- Safe-Haven 자산 ---")
        for ticker, stat in safe_haven.asset_stats.items():
            if not stat.get('available'):
                continue

            name = stat['name']
            current = stat.get('current')
            chg_1w = stat.get('chg_1w')

            if current is None:
                continue

            # 포맷팅
            if ticker == 'KRW=X':
                price_str = f"{current:,.0f}"
            elif ticker in ('^VIX', '^TNX'):
                price_str = f"{current:.1f}"
            else:
                price_str = f"${current:,.1f}"

            chg_str = f"{chg_1w:+.1f}% 1주" if chg_1w is not None else "N/A"

            signal = GeopoliticalExcelBuilder._get_asset_signal(ticker, stat)

            padding = " " * (14 - len(name))
            print(f"  {name}:{padding}{price_str} ({chg_str})  {signal}")

        # 섹터 영향
        print("\n  --- 섹터 영향 ---")
        for sector_name in sector_analyzer.SECTOR_STOCKS:
            exposure = sector_analyzer.sector_exposure.get(sector_name, '낮음')
            impacts = sector_analyzer.sector_impacts.get(sector_name, [])

            # 주요 영향 요약
            if impacts:
                main_impact = impacts[0]
                desc = main_impact['description']
                impact_tag = main_impact['impact']
            else:
                desc = "현재 직접적 지정학 리스크 제한적"
                impact_tag = "중립"

            padding = " " * (8 - len(sector_name))
            print(f"  {sector_name}:{padding}[{impact_tag}] {desc}")


# ====================================================================
# 6. 메인 실행
# ====================================================================
def main():
    parser = argparse.ArgumentParser(description="지정학 리스크 분석 시스템")
    parser.add_argument('--category', default=None, help='특정 카테고리만 분석 (예: 관세, 전쟁)')
    parser.add_argument('--days', type=int, default=3, help='최근 N일 뉴스 수집 (기본: 3)')
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("  지정학 리스크 분석 시스템")
    print("=" * 60)

    # 1. 뉴스 수집 및 리스크 점수 산정
    news_collector = GeopoliticalNewsCollector(
        days=args.days,
        category_filter=args.category,
    )
    news_collector.fetch_all_feeds()
    news_collector.score_news()

    # 2. 안전자산 모니터링
    safe_haven = SafeHavenMonitor()
    safe_haven.fetch_all()
    safe_haven.calculate_stats()

    # 뉴스 리스크 점수 정규화
    news_risk_normalized = news_collector.get_normalized_news_risk()

    # 리스크 온도계 계산
    risk_score, risk_level = safe_haven.calculate_risk_thermometer(news_risk_normalized)

    # 3. 섹터 영향 분석
    sector_analyzer = SectorImpactAnalyzer(news_collector.category_scores)
    sector_analyzer.analyze()

    # 4. 콘솔 출력
    ConsoleReporter.print_report(news_collector, safe_haven, sector_analyzer)

    # 5. 종합 판단 텍스트 생성
    summary_text = _generate_summary(risk_score, risk_level, news_collector, safe_haven)

    # 6. Excel 보고서 생성
    print("\n[보고서] Excel 생성 중...")
    builder = GeopoliticalExcelBuilder()

    builder.add_dashboard_sheet(
        risk_score=risk_score,
        risk_level=risk_level,
        asset_stats=safe_haven.asset_stats,
        summary_text=summary_text,
    )

    builder.add_category_sheet(
        category_scores=news_collector.category_scores,
        risk_categories=news_collector.RISK_CATEGORIES,
    )

    builder.add_news_sheet(
        scored_news=news_collector.scored_news,
    )

    builder.add_sector_sheet(
        sector_analyzer=sector_analyzer,
        category_scores=news_collector.category_scores,
    )

    today_str = datetime.now().strftime('%Y%m%d')
    filename = f"지정학리스크_{today_str}.xlsx"
    filepath = builder.save(filename)

    print(f"\n  [저장 완료] {filepath}")
    print()
    print(f"  - Excel 파일을 열어 지정학 리스크 분석 결과를 확인하세요.")
    print(f"  - financial_analyzer.py, fed_macro_analyzer.py와 함께 종합 분석에 활용하세요.")
    print()


def _generate_summary(risk_score, risk_level, news_collector, safe_haven):
    """종합 판단 텍스트 생성"""
    parts = []

    parts.append(f"현재 지정학 리스크 수준은 [{risk_level}] ({risk_score}/100)입니다.")

    # 가장 높은 리스크 카테고리
    if news_collector.category_scores:
        top_cat = max(news_collector.category_scores.items(), key=lambda x: x[1]['score'])
        if top_cat[1]['score'] > 0:
            parts.append(
                f"가장 높은 리스크 카테고리는 '{top_cat[0]}'으로 "
                f"관련 뉴스 {top_cat[1]['news_count']}건이 감지되었습니다."
            )

    # 안전자산 동향
    gold = safe_haven.asset_stats.get('GC=F', {})
    vix = safe_haven.asset_stats.get('^VIX', {})
    if gold.get('available') and gold.get('chg_1w') is not None:
        gold_chg = gold['chg_1w']
        if gold_chg > 1:
            parts.append(f"금 가격 주간 {gold_chg:+.1f}% 상승으로 안전자산 선호 심리가 강화되고 있습니다.")
        elif gold_chg < -1:
            parts.append(f"금 가격 주간 {gold_chg:+.1f}% 하락으로 리스크 선호 분위기입니다.")

    if vix.get('available') and vix.get('current') is not None:
        vix_val = vix['current']
        if vix_val > 25:
            parts.append(f"VIX {vix_val:.1f}로 시장 공포 수준이 높습니다. 방어적 포지션을 권고합니다.")
        elif vix_val < 15:
            parts.append(f"VIX {vix_val:.1f}로 시장 안정 구간입니다.")

    return ' '.join(parts)


if __name__ == "__main__":
    main()
