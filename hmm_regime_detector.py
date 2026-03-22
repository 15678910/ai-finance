"""
HMM 마켓 레짐 탐지 시스템 - Hidden Markov Model 기반
=====================================================
사용법:
  1. pip install hmmlearn yfinance pandas openpyxl numpy scikit-learn
  2. python hmm_regime_detector.py --ticker 005930  (삼성전자)
  3. python hmm_regime_detector.py --ticker AAPL    (애플, 미국주식)

생성 파일:
  - output/[종목코드]_HMM레짐분석_[날짜].xlsx
"""

import os
import sys
import argparse
import warnings
from datetime import datetime
from pathlib import Path

warnings.filterwarnings('ignore')

# -- 필수 라이브러리 설치 확인 ----------------------------------------
def check_and_install():
    required = {
        'hmmlearn': 'hmmlearn',
        'yfinance': 'yfinance',
        'pandas': 'pandas',
        'openpyxl': 'openpyxl',
        'numpy': 'numpy',
        'sklearn': 'scikit-learn',
    }
    missing = []
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[설치 필요] pip install {' '.join(missing)}")
        sys.exit(1)

check_and_install()

import numpy as np
import pandas as pd
import yfinance as yf
from hmmlearn.hmm import GaussianHMM
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- 출력 폴더 생성 ---------------------------------------------------
Path("output").mkdir(exist_ok=True)


# ====================================================================
# 1. HMM 마켓 레짐 탐지기
# ====================================================================
class MarketRegimeDetector:

    # 7개 레짐 라벨 (평균 수익률 내림차순으로 매핑)
    REGIME_LABELS = [
        "강한상승", "상승", "약한상승", "횡보",
        "약한하락", "하락", "강한하락"
    ]

    # 레짐별 투자 전략
    STRATEGIES = {
        "강한상승": "적극 매수. 포지션 확대 (80-100%). 모멘텀 추종 전략.",
        "상승": "매수 유지. 포지션 (60-80%). 눌림목 매수 전략.",
        "약한상승": "선별적 매수. 포지션 (40-60%). 우량주 중심.",
        "횡보": "관망. 포지션 축소 (20-40%). 박스권 매매.",
        "약한하락": "방어적 운용. 포지션 (10-30%). 현금 비중 확대.",
        "하락": "매도 또는 헤지. 포지션 최소화 (0-10%). 인버스 ETF 검토.",
        "강한하락": "전량 매도/숏. 포지션 0%. 안전자산 이동. 반등 시그널 대기.",
    }

    # 레짐별 포지션 비중
    POSITION_SIZES = {
        "강한상승": "80-100%",
        "상승": "60-80%",
        "약한상승": "40-60%",
        "횡보": "20-40%",
        "약한하락": "10-30%",
        "하락": "0-10%",
        "강한하락": "0%",
    }

    # 레짐별 리스크 레벨
    RISK_LEVELS = {
        "강한상승": "공격적",
        "상승": "적극적",
        "약한상승": "중립 편향 적극",
        "횡보": "중립",
        "약한하락": "보수적",
        "하락": "방어적",
        "강한하락": "극방어적",
    }

    def __init__(self, ticker: str):
        # 국내 종목은 .KS 추가
        if ticker.isdigit():
            self.ticker_raw = ticker
            # 코스닥/코스피 자동 감지
            self.ticker = self._detect_market(ticker)
            self.is_korean = True
        else:
            self.ticker_raw = ticker
            self.ticker = ticker
            self.is_korean = False

        self.stock = yf.Ticker(self.ticker)
        self.info = self.stock.info
        self.company_name = self.info.get('longName') or self.info.get('shortName', ticker)
        self.model = None
        self.scaler = StandardScaler()
        self.features_df = None
        self.state_map = {}  # HMM state -> regime label

    @staticmethod
    @staticmethod
    def _detect_market(ticker_code):
        """코스피(.KS) / 코스닥(.KQ) 자동 감지"""
        for suffix in [".KS", ".KQ"]:
            try:
                t = yf.Ticker(ticker_code + suffix)
                hist = t.history(period="5d")
                if not hist.empty and len(hist) > 0:
                    info = t.info or {}
                    name = info.get('longName') or info.get('shortName', '')
                    if name and not name.startswith(ticker_code):
                        return ticker_code + suffix
            except Exception:
                pass
        print(f"  [!] 시장 자동 감지 실패, 코스피(.KS)로 시도합니다.")
        return ticker_code + ".KS"

    def prepare_features(self) -> pd.DataFrame:
        """피처 준비: 수익률, 변동성, 평균수익률, 거래량변화율"""
        print(f"\n[분석] [{self.ticker}] HMM 레짐 분석 시작...")
        print("  ├-- 주가 데이터 수집 중...")

        df = self.stock.history(period="2y")
        if df.empty or len(df) < 30:
            print("  [오류] 충분한 주가 데이터가 없습니다.")
            sys.exit(1)

        if df.index.tz is not None:
            df.index = df.index.tz_localize(None)

        # 피처 계산
        df['Return'] = df['Close'].pct_change()
        df['Volatility_5d'] = df['Return'].rolling(window=5).std()
        df['MeanReturn_5d'] = df['Return'].rolling(window=5).mean()
        df['VolumeChange'] = df['Volume'].pct_change()

        # 결측치 제거
        df = df.dropna(subset=['Return', 'Volatility_5d', 'MeanReturn_5d', 'VolumeChange'])

        # 극단값 클리핑 (거래량 변화율)
        df['VolumeChange'] = df['VolumeChange'].clip(-5, 5)

        self.features_df = df
        print(f"  ├-- 피처 준비 완료 (데이터 {len(df)}일)")
        return df

    def fit_hmm(self):
        """HMM 모델 학습"""
        print("  ├-- HMM 모델 학습 중 (7 states, 200 iterations)...")

        feature_cols = ['Return', 'Volatility_5d', 'MeanReturn_5d', 'VolumeChange']
        X = self.features_df[feature_cols].values

        # 정규화
        X_scaled = self.scaler.fit_transform(X)

        # HMM 학습
        self.model = GaussianHMM(
            n_components=7,
            covariance_type="full",
            n_iter=200,
            random_state=42,
            verbose=False,
        )
        self.model.fit(X_scaled)

        # 상태 예측
        states = self.model.predict(X_scaled)
        self.features_df['State'] = states

        # 상태별 평균 수익률 계산 후 라벨 매핑
        state_returns = {}
        for s in range(7):
            mask = self.features_df['State'] == s
            if mask.sum() > 0:
                state_returns[s] = self.features_df.loc[mask, 'Return'].mean()
            else:
                state_returns[s] = 0.0

        # 평균 수익률 내림차순 정렬 -> 라벨 매핑
        sorted_states = sorted(state_returns.keys(), key=lambda s: state_returns[s], reverse=True)
        for rank, s in enumerate(sorted_states):
            self.state_map[s] = self.REGIME_LABELS[rank]

        # 라벨 컬럼 추가
        self.features_df['Regime'] = self.features_df['State'].map(self.state_map)

        print(f"  ├-- 모델 학습 완료 (수렴 점수: {self.model.score(X_scaled):.2f})")
        return self.model

    def get_current_regime(self) -> dict:
        """현재 레짐 정보 반환"""
        latest = self.features_df.iloc[-1]
        current_state = int(latest['State'])
        current_regime = self.state_map[current_state]

        # 현재 레짐의 통계
        regime_mask = self.features_df['Regime'] == current_regime
        regime_data = self.features_df.loc[regime_mask]

        return {
            'regime': current_regime,
            'state_id': current_state,
            'date': latest.name.strftime('%Y-%m-%d'),
            'close': latest['Close'],
            'mean_return': regime_data['Return'].mean() * 100,
            'volatility': regime_data['Volatility_5d'].mean() * 100,
            'days_in_regime': int(regime_mask.sum()),
            'pct_of_total': regime_mask.sum() / len(self.features_df) * 100,
        }

    def get_transition_probs(self) -> pd.DataFrame:
        """전환확률 매트릭스 반환 (레짐 라벨 기준)"""
        transmat = self.model.transmat_

        # HMM state 인덱스 -> 라벨 순서로 재배열
        label_to_states = {}
        for s, label in self.state_map.items():
            label_to_states[label] = s

        ordered_states = [label_to_states[label] for label in self.REGIME_LABELS]

        # 재배열된 전환확률
        reordered = np.zeros((7, 7))
        for i, si in enumerate(ordered_states):
            for j, sj in enumerate(ordered_states):
                reordered[i, j] = transmat[si, sj]

        df = pd.DataFrame(
            reordered,
            index=self.REGIME_LABELS,
            columns=self.REGIME_LABELS,
        )
        return df

    def get_next_regime_probs(self) -> dict:
        """현재 상태에서 다음 레짐으로의 전환 확률"""
        current_state = int(self.features_df.iloc[-1]['State'])
        probs = self.model.transmat_[current_state]

        result = {}
        for s in range(7):
            label = self.state_map[s]
            result[label] = probs[s] * 100

        # 라벨 순서대로 정렬
        ordered = {}
        for label in self.REGIME_LABELS:
            ordered[label] = result[label]
        return ordered

    def get_strategy(self, regime: str = None) -> str:
        """레짐별 투자 전략"""
        if regime is None:
            regime = self.get_current_regime()['regime']
        return self.STRATEGIES.get(regime, "데이터 부족")

    def get_regime_stats(self) -> list:
        """모든 레짐 통계 반환"""
        stats = []
        for label in self.REGIME_LABELS:
            mask = self.features_df['Regime'] == label
            if mask.sum() == 0:
                stats.append({
                    'regime': label,
                    'count': 0,
                    'pct': 0,
                    'mean_return': 0,
                    'volatility': 0,
                    'strategy': self.STRATEGIES[label],
                    'position': self.POSITION_SIZES[label],
                    'risk_level': self.RISK_LEVELS[label],
                })
                continue

            data = self.features_df.loc[mask]
            stats.append({
                'regime': label,
                'count': int(mask.sum()),
                'pct': mask.sum() / len(self.features_df) * 100,
                'mean_return': data['Return'].mean() * 100,
                'volatility': data['Volatility_5d'].mean() * 100,
                'strategy': self.STRATEGIES[label],
                'position': self.POSITION_SIZES[label],
                'risk_level': self.RISK_LEVELS[label],
            })
        return stats


# ====================================================================
# 2. Excel 보고서 생성
# ====================================================================
class RegimeExcelBuilder:

    # 색상 팔레트 (financial_analyzer.py 동일)
    COLORS = {
        'header_dark': '0D1B2A',
        'header_blue': '1B4F72',
        'accent': '2E86C1',
        'light_blue': 'D6EAF8',
        'white': 'FFFFFF',
        'light_gray': 'F2F3F4',
        'positive': 'EAFAF1',
        'negative': 'FDEDEC',
        'gold': 'F39C12',
    }

    # 레짐별 행 색상
    REGIME_COLORS = {
        "강한상승": '1A5276',
        "상승": '2E86C1',
        "약한상승": '85C1E9',
        "횡보": 'F7DC6F',
        "약한하락": 'F0B27A',
        "하락": 'E74C3C',
        "강한하락": '922B21',
    }

    REGIME_FONT_COLORS = {
        "강한상승": 'FFFFFF',
        "상승": 'FFFFFF',
        "약한상승": '0D1B2A',
        "횡보": '0D1B2A',
        "약한하락": '0D1B2A',
        "하락": 'FFFFFF',
        "강한하락": 'FFFFFF',
    }

    def __init__(self, ticker: str, company_name: str):
        self.wb = Workbook()
        self.ticker = ticker
        self.company_name = company_name
        self.wb.remove(self.wb.active)

    def _style_header(self, ws, row, col, value, color='header_dark', font_color='FFFFFF', bold=True, size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=self.COLORS.get(color, color))
        cell.font = Font(bold=bold, color=font_color, size=size, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def _style_data(self, ws, row, col, value, bg_color=None, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if number_format:
            cell.number_format = number_format
        return cell

    def _prob_color(self, prob):
        """전환확률 값에 따른 배경색 (높을수록 진한 파랑)"""
        if prob >= 0.5:
            return '1A5276'
        elif prob >= 0.3:
            return '2E86C1'
        elif prob >= 0.2:
            return '5DADE2'
        elif prob >= 0.1:
            return '85C1E9'
        elif prob >= 0.05:
            return 'AED6F1'
        else:
            return 'D6EAF8'

    def _prob_font_color(self, prob):
        """전환확률 값에 따른 글자색"""
        if prob >= 0.3:
            return 'FFFFFF'
        return '0D1B2A'

    def add_regime_summary_sheet(self, current_regime: dict, regime_stats: list, strategy: str):
        """시트 1: 레짐 요약"""
        ws = self.wb.create_sheet("레짐 요약")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = f"{self.company_name} ({self.ticker}) -- HMM 레짐 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 날짜
        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 현재 레짐 정보
        ws.merge_cells('B5:G5')
        self._style_header(ws, 5, 2, "현재 시장 레짐", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        regime_color = self.REGIME_COLORS.get(current_regime['regime'], 'accent')
        regime_font = self.REGIME_FONT_COLORS.get(current_regime['regime'], 'FFFFFF')

        info_items = [
            ('분석 기준일', current_regime['date']),
            ('현재 종가', f"{current_regime['close']:,.0f}"),
            ('현재 레짐', current_regime['regime']),
            ('평균 일간 수익률', f"{current_regime['mean_return']:.4f}%"),
            ('평균 변동성', f"{current_regime['volatility']:.4f}%"),
            ('레짐 일수', f"{current_regime['days_in_regime']}일 ({current_regime['pct_of_total']:.1f}%)"),
            ('투자 전략', strategy),
        ]

        row = 6
        for key, value in info_items:
            ws.cell(row=row, column=2, value=key).font = Font(bold=True, size=10, name='맑은 고딕')
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)

            ws.merge_cells(f'C{row}:G{row}')
            data_cell = ws.cell(row=row, column=3, value=value)
            data_cell.font = Font(size=10, name='맑은 고딕')
            data_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            if key == '현재 레짐':
                data_cell.fill = PatternFill("solid", fgColor=regime_color)
                data_cell.font = Font(size=11, name='맑은 고딕', bold=True, color=regime_font)
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
            elif key == '투자 전략':
                data_cell.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
                data_cell.font = Font(size=10, name='맑은 고딕', bold=True, color='1B4F72')

            bg = 'light_gray' if row % 2 == 0 else 'white'
            if key not in ('현재 레짐', '투자 전략'):
                data_cell.fill = PatternFill("solid", fgColor=self.COLORS[bg])

            ws.row_dimensions[row].height = 24
            row += 1

        # 전체 레짐 통계 요약
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "전체 레짐 통계", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        stat_headers = ['레짐', '일수', '비율(%)', '평균수익률(%)', '변동성(%)', '리스크']
        for col, h in enumerate(stat_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        for stat in regime_stats:
            rc = self.REGIME_COLORS.get(stat['regime'], 'F2F3F4')
            rf = self.REGIME_FONT_COLORS.get(stat['regime'], '0D1B2A')

            cell = ws.cell(row=row, column=2, value=stat['regime'])
            cell.fill = PatternFill("solid", fgColor=rc)
            cell.font = Font(size=10, name='맑은 고딕', bold=True, color=rf)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            self._style_data(ws, row, 3, stat['count'])
            self._style_data(ws, row, 4, round(stat['pct'], 1), number_format='0.0')
            self._style_data(ws, row, 5, round(stat['mean_return'], 4), number_format='0.0000')
            self._style_data(ws, row, 6, round(stat['volatility'], 4), number_format='0.0000')
            self._style_data(ws, row, 7, stat['risk_level'])
            ws.row_dimensions[row].height = 22
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18
        for c in 'CDEFG':
            ws.column_dimensions[c].width = 18

    def add_transition_matrix_sheet(self, trans_df: pd.DataFrame):
        """시트 2: 전환확률 매트릭스"""
        ws = self.wb.create_sheet("전환확률 매트릭스")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:I2')
        title = ws['B2']
        title.value = "레짐 전환확률 매트릭스 (행: 현재 상태 -> 열: 다음 상태)"
        title.font = Font(bold=True, size=12, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 30

        labels = list(trans_df.index)

        # 좌상단 빈 셀
        self._style_header(ws, 4, 2, "현재 \\ 다음", 'header_dark', size=10)

        # 열 헤더
        for j, label in enumerate(labels):
            rc = self.REGIME_COLORS.get(label, 'header_blue')
            rf = self.REGIME_FONT_COLORS.get(label, 'FFFFFF')
            cell = ws.cell(row=4, column=3 + j, value=label)
            cell.fill = PatternFill("solid", fgColor=rc)
            cell.font = Font(bold=True, size=10, name='맑은 고딕', color=rf)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[4].height = 28

        # 데이터
        for i, from_label in enumerate(labels):
            # 행 헤더
            rc = self.REGIME_COLORS.get(from_label, 'header_blue')
            rf = self.REGIME_FONT_COLORS.get(from_label, 'FFFFFF')
            cell = ws.cell(row=5 + i, column=2, value=from_label)
            cell.fill = PatternFill("solid", fgColor=rc)
            cell.font = Font(bold=True, size=10, name='맑은 고딕', color=rf)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            for j, to_label in enumerate(labels):
                prob = trans_df.loc[from_label, to_label]
                bg = self._prob_color(prob)
                fc = self._prob_font_color(prob)
                cell = ws.cell(row=5 + i, column=3 + j, value=round(prob, 4))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(size=10, name='맑은 고딕', color=fc)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.00%'

            ws.row_dimensions[5 + i].height = 24

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 14
        for j in range(7):
            ws.column_dimensions[get_column_letter(3 + j)].width = 12

        # 범례
        legend_row = 13
        ws.merge_cells(f'B{legend_row}:I{legend_row}')
        ws.cell(row=legend_row, column=2, value="* 색상이 진할수록 전환 확률이 높음").font = Font(
            size=9, color='7F8C8D', name='맑은 고딕')

    def add_daily_regime_sheet(self, features_df: pd.DataFrame):
        """시트 3: 일별 레짐"""
        ws = self.wb.create_sheet("일별 레짐")
        ws.sheet_view.showGridLines = False

        headers = ['날짜', '종가', '수익률(%)', '변동성(%)', '레짐']
        for col, h in enumerate(headers, 1):
            self._style_header(ws, 1, col, h, 'header_dark')
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.row_dimensions[1].height = 28

        # 최근 252일
        recent = features_df.tail(252).copy()

        for row_idx, (date, row_data) in enumerate(recent.iterrows(), 2):
            regime = row_data['Regime']
            rc = self.REGIME_COLORS.get(regime, 'F2F3F4')
            rf = self.REGIME_FONT_COLORS.get(regime, '0D1B2A')

            # 날짜
            cell = ws.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d'))
            cell.font = Font(size=10, name='맑은 고딕')
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 종가
            self._style_data(ws, row_idx, 2, round(row_data['Close'], 0), number_format='#,##0')

            # 수익률
            ret_val = round(row_data['Return'] * 100, 4)
            ret_bg = 'positive' if ret_val > 0 else ('negative' if ret_val < 0 else 'white')
            self._style_data(ws, row_idx, 3, ret_val, bg_color=ret_bg, number_format='0.0000')

            # 변동성
            vol_val = round(row_data['Volatility_5d'] * 100, 4)
            self._style_data(ws, row_idx, 4, vol_val, number_format='0.0000')

            # 레짐 (색상)
            regime_cell = ws.cell(row=row_idx, column=5, value=regime)
            regime_cell.fill = PatternFill("solid", fgColor=rc)
            regime_cell.font = Font(size=10, name='맑은 고딕', bold=True, color=rf)
            regime_cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[row_idx].height = 18

    def add_strategy_sheet(self, regime_stats: list):
        """시트 4: 투자전략"""
        ws = self.wb.create_sheet("투자전략")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "레짐별 투자 전략 가이드"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 헤더
        headers = ['레짐', '설명', '포지션 비중', '투자 전략', '리스크 레벨']
        for col, h in enumerate(headers, 2):
            self._style_header(ws, 4, col, h, 'header_dark', size=10)
        ws.row_dimensions[4].height = 28

        descriptions = {
            "강한상승": "시장이 강한 상승 모멘텀을 보이는 국면. 높은 수익률과 비교적 안정적 변동성.",
            "상승": "완만한 상승 추세. 양호한 시장 심리와 매수세 유입.",
            "약한상승": "미약한 상승세. 제한적 상승폭, 선별적 접근 필요.",
            "횡보": "방향성 없는 박스권. 변동성 낮고 매매 기회 제한적.",
            "약한하락": "초기 하락 신호. 방어적 전환 시점.",
            "하락": "뚜렷한 하락 추세. 매도 압력 우세, 리스크 관리 필수.",
            "강한하락": "급격한 하락. 패닉 매도 가능성, 자본 보전 최우선.",
        }

        row = 5
        for stat in regime_stats:
            regime = stat['regime']
            rc = self.REGIME_COLORS.get(regime, 'F2F3F4')
            rf = self.REGIME_FONT_COLORS.get(regime, '0D1B2A')

            # 레짐
            cell = ws.cell(row=row, column=2, value=regime)
            cell.fill = PatternFill("solid", fgColor=rc)
            cell.font = Font(size=10, name='맑은 고딕', bold=True, color=rf)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 설명
            desc_cell = ws.cell(row=row, column=3, value=descriptions.get(regime, ''))
            desc_cell.font = Font(size=10, name='맑은 고딕')
            desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

            # 포지션 비중
            self._style_data(ws, row, 4, stat['position'])

            # 전략
            strat_cell = ws.cell(row=row, column=5, value=stat['strategy'])
            strat_cell.font = Font(size=10, name='맑은 고딕')
            strat_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

            # 리스크
            self._style_data(ws, row, 6, stat['risk_level'])

            ws.row_dimensions[row].height = 40
            row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 14

    def save(self, filepath: str):
        self.wb.save(filepath)
        print(f"\n  [저장 완료] {filepath}")


# ====================================================================
# 3. 콘솔 출력
# ====================================================================
def print_console_report(detector: MarketRegimeDetector):
    """콘솔에 현재 레짐 및 전환확률 출력"""
    current = detector.get_current_regime()
    next_probs = detector.get_next_regime_probs()
    strategy = detector.get_strategy()

    print("\n" + "=" * 60)
    print(f"  HMM 마켓 레짐 분석 결과 - {detector.company_name}")
    print("=" * 60)

    print(f"\n  기준일: {current['date']}")
    print(f"  종가: {current['close']:,.0f}")
    print(f"  현재 레짐: [{current['regime']}]")
    print(f"  평균 수익률: {current['mean_return']:.4f}%")
    print(f"  평균 변동성: {current['volatility']:.4f}%")
    print(f"  레짐 일수: {current['days_in_regime']}일 ({current['pct_of_total']:.1f}%)")

    print(f"\n  투자 전략: {strategy}")

    print(f"\n  --- 다음 레짐 전환 확률 ---")
    for label, prob in next_probs.items():
        bar = "#" * int(prob / 2)
        print(f"  {label:6s} : {prob:6.2f}% {bar}")

    print("\n" + "=" * 60)


# ====================================================================
# 4. 메인 실행
# ====================================================================
def main():
    parser = argparse.ArgumentParser(description="HMM 마켓 레짐 탐지 시스템")
    parser.add_argument('--ticker', required=True, help="종목 코드 (예: 005930, AAPL)")
    args = parser.parse_args()

    ticker = args.ticker

    # 1. 데이터 준비 및 HMM 학습
    detector = MarketRegimeDetector(ticker)
    detector.prepare_features()
    detector.fit_hmm()

    # 2. 분석 결과
    current = detector.get_current_regime()
    trans_df = detector.get_transition_probs()
    next_probs = detector.get_next_regime_probs()
    strategy = detector.get_strategy()
    regime_stats = detector.get_regime_stats()

    # 3. 콘솔 출력
    print_console_report(detector)

    # 4. Excel 저장
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"output/{ticker}_HMM레짐분석_{date_str}.xlsx"

    builder = RegimeExcelBuilder(ticker, detector.company_name)
    builder.add_regime_summary_sheet(current, regime_stats, strategy)
    builder.add_transition_matrix_sheet(trans_df)
    builder.add_daily_regime_sheet(detector.features_df)
    builder.add_strategy_sheet(regime_stats)
    builder.save(filename)

    print(f"\n  다음 단계: output 폴더의 Excel 파일을 확인하세요.")


if __name__ == "__main__":
    main()
