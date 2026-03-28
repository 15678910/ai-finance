"""
generate_dashboard_data.py
일일 분석 결과를 읽어 웹 대시보드용 docs/data.json을 생성합니다.

Usage:
    python generate_dashboard_data.py
    python generate_dashboard_data.py --date 20260322
    python generate_dashboard_data.py --date 20260322 --daily-dir output/daily/20260322
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from glob import glob

# ---------------------------------------------------------------------------
# 헤드라인 번역 (키워드 기반)
# ---------------------------------------------------------------------------

def translate_headline(title: str) -> str:
    """영어 뉴스 제목을 한글로 번역 (키워드 기반)"""
    if not title:
        return title

    # 구문 단위 치환 (긴 것부터)
    PHRASE_MAP = {
        "North Korea": "북한",
        "South Korea": "한국",
        "United States": "미국",
        "Saudi Arabia": "사우디아라비아",
        "ceasefire in war": "전쟁 휴전",
        "state media reports": "국영매체 보도",
        "state media": "국영매체",
        "oil prices": "유가",
        "Oil giants": "석유 대기업들",
        "energy shortages": "에너지 부족",
        "trade war": "무역전쟁",
        "nuclear weapons": "핵무기",
        "drone strike": "드론 공격",
        "drone attack": "드론 공격",
        "peace deal": "평화 협정",
        "peace talks": "평화 협상",
        "15-point plan": "15개항 계획",
        "end war": "전쟁 종결",
        "to end": "종결",
        "won't accept": "거부",
        "have been overturned": "전복되었다",
        "has received": "수령했다",
        "raises the alarm": "경고 울려",
        "raise the alarm": "경고 울려",
        "drags on": "장기화",
        "largest attack": "최대규모 공격",
        "24-hour period": "24시간 동안",
        "over 24-ho": "24시간 동안",
        "international conflicts": "국제 분쟁",
        "could accelerate": "가속화할 수 있다",
        "a shift into": "전환을",
        "President says": "대통령 발언",
        "report says": "보도에 따르면",
    }

    # 단어 단위 치환
    WORD_MAP = {
        "Iran": "이란",
        "Iranian": "이란의",
        "Russia": "러시아",
        "Russian": "러시아의",
        "Ukraine": "우크라이나",
        "U.S.": "미국",
        "US": "미국",
        "China": "중국",
        "Chinese": "중국의",
        "Japan": "일본",
        "Israel": "이스라엘",
        "Gaza": "가자",
        "Hamas": "하마스",
        "Taiwan": "대만",
        "NATO": "나토",
        "Trump": "트럼프",
        "Biden": "바이든",
        "Asia": "아시아",
        "Europe": "유럽",
        "Korea": "한국",
        "Syria": "시리아",
        "war": "전쟁",
        "attack": "공격",
        "attacks": "공격",
        "missile": "미사일",
        "missiles": "미사일",
        "drone": "드론",
        "drones": "드론",
        "launches": "발사",
        "ceasefire": "휴전",
        "sanctions": "제재",
        "tariff": "관세",
        "tariffs": "관세",
        "military": "군사",
        "conflict": "분쟁",
        "defense": "방어",
        "weapons": "무기",
        "nuclear": "핵",
        "oil": "석유",
        "trade": "무역",
        "threatens": "위협",
        "invasion": "침공",
        "bomb": "폭탄",
        "effort": "노력",
        "plan": "계획",
        "norms": "규범",
        "shows": "보여준다",
        "moment": "국면",
        "shift": "전환",
        "renewable": "재생에너지",
        "accelerate": "가속화",
        "President": "대통령",
        "injure": "부상",
        "kill": "사망",
        "killed": "사망",
        "troops": "병력",
        "soldiers": "군인",
        "bombing": "폭격",
        "power": "전력/권력",
        "plant": "시설",
        "fighting": "전투",
        "illegally": "불법적으로",
        "towns": "마을",
        "prices": "가격",
        "stabilize": "안정화",
        "investors": "투자자",
        "weigh": "저울질",
        "How": "어떻게",
        "the": "",
        "The": "",
        "a": "",
        "an": "",
        "of": "",
        "in": "",
        "at": "",
        "on": "",
        "to": "",
        "for": "",
        "by": "",
        "as": "",
        "is": "",
        "are": "",
        "was": "",
        "were": "",
        "has": "",
        "have": "",
        "had": "",
        "been": "",
        "being": "",
        "with": "",
        "from": "",
        "that": "",
        "this": "",
        "and": "",
        "or": "",
        "but": "",
        "not": "",
        "its": "",
        "their": "",
        "over": "",
        "into": "",
        "near": "인근",
    }

    import re
    result = title

    # 1단계: 구문 치환 (긴 것부터)
    for eng, kor in sorted(PHRASE_MAP.items(), key=lambda x: -len(x[0])):
        result = result.replace(eng, kor)

    # 2단계: 단어 치환 (정확한 단어 경계에서만)
    for eng, kor in sorted(WORD_MAP.items(), key=lambda x: -len(x[0])):
        result = re.sub(r'\b' + re.escape(eng) + r'\b', kor, result)

    # 3단계: 다중 공백 정리
    result = re.sub(r'\s+', ' ', result).strip()

    return result


# ---------------------------------------------------------------------------
# 상수 / 기본값
# ---------------------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

FALLBACK_NAME_TO_TICKER = {
    "삼성전자": "005930",
    "SK하이닉스": "000660",
    "네이버": "035420",
    "카카오": "035720",
    "한화에어로스페이스": "012450",
    "LIG넥스원": "079550",
    "현대로템": "064350",
    "LG에너지솔루션": "373220",
    "삼성SDI": "006400",
    "에코프로비엠": "247540",
    "비트코인": "BTC-USD",
    "이더리움": "ETH-USD",
    "리플": "XRP-USD",
}

# sectors.json의 섹터 키 → 실제 서브디렉터리 이름 매핑
SECTOR_DIR_MAP = {
    "IT/반도체": "IT_반도체",
    "에너지": "에너지",
    "방산": "방산",
    "배터리/2차전지": "배터리_2차전지",
    "바이오/헬스케어": "바이오_헬스케어",
    "미국 빅테크": "미국_빅테크",
    "암호화폐": "암호화폐",
}

# ---------------------------------------------------------------------------
# 헬퍼: sectors.json 로드 → name_to_ticker 딕셔너리 생성
# ---------------------------------------------------------------------------

def load_name_to_ticker(config_path: str) -> dict:
    """sectors.json을 읽어 {한국어이름: ticker} 딕셔너리를 반환합니다."""
    mapping = {}
    try:
        with open(config_path, encoding="utf-8") as f:
            data = json.load(f)

        sectors_data = data.get("sectors", data)

        for sector_key, sector_val in sectors_data.items():
            tickers = sector_val.get("tickers", {})
            for ticker, kor_name in tickers.items():
                mapping[kor_name] = ticker

        print(f"[INFO] sectors.json 로드 완료: {len(mapping)}개 종목 매핑")
    except FileNotFoundError:
        print(f"[WARN] sectors.json 없음: {config_path} → fallback 매핑 사용")
    except Exception as e:
        print(f"[WARN] sectors.json 로드 실패 ({e}) → fallback 매핑 사용")

    if not mapping:
        mapping = FALLBACK_NAME_TO_TICKER.copy()

    return mapping


# ---------------------------------------------------------------------------
# 헬퍼: 종합요약 텍스트 파싱
# ---------------------------------------------------------------------------

def parse_summary_txt(filepath: str) -> dict:
    """종합요약_{date}.txt를 파싱해 매크로 정보와 섹터별 종목 데이터를 반환합니다."""
    result = {
        "date": "",
        "generated_at": "",
        "macro": {},
        "sectors": [],
    }

    try:
        with open(filepath, encoding="utf-8") as f:
            text = f.read()
    except FileNotFoundError:
        print(f"[ERROR] 종합요약 파일 없음: {filepath}")
        return result
    except Exception as e:
        print(f"[ERROR] 종합요약 파일 읽기 실패: {e}")
        return result

    lines = text.splitlines()

    # --- 날짜 파싱 (첫 줄) ---
    date_match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", lines[0] if lines else "")
    if date_match:
        result["date"] = date_match.group(1)

    # --- 생성 시각 파싱 ---
    gen_match = re.search(r"\[생성 시각\]\s*(.+)", text)
    if gen_match:
        result["generated_at"] = gen_match.group(1).strip()

    # --- 매크로 환경 파싱 ---
    macro = {}
    cycle_m = re.search(r"경기 사이클:\s*(.+)", text)
    rate_m = re.search(r"금리:\s*([^\s(]+)(?:\s*\(FFR\s*([\d.]+%)\))?", text)
    infl_m = re.search(r"인플레이션:\s*([^\s(]+)(?:\s*\(CPI\s*([\d.]+%)\))?", text)

    if cycle_m:
        macro["cycle"] = cycle_m.group(1).strip()
    if rate_m:
        macro["rate"] = rate_m.group(1).strip()
        macro["ffr"] = rate_m.group(2).strip() if rate_m.group(2) else "N/A"
    if infl_m:
        macro["inflation"] = infl_m.group(1).strip()
        macro["cpi"] = infl_m.group(2).strip() if infl_m.group(2) else "N/A"

    result["macro"] = macro

    # --- 섹터별 요약 파싱 ---
    sectors = []
    current_sector = None

    in_sector_block = False
    for line in lines:
        stripped = line.strip()

        if stripped == "[섹터별 요약]":
            in_sector_block = True
            continue

        if not in_sector_block:
            continue

        if stripped.startswith("[생성 시각]"):
            break

        if not stripped:
            continue

        if re.match(r"^[^\s].+:$", stripped):
            sector_name = stripped.rstrip(":")
            current_sector = {"name": sector_name, "stocks": []}
            sectors.append(current_sector)
            continue

        stock_m = re.match(
            r"^(.+?)\s*-\s*레짐:\s*(.+?)\s*/\s*심리:\s*(.+?)\(([+-]?[\d.]+)\)\s*$",
            stripped,
        )
        if stock_m and current_sector is not None:
            name = stock_m.group(1).strip()
            regime = stock_m.group(2).strip()
            sentiment_label = stock_m.group(3).strip()
            score_raw = stock_m.group(4).strip()
            score_str = score_raw if score_raw.startswith(("+", "-")) else f"+{score_raw}"
            current_sector["stocks"].append(
                {
                    "name": name,
                    "regime": regime,
                    "sentiment_label": sentiment_label,
                    "sentiment_score": score_str,
                }
            )

    result["sectors"] = sectors
    return result


# ---------------------------------------------------------------------------
# 헬퍼: metrics.json 파싱
# ---------------------------------------------------------------------------

def load_metrics(metrics_path: str) -> dict:
    """티커_metrics.json 파일을 읽어 필요한 필드만 추출합니다."""
    defaults = {
        "price": None,
        "market_cap": None,
        "per": "N/A",
        "forward_per": None,
        "roe": None,
        "revenue_growth": None,
        "high_52w": None,
        "low_52w": None,
    }
    try:
        with open(metrics_path, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"[WARN] metrics.json 읽기 실패 ({metrics_path}): {e}")
        return defaults

    def _val(key, fallback=None):
        v = data.get(key, fallback)
        return v if v not in (None, "", "N/A") else fallback

    return {
        "price": _val("현재주가"),
        "market_cap": _val("시가총액(억)"),
        "per": data.get("PER", "N/A"),
        "forward_per": _val("Forward PER"),
        "roe": _val("ROE(%)"),
        "revenue_growth": _val("매출성장률(%)"),
        "high_52w": _val("52주최고"),
        "low_52w": _val("52주최저"),
    }


# ---------------------------------------------------------------------------
# 헬퍼: daily_dir에서 티커→metrics 경로 인덱스 빌드
# ---------------------------------------------------------------------------

def build_ticker_metrics_index(daily_dir: str) -> dict:
    """
    daily_dir 하위 모든 서브디렉터리를 스캔해 {ticker: metrics_path} 딕셔너리를 반환합니다.
    파일명 패턴: {ticker}_metrics.json
    """
    index = {}
    try:
        for entry in os.scandir(daily_dir):
            if not entry.is_dir():
                continue
            subdir = entry.path
            pattern = os.path.join(subdir, "*_metrics.json")
            for fpath in glob(pattern):
                fname = os.path.basename(fpath)
                ticker = fname.replace("_metrics.json", "")
                index[ticker] = fpath
    except Exception as e:
        print(f"[WARN] metrics 인덱스 빌드 실패: {e}")
    print(f"[INFO] metrics 인덱스: {len(index)}개 티커 발견")
    return index


# ---------------------------------------------------------------------------
# 헬퍼: Excel 셀에서 숫자 추출
# ---------------------------------------------------------------------------

def _to_float(val, default=None):
    """셀 값을 float으로 변환. 문자열에서 숫자 부분만 추출."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        # "3.64%", "+2.40", "-0.10%p", "31442.48십억$" 등에서 숫자 추출
        cleaned = val.replace(",", "").replace("%", "").replace("p", "").replace("십억$", "").replace("$", "")
        cleaned = cleaned.strip()
        try:
            return float(cleaned)
        except ValueError:
            return default
    return default


def _to_str(val, default="N/A"):
    """셀 값을 문자열로 변환."""
    if val is None:
        return default
    return str(val).strip()


def _find_cell_by_label(ws, label, search_col=2, value_col=3, max_row=50):
    """시트에서 특정 라벨을 찾아 인접 셀 값을 반환합니다."""
    for r in range(1, min(ws.max_row + 1, max_row + 1)):
        cell_val = ws.cell(r, search_col).value
        if cell_val and str(cell_val).strip() == label:
            return ws.cell(r, value_col).value
    return None


def _find_row_by_label(ws, label, search_col=2, max_row=50):
    """시트에서 특정 라벨이 있는 행 번호를 반환합니다."""
    for r in range(1, min(ws.max_row + 1, max_row + 1)):
        cell_val = ws.cell(r, search_col).value
        if cell_val and label in str(cell_val).strip():
            return r
    return None


# ---------------------------------------------------------------------------
# 매크로 상세 데이터 추출 (macro_detail)
# ---------------------------------------------------------------------------

def extract_macro_detail(daily_dir: str, date_str: str) -> dict:
    """매크로분석 Excel에서 상세 매크로 데이터를 추출합니다."""
    result = {
        "ffr": {"current": None, "3m_ago": None, "6m_ago": None, "1y_ago": None, "trend": "N/A"},
        "unemployment": {"current": None, "change_3m": None, "trend": "N/A"},
        "gdp": {"current": None, "change_3m": None, "trend": "N/A"},
        "vix": {"current": None, "change_3m": None, "trend": "N/A"},
        "yield_curve": {"spread_10y2y": None, "status": "N/A", "trend": "N/A"},
        "breakeven_inflation_10y": None,
        "cpi_yoy": None,
        "real_rate": None,
        "inflation_status": "N/A",
        "inflation_risk": "N/A",
        "rate_cycle": {"stage": "N/A", "description": "N/A"},
        "economic_cycle": {"stage": "N/A", "score": None, "description": "N/A"},
        "key_signals": [],
        "dollar_index": {"current": None, "trend": "N/A"},
        "treasury_10y": {"current": None, "trend": "N/A"},
        "treasury_2y": {"current": None, "trend": "N/A"},
        "asset_outlook": [],
        "fomc": {},
    }

    macro_path = os.path.join(daily_dir, "macro", f"매크로분석_{date_str}.xlsx")
    if not os.path.exists(macro_path):
        print(f"[WARN] 매크로분석 파일 없음: {macro_path}")
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(macro_path, data_only=True)
    except Exception as e:
        print(f"[WARN] 매크로분석 Excel 로드 실패: {e}")
        return result

    # --- Sheet 1: 매크로 대시보드 ---
    try:
        ws = wb["매크로 대시보드"]
        # Row 7~16: 지표 데이터 (B=지표명, C=현재값, D=3개월전, E=6개월전, F=1년전, G=3M변화, H=추세)
        indicators = {}
        for r in range(7, ws.max_row + 1):
            label = _to_str(ws.cell(r, 2).value, "")
            if not label:
                continue
            indicators[label] = {
                "current": ws.cell(r, 3).value,
                "3m": ws.cell(r, 4).value,
                "6m": ws.cell(r, 5).value,
                "1y": ws.cell(r, 6).value,
                "change": _to_str(ws.cell(r, 7).value, ""),
                "trend": _to_str(ws.cell(r, 8).value, "N/A"),
            }

        if "기준금리(FFR)" in indicators:
            d = indicators["기준금리(FFR)"]
            result["ffr"] = {
                "current": _to_float(d["current"]),
                "3m_ago": _to_float(d["3m"]),
                "6m_ago": _to_float(d["6m"]),
                "1y_ago": _to_float(d["1y"]),
                "trend": d["trend"],
            }

        if "실업률" in indicators:
            d = indicators["실업률"]
            result["unemployment"] = {
                "current": _to_float(d["current"]),
                "change_3m": _to_str(d["change"]),
                "trend": d["trend"],
            }

        if "실질GDP" in indicators:
            d = indicators["실질GDP"]
            result["gdp"] = {
                "current": _to_float(d["current"]),
                "change_3m": _to_str(d["change"]),
                "trend": d["trend"],
            }

        if "VIX지수" in indicators:
            d = indicators["VIX지수"]
            result["vix"] = {
                "current": _to_float(d["current"]),
                "change_3m": _to_str(d["change"]),
                "trend": d["trend"],
            }

        if "장단기금리차(10Y-2Y)" in indicators:
            d = indicators["장단기금리차(10Y-2Y)"]
            result["yield_curve"]["spread_10y2y"] = _to_float(d["current"])
            result["yield_curve"]["trend"] = d["trend"]

        if "기대인플레이션(10Y)" in indicators:
            d = indicators["기대인플레이션(10Y)"]
            result["breakeven_inflation_10y"] = _to_float(d["current"])

        if "10년국채금리" in indicators:
            d = indicators["10년국채금리"]
            result["treasury_10y"] = {
                "current": _to_float(d["current"]),
                "trend": d["trend"],
            }

        if "2년국채금리" in indicators:
            d = indicators["2년국채금리"]
            result["treasury_2y"] = {
                "current": _to_float(d["current"]),
                "trend": d["trend"],
            }

        if "달러인덱스(TWI)" in indicators:
            d = indicators["달러인덱스(TWI)"]
            result["dollar_index"] = {
                "current": _to_float(d["current"]),
                "trend": d["trend"],
            }
    except Exception as e:
        print(f"[WARN] 매크로 대시보드 시트 파싱 실패: {e}")

    # --- Sheet 2: 금리-인플레이션 ---
    try:
        ws2 = wb["금리-인플레이션"]
        rate_cycle_stage = _find_cell_by_label(ws2, "금리 사이클 단계")
        rate_cycle_desc = _find_cell_by_label(ws2, "사이클 설명")
        yield_status = _find_cell_by_label(ws2, "수익률 곡선 상태")
        cpi_yoy = _find_cell_by_label(ws2, "CPI 전년비 변화율 (YoY)")
        real_rate = _find_cell_by_label(ws2, "실질금리 (FFR - 기대인플레)")
        inflation_status = _find_cell_by_label(ws2, "인플레이션 상태")
        inflation_risk = _find_cell_by_label(ws2, "인플레이션 리스크")

        result["rate_cycle"] = {
            "stage": _to_str(rate_cycle_stage),
            "description": _to_str(rate_cycle_desc),
        }
        result["yield_curve"]["status"] = _to_str(yield_status)
        result["cpi_yoy"] = _to_str(cpi_yoy)
        result["real_rate"] = _to_str(real_rate)
        result["inflation_status"] = _to_str(inflation_status)
        result["inflation_risk"] = _to_str(inflation_risk)
    except Exception as e:
        print(f"[WARN] 금리-인플레이션 시트 파싱 실패: {e}")

    # --- Sheet 3: 경기 사이클 ---
    try:
        ws3 = wb["경기 사이클"]
        cycle_stage = _find_cell_by_label(ws3, "현재 경기 단계")
        cycle_score = _find_cell_by_label(ws3, "종합 점수")
        cycle_desc = _find_cell_by_label(ws3, "단계 설명")

        result["economic_cycle"] = {
            "stage": _to_str(cycle_stage),
            "score": _to_str(cycle_score),
            "description": _to_str(cycle_desc),
        }

        # 핵심 경기 신호 (rows after 신호/해석 header)
        signals = []
        # Find the exact header row where B="신호" and C="해석"
        signal_header_row = None
        for r in range(1, ws3.max_row + 1):
            b_val = ws3.cell(r, 2).value
            c_val = ws3.cell(r, 3).value
            if b_val and str(b_val).strip() == "신호" and c_val and "해석" in str(c_val):
                signal_header_row = r
                break
        if signal_header_row:
            for r in range(signal_header_row + 1, min(ws3.max_row + 1, signal_header_row + 20)):
                sig = ws3.cell(r, 2).value
                interp = ws3.cell(r, 3).value
                direction = ws3.cell(r, 4).value
                if not sig:
                    break
                signals.append({
                    "signal": _to_str(sig),
                    "interpretation": _to_str(interp),
                    "direction": _to_str(direction, "neutral"),
                })
        result["key_signals"] = signals
    except Exception as e:
        print(f"[WARN] 경기 사이클 시트 파싱 실패: {e}")

    # --- Sheet 4: 자산별 전망 ---
    try:
        ws4 = wb["자산별 전망"]
        outlooks = []
        # Find the exact header row where B="자산 클래스" (not "자산 클래스별 전망")
        header_row = None
        for r in range(1, ws4.max_row + 1):
            b_val = ws4.cell(r, 2).value
            if b_val and str(b_val).strip() == "자산 클래스":
                header_row = r
                break
        if header_row:
            for r in range(header_row + 1, ws4.max_row + 1):
                asset = ws4.cell(r, 2).value
                if not asset or asset.startswith("*"):
                    break
                outlooks.append({
                    "asset": _to_str(asset),
                    "outlook": _to_str(ws4.cell(r, 3).value),
                    "strategy": _to_str(ws4.cell(r, 4).value),
                    "risk": _to_str(ws4.cell(r, 5).value),
                })
        result["asset_outlook"] = outlooks
    except Exception as e:
        print(f"[WARN] 자산별 전망 시트 파싱 실패: {e}")

    # --- Sheet 5: FOMC 분석 ---
    try:
        ws5 = wb["FOMC 분석"]

        # 현재 금리 정보 (label-based lookup)
        target_upper_raw = _find_cell_by_label(ws5, "FF목표금리 상단 (DFEDTARU)", max_row=60)
        target_lower_raw = _find_cell_by_label(ws5, "FF목표금리 하단 (DFEDTARL)", max_row=60)
        effective_raw = _find_cell_by_label(ws5, "실효 연방기금금리 (FEDFUNDS)", max_row=60)
        market_exp_raw = _find_cell_by_label(ws5, "시장 금리 기대", max_row=60)

        # 다음 회의 정보
        next_meeting_raw = _find_cell_by_label(ws5, "다음 FOMC 결정일", max_row=60)
        dday_raw = _find_cell_by_label(ws5, "D-day (남은 일수)", max_row=60)

        # 시장 기대: "동결 (Hold)" → "hold", "인하 기대 (Cut Expected)" → "cut", "인상 기대 (Hike Expected)" → "hike"
        market_exp_str = _to_str(market_exp_raw, "")
        if "Cut" in market_exp_str or "인하" in market_exp_str:
            market_expectation = "cut"
        elif "Hike" in market_exp_str or "인상" in market_exp_str:
            market_expectation = "hike"
        else:
            market_expectation = "hold"

        # D-day 숫자 추출: "D-7" → 7
        days_until = None
        dday_str = _to_str(dday_raw, "")
        dday_match = re.search(r"D-(\d+)", dday_str)
        if dday_match:
            days_until = int(dday_match.group(1))

        # 최근 금리 결정 이력: 헤더 행("결정월") 찾아서 그 다음 행부터 파싱
        recent_decisions = []
        decision_header_row = None
        for r in range(1, ws5.max_row + 1):
            b_val = ws5.cell(r, 2).value
            if b_val and str(b_val).strip() == "결정월":
                decision_header_row = r
                break
        if decision_header_row:
            for r in range(decision_header_row + 1, min(ws5.max_row + 1, decision_header_row + 20)):
                date_val = ws5.cell(r, 2).value
                if not date_val:
                    break
                rate_val = ws5.cell(r, 3).value
                change_val = ws5.cell(r, 4).value
                action_val = ws5.cell(r, 5).value
                recent_decisions.append({
                    "date": _to_str(date_val),
                    "rate": _to_str(rate_val),
                    "change": _to_str(change_val),
                    "direction": _to_str(action_val),
                })

        result["fomc"] = {
            "next_meeting": _to_str(next_meeting_raw),
            "days_until": days_until,
            "target_upper": _to_float(target_upper_raw),
            "target_lower": _to_float(target_lower_raw),
            "effective_rate": _to_float(effective_raw),
            "market_expectation": market_expectation,
            "recent_decisions": recent_decisions,
        }
    except Exception as e:
        print(f"[WARN] FOMC 분석 시트 파싱 실패: {e}")

    wb.close()
    print(f"[INFO] 매크로 상세 데이터 추출 완료")
    return result


# ---------------------------------------------------------------------------
# 지정학 리스크 데이터 추출 (geopolitical)
# ---------------------------------------------------------------------------

def extract_geopolitical(daily_dir: str, date_str: str) -> dict:
    """지정학리스크 Excel에서 리스크 데이터를 추출합니다."""
    result = {
        "risk_score": None,
        "risk_level": "N/A",
        "summary": "",
        "categories": [],
        "safe_haven_assets": [],
        "top_news": [],
    }

    geo_path = os.path.join(daily_dir, "macro", f"지정학리스크_{date_str}.xlsx")
    if not os.path.exists(geo_path):
        print(f"[WARN] 지정학리스크 파일 없음: {geo_path}")
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(geo_path, data_only=True)
    except Exception as e:
        print(f"[WARN] 지정학리스크 Excel 로드 실패: {e}")
        return result

    # --- Sheet 1: 리스크 대시보드 ---
    try:
        ws = wb["리스크 대시보드"]

        # Row 6: [주의]  57.1/100 - parse score and level
        thermometer = _to_str(ws.cell(6, 2).value, "")
        level_m = re.match(r"\[(.+?)\]\s+([\d.]+)/100", thermometer)
        if level_m:
            result["risk_level"] = level_m.group(1)
            result["risk_score"] = _to_float(level_m.group(2))

        # Safe-Haven assets: rows 11-17 (B=자산, C=현재가, D=1주변화, E=1월변화, F=시그널)
        assets = []
        for r in range(11, ws.max_row + 1):
            asset_name = ws.cell(r, 2).value
            if not asset_name:
                break
            price = ws.cell(r, 3).value
            week_change = _to_str(ws.cell(r, 4).value, "")
            month_change = _to_str(ws.cell(r, 5).value, "")
            signal = _to_str(ws.cell(r, 6).value, "")
            assets.append({
                "name": _to_str(asset_name),
                "current_value": _to_float(price),
                "week_change": week_change,
                "month_change": month_change,
                "signal": signal,
            })
        result["safe_haven_assets"] = assets

        # Row 20: 종합 판단 text
        summary_text = _to_str(ws.cell(20, 2).value, "")
        result["summary"] = summary_text
    except Exception as e:
        print(f"[WARN] 리스크 대시보드 시트 파싱 실패: {e}")

    # --- Sheet 2: 리스크 카테고리 분석 ---
    try:
        ws2 = wb["리스크 카테고리 분석"]
        categories = []
        # Rows 6-11: B=카테고리, C=리스크점수, D=관련뉴스, E=핵심뉴스, F=영향섹터
        for r in range(6, ws2.max_row + 1):
            cat_name = ws2.cell(r, 2).value
            if not cat_name:
                break
            score = _to_float(ws2.cell(r, 3).value, 0)
            news_count = _to_str(ws2.cell(r, 4).value, "0건")
            key_news = _to_str(ws2.cell(r, 5).value, "")
            affected_sectors = _to_str(ws2.cell(r, 6).value, "")
            categories.append({
                "category": _to_str(cat_name),
                "score": score,
                "news_count": news_count,
                "key_news": key_news,
                "affected_sectors": affected_sectors,
            })
        result["categories"] = categories
    except Exception as e:
        print(f"[WARN] 리스크 카테고리 시트 파싱 실패: {e}")

    # --- Sheet 3: 주요 뉴스 (top 5) ---
    try:
        ws3 = wb["주요 뉴스"]
        news = []
        for r in range(6, min(ws3.max_row + 1, 11)):  # top 5
            date_val = ws3.cell(r, 2).value
            title = ws3.cell(r, 3).value
            if not title:
                break
            source = _to_str(ws3.cell(r, 4).value, "")
            category = _to_str(ws3.cell(r, 5).value, "")
            risk_score = _to_float(ws3.cell(r, 6).value, 0)
            # col 7: 한글제목, col 8: 링크 (신규 컬럼 - 구버전 파일 호환)
            title_kr_val = ws3.cell(r, 7).value
            link_val = ws3.cell(r, 8).value
            title_str = _to_str(title)
            title_kr = _to_str(title_kr_val, "") if title_kr_val else translate_headline(title_str)
            news.append({
                "date": _to_str(date_val),
                "title": title_str,
                "title_kr": title_kr,
                "link": _to_str(link_val, "") if link_val else "",
                "source": source,
                "category": category,
                "risk_score": risk_score,
            })
        result["top_news"] = news
    except Exception as e:
        print(f"[WARN] 주요 뉴스 시트 파싱 실패: {e}")

    wb.close()
    print(f"[INFO] 지정학 리스크 데이터 추출 완료")
    return result


# ---------------------------------------------------------------------------
# 포트폴리오 데이터 추출 (portfolios)
# ---------------------------------------------------------------------------

def extract_portfolios(daily_dir: str, date_str: str) -> dict:
    """포트폴리오 Excel 파일들에서 섹터별 포트폴리오 데이터를 추출합니다."""
    result = {
        "sectors": [],
        "best_sector": None,
        "worst_sector": None,
    }

    portfolio_dir = os.path.join(daily_dir, "포트폴리오")
    if not os.path.isdir(portfolio_dir):
        print(f"[WARN] 포트폴리오 디렉터리 없음: {portfolio_dir}")
        return result

    try:
        import openpyxl
    except ImportError:
        print(f"[WARN] openpyxl 미설치")
        return result

    portfolio_files = glob(os.path.join(portfolio_dir, f"*_포트폴리오_{date_str}.xlsx"))
    if not portfolio_files:
        print(f"[WARN] 포트폴리오 파일 없음: {portfolio_dir}")
        return result

    best_sharpe = -999
    worst_sharpe = 999
    best_name = None
    worst_name = None

    for pf_path in sorted(portfolio_files):
        fname = os.path.basename(pf_path)
        # Extract sector name: "IT_반도체_포트폴리오_20260323.xlsx" → "IT_반도체"
        sector_name = fname.replace(f"_포트폴리오_{date_str}.xlsx", "")

        try:
            wb = openpyxl.load_workbook(pf_path, data_only=True)
        except Exception as e:
            print(f"[WARN] 포트폴리오 Excel 로드 실패 ({fname}): {e}")
            continue

        sector_data = {
            "name": sector_name,
            "sharpe_ratio": None,
            "annual_return": None,
            "annual_volatility": None,
            "max_drawdown": None,
            "cumulative_return": None,
            "beta": None,
            "var_95": None,
            "var_99": None,
            "cvar_95": None,
            "allocations": {
                "aggressive": {},
                "balanced": {},
                "stable": {},
            },
            "assets": [],
        }

        # --- Sheet: 포트폴리오 요약 ---
        try:
            ws = wb["포트폴리오 요약"]

            # Portfolio performance metrics
            sector_data["annual_return"] = _to_str(_find_cell_by_label(ws, "연간 수익률"))
            sector_data["annual_volatility"] = _to_str(_find_cell_by_label(ws, "연간 변동성"))

            sharpe_val = _find_cell_by_label(ws, "Sharpe Ratio")
            sector_data["sharpe_ratio"] = _to_float(sharpe_val)

            sector_data["cumulative_return"] = _to_str(_find_cell_by_label(ws, "누적 수익률"))
            sector_data["max_drawdown"] = _to_str(_find_cell_by_label(ws, "최대 낙폭(MDD)"))

            # Risk metrics
            sector_data["var_95"] = _to_str(_find_cell_by_label(ws, "VaR(95%)"))
            sector_data["var_99"] = _to_str(_find_cell_by_label(ws, "VaR(99%)"))
            sector_data["cvar_95"] = _to_str(_find_cell_by_label(ws, "CVaR(95%)"))
            sector_data["beta"] = _to_float(_find_cell_by_label(ws, "Beta"))

            # Asset list (rows 7+ until empty)
            assets = []
            for r in range(7, ws.max_row + 1):
                asset_name = ws.cell(r, 2).value
                if not asset_name or asset_name in ("포트폴리오 성과", "포트폴리오 비교", "핵심 리스크 지표"):
                    if asset_name and asset_name != ws.cell(7, 2).value:
                        break
                    if not asset_name:
                        break
                ticker = _to_str(ws.cell(r, 3).value, "")
                weight = _to_str(ws.cell(r, 4).value, "")
                if ticker and "." in ticker:  # looks like a ticker
                    assets.append({
                        "name": _to_str(asset_name),
                        "ticker": ticker,
                        "weight": weight,
                    })
            sector_data["assets"] = assets
        except Exception as e:
            print(f"[WARN] 포트폴리오 요약 시트 파싱 실패 ({sector_name}): {e}")

        # --- Sheet: 최적화 결과 ---
        try:
            ws_opt = wb["최적화 결과"]
            # Row 6: header with asset names in C6, D6, E6, ...
            # Row 7: 현재 weights
            # Row 8: 공격형 (aggressive)
            # Row 9: 균형형 (balanced)
            # Row 10: 안정형 (stable)

            # Get asset column headers
            asset_cols = {}
            for c in range(3, ws_opt.max_column + 1):
                header = ws_opt.cell(6, c).value
                if header and header not in ("수익률", "변동성", "Sharpe", "VaR(95%)"):
                    asset_cols[c] = _to_str(header)
                elif header in ("수익률", "변동성", "Sharpe", "VaR(95%)"):
                    break

            type_map = {
                8: "aggressive",
                9: "balanced",
                10: "stable",
            }

            for row_num, alloc_type in type_map.items():
                type_label = ws_opt.cell(row_num, 2).value
                if not type_label:
                    continue
                weights = {}
                for c, asset_name in asset_cols.items():
                    w = _to_str(ws_opt.cell(row_num, c).value, "0%")
                    weights[asset_name] = w

                # Find the return/vol/sharpe columns (after asset columns)
                # They are at fixed positions relative to asset count
                meta_start = max(asset_cols.keys()) + 1 if asset_cols else 6
                alloc_return = _to_str(ws_opt.cell(row_num, meta_start).value, "")
                alloc_vol = _to_str(ws_opt.cell(row_num, meta_start + 1).value, "")
                alloc_sharpe = _to_float(ws_opt.cell(row_num, meta_start + 2).value)
                alloc_var = _to_str(ws_opt.cell(row_num, meta_start + 3).value, "")

                sector_data["allocations"][alloc_type] = {
                    "weights": weights,
                    "return": alloc_return,
                    "volatility": alloc_vol,
                    "sharpe": alloc_sharpe,
                    "var_95": alloc_var,
                }
        except Exception as e:
            print(f"[WARN] 최적화 결과 시트 파싱 실패 ({sector_name}): {e}")

        wb.close()

        # Track best/worst Sharpe
        if sector_data["sharpe_ratio"] is not None:
            if sector_data["sharpe_ratio"] > best_sharpe:
                best_sharpe = sector_data["sharpe_ratio"]
                best_name = sector_name
            if sector_data["sharpe_ratio"] < worst_sharpe:
                worst_sharpe = sector_data["sharpe_ratio"]
                worst_name = sector_name

        result["sectors"].append(sector_data)
        print(f"[INFO]   포트폴리오 추출: {sector_name} (Sharpe: {sector_data['sharpe_ratio']})")

    result["best_sector"] = {
        "name": best_name,
        "sharpe_ratio": best_sharpe if best_name else None,
    }
    result["worst_sector"] = {
        "name": worst_name,
        "sharpe_ratio": worst_sharpe if worst_name else None,
    }

    print(f"[INFO] 포트폴리오 데이터 추출 완료: {len(result['sectors'])}개 섹터")
    return result


# ---------------------------------------------------------------------------
# 인사이트 생성 (insights)
# ---------------------------------------------------------------------------

def generate_insights(macro: dict, macro_detail: dict, geopolitical: dict, portfolios: dict, sectors: list) -> dict:
    """모든 데이터를 종합해 인사이트를 생성합니다."""
    result = {
        "market_assessment": "",
        "top_insights": [],
        "key_risks": [],
        "best_opportunities": [],
    }

    # --- Market Assessment ---
    cycle = macro_detail.get("economic_cycle", {}).get("stage", "N/A")
    rate_stage = macro_detail.get("rate_cycle", {}).get("stage", "N/A")
    vix_val = macro_detail.get("vix", {}).get("current")
    geo_score = geopolitical.get("risk_score")
    geo_level = geopolitical.get("risk_level", "N/A")

    vix_desc = ""
    if vix_val is not None:
        if vix_val < 15:
            vix_desc = "낮은 변동성"
        elif vix_val < 20:
            vix_desc = "안정적 시장"
        elif vix_val < 30:
            vix_desc = "경계 수준의 변동성"
        else:
            vix_desc = "높은 공포 수준"

    assessment_parts = []
    if cycle != "N/A":
        assessment_parts.append(f"경기 {cycle}")
    if rate_stage != "N/A":
        assessment_parts.append(f"금리 {rate_stage}")
    if vix_desc:
        assessment_parts.append(f"VIX {vix_val:.1f} ({vix_desc})")
    if geo_score is not None:
        assessment_parts.append(f"지정학 리스크 {geo_level}({geo_score:.1f}/100)")

    if assessment_parts:
        result["market_assessment"] = f"현재 시장: {', '.join(assessment_parts)}."
    else:
        result["market_assessment"] = "시장 데이터 부족."

    # --- Top Insights ---
    insights = []

    # Insight from rate cycle
    if rate_stage != "N/A":
        desc = macro_detail.get("rate_cycle", {}).get("description", "")
        if desc and desc != "N/A":
            insights.append(f"금리 {rate_stage}: {desc}")

    # Insight from yield curve
    yc_status = macro_detail.get("yield_curve", {}).get("status", "N/A")
    spread = macro_detail.get("yield_curve", {}).get("spread_10y2y")
    if spread is not None and yc_status != "N/A":
        insights.append(f"장단기 금리차 {spread:.2f}%p - {yc_status}")

    # Insight from best portfolio sector
    best = portfolios.get("best_sector", {})
    if best and best.get("name"):
        insights.append(f"최고 Sharpe 섹터: {best['name']} ({best['sharpe_ratio']:.2f})")

    # Insight from geopolitical
    if geo_score is not None and geo_score >= 50:
        top_cat = None
        top_score = 0
        for cat in geopolitical.get("categories", []):
            if cat.get("score", 0) > top_score:
                top_score = cat["score"]
                top_cat = cat.get("category", "")
        if top_cat:
            insights.append(f"지정학 주의: {top_cat} 리스크 점수 {top_score}")

    result["top_insights"] = insights[:3]

    # --- Key Risks ---
    risks = []

    if vix_val is not None and vix_val >= 20:
        risks.append(f"VIX {vix_val:.1f} - 시장 변동성 경계 수준")

    if geo_score is not None and geo_score >= 40:
        risks.append(f"지정학 리스크 {geo_level} ({geo_score:.1f}/100)")

    inflation_risk = macro_detail.get("inflation_risk", "N/A")
    if inflation_risk not in ("N/A", "낮음"):
        infl_status = macro_detail.get("inflation_status", "")
        risks.append(f"인플레이션 리스크: {inflation_risk} ({infl_status})")

    # Unemployment
    unemp = macro_detail.get("unemployment", {}).get("current")
    if unemp is not None and unemp >= 4.5:
        risks.append(f"실업률 {unemp}% - 고용 시장 약화 우려")

    result["key_risks"] = risks[:3]

    # --- Best Opportunities ---
    opportunities = []

    # From asset outlook
    for outlook in macro_detail.get("asset_outlook", []):
        if outlook.get("outlook") == "긍정":
            opportunities.append({
                "asset": outlook["asset"],
                "reason": outlook.get("strategy", ""),
            })

    # From portfolio best sectors (top 3 by Sharpe)
    pf_sectors = sorted(
        [s for s in portfolios.get("sectors", []) if s.get("sharpe_ratio") is not None],
        key=lambda x: x["sharpe_ratio"],
        reverse=True,
    )
    for s in pf_sectors[:3]:
        opportunities.append({
            "asset": s["name"],
            "reason": f"Sharpe {s['sharpe_ratio']:.2f}, 연간수익률 {s.get('annual_return', 'N/A')}",
        })

    result["best_opportunities"] = opportunities[:5]

    return result


# ---------------------------------------------------------------------------
# 메인 로직
# ---------------------------------------------------------------------------

def generate(date_str: str, daily_dir: str, output_path: str) -> bool:
    """데이터를 수집해 output_path에 JSON을 저장합니다."""

    print(f"[INFO] 날짜: {date_str}")
    print(f"[INFO] 일일 디렉터리: {daily_dir}")
    print(f"[INFO] 출력 경로: {output_path}")

    # 1) name → ticker 매핑 로드
    config_path = os.path.join(SCRIPT_DIR, "config", "sectors.json")
    name_to_ticker = load_name_to_ticker(config_path)

    # 2) 종합요약 파싱
    summary_filename = f"종합요약_{date_str}.txt"
    summary_path = os.path.join(daily_dir, summary_filename)
    print(f"[INFO] 종합요약 파싱 중: {summary_path}")
    parsed = parse_summary_txt(summary_path)

    if not parsed["date"]:
        parsed["date"] = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
    if not parsed["generated_at"]:
        parsed["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 3) metrics 인덱스 빌드
    ticker_to_metrics_path = build_ticker_metrics_index(daily_dir)

    # 4) 종목별 metrics 병합
    for sector in parsed["sectors"]:
        enriched_stocks = []
        for stock in sector["stocks"]:
            kor_name = stock["name"]
            ticker = name_to_ticker.get(kor_name)

            stock_entry = {
                "name": kor_name,
                "ticker": ticker or "UNKNOWN",
                "regime": stock["regime"],
                "sentiment_label": stock["sentiment_label"],
                "sentiment_score": stock["sentiment_score"],
                "price": None,
                "market_cap": None,
                "per": "N/A",
                "forward_per": None,
                "roe": None,
                "revenue_growth": None,
                "high_52w": None,
                "low_52w": None,
            }

            if ticker and ticker in ticker_to_metrics_path:
                metrics = load_metrics(ticker_to_metrics_path[ticker])
                stock_entry.update(metrics)
                print(f"[INFO]   {kor_name} ({ticker}): metrics 로드 완료")
            else:
                if ticker:
                    print(f"[WARN]   {kor_name} ({ticker}): metrics.json 없음")
                else:
                    print(f"[WARN]   {kor_name}: ticker 매핑 없음")

            enriched_stocks.append(stock_entry)
        sector["stocks"] = enriched_stocks

    # 5) 매크로 상세 데이터 추출
    print(f"[INFO] 매크로 상세 데이터 추출 중...")
    macro_detail = extract_macro_detail(daily_dir, date_str)

    # 6) 지정학 리스크 데이터 추출
    print(f"[INFO] 지정학 리스크 데이터 추출 중...")
    geopolitical = extract_geopolitical(daily_dir, date_str)

    # 7) 포트폴리오 데이터 추출
    print(f"[INFO] 포트폴리오 데이터 추출 중...")
    portfolios = extract_portfolios(daily_dir, date_str)

    # 8) 인사이트 생성
    print(f"[INFO] 인사이트 생성 중...")
    insights = generate_insights(
        parsed["macro"], macro_detail, geopolitical, portfolios, parsed["sectors"]
    )

    # 9) 최종 JSON 구조 조립
    output_data = {
        "date": parsed["date"],
        "generated_at": parsed["generated_at"],
        "macro": parsed["macro"],
        "macro_detail": macro_detail,
        "geopolitical": geopolitical,
        "portfolios": portfolios,
        "insights": insights,
        "sectors": parsed["sectors"],
    }

    # 10) docs/ 디렉터리 생성 및 파일 저장
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        print(f"[OK] docs/data.json 생성 완료: {output_path}")
        return True
    except Exception as e:
        print(f"[ERROR] JSON 저장 실패: {e}")
        return False


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    today = datetime.now().strftime("%Y%m%d")
    parser = argparse.ArgumentParser(
        description="일일 분석 결과를 읽어 웹 대시보드용 docs/data.json을 생성합니다."
    )
    parser.add_argument(
        "--date",
        default=today,
        help=f"분석 날짜 (YYYYMMDD, 기본값: {today})",
    )
    parser.add_argument(
        "--daily-dir",
        default=None,
        help="일일 출력 디렉터리 경로 (기본값: output/daily/{date})",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    date_str = args.date

    if args.daily_dir:
        daily_dir = args.daily_dir
    else:
        daily_dir = os.path.join(SCRIPT_DIR, "output", "daily", date_str)

    output_path = os.path.join(SCRIPT_DIR, "docs", "data.json")

    success = generate(date_str, daily_dir, output_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
