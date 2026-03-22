"""
뉴스 및 시장 심리 분석 시스템 - 무료 데이터 기반
===================================================
사용법:
  1. pip install yfinance pandas numpy openpyxl
  2. python news_sentiment_analyzer.py --ticker 005930  (삼성전자)
  3. python news_sentiment_analyzer.py --ticker AAPL    (애플, 미국주식)
  4. python news_sentiment_analyzer.py --ticker BTC-USD  (비트코인)

생성 파일:
  - output/[종목코드]_심리분석_[날짜].xlsx
"""

import os
import sys
import argparse
import warnings
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings('ignore')

# -- 필수 라이브러리 설치 확인 ----------------------------------------
def check_and_install():
    required = {
        'yfinance': 'yfinance',
        'pandas': 'pandas',
        'numpy': 'numpy',
        'openpyxl': 'openpyxl',
    }
    missing = []
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[설치 필요] pip install {' '.join(missing)}")
        sys.exit(1)

check_and_install()

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -- 출력 폴더 생성 ---------------------------------------------------
Path("output").mkdir(exist_ok=True)


# ====================================================================
# 1. 뉴스 감성 수집기
# ====================================================================
class NewsSentimentCollector:

    # 긍정 키워드 (한국어 + 영어)
    POSITIVE_KEYWORDS = [
        # 한국어
        '상승', '급등', '호재', '실적개선', '매수', '신고가', '호실적',
        '흑자', '성장', '반등', '돌파', '최고', '강세', '수혜',
        '증가', '확대', '개선', '기대', '낙관', '추천',
        # 영어
        'surge', 'rally', 'beat', 'upgrade', 'bullish', 'growth',
        'record', 'gain', 'rise', 'jump', 'soar', 'outperform',
        'buy', 'positive', 'strong', 'boost', 'profit', 'recovery',
        'breakout', 'upside', 'optimistic', 'expand', 'exceeded',
    ]

    # 부정 키워드 (한국어 + 영어)
    NEGATIVE_KEYWORDS = [
        # 한국어
        '하락', '급락', '악재', '실적부진', '매도', '신저가', '적자',
        '감소', '축소', '위기', '우려', '경고', '약세', '리스크',
        '손실', '부진', '둔화', '하향', '비관', '폭락',
        # 영어
        'plunge', 'crash', 'miss', 'downgrade', 'bearish', 'decline',
        'risk', 'warning', 'fall', 'drop', 'sell', 'loss', 'weak',
        'cut', 'negative', 'concern', 'fear', 'recession', 'default',
        'downside', 'slump', 'layoff', 'lawsuit', 'investigation',
    ]

    def __init__(self, ticker: str, is_crypto: bool = False):
        self.ticker = ticker
        self.is_crypto = is_crypto
        self.news_data = []

    def fetch_news(self) -> list:
        """yfinance에서 뉴스 헤드라인 수집"""
        print("  |- 뉴스 헤드라인 수집 중...")
        try:
            stock = yf.Ticker(self.ticker)
            news_list = stock.news
            if not news_list:
                print("  |  [참고] 뉴스 데이터 없음 - 다른 지표로 분석 진행")
                return []

            results = []
            for item in news_list:
                title = item.get('title', '')
                # 발행일 처리
                pub_ts = item.get('providerPublishTime', 0)
                if pub_ts:
                    pub_date = datetime.fromtimestamp(pub_ts).strftime('%Y-%m-%d')
                else:
                    pub_date = 'N/A'

                # 감성 분석
                score, pos_kw, neg_kw = self._score_headline(title)

                results.append({
                    'date': pub_date,
                    'title': title,
                    'score': score,
                    'positive_keywords': pos_kw,
                    'negative_keywords': neg_kw,
                    'publisher': item.get('publisher', 'N/A'),
                })

            self.news_data = results
            print(f"  |  수집 완료: {len(results)}건")
            return results

        except Exception as e:
            print(f"  |  [경고] 뉴스 수집 실패: {e}")
            return []

    def _score_headline(self, title: str) -> tuple:
        """헤드라인 감성 점수 계산"""
        title_lower = title.lower()
        pos_found = [kw for kw in self.POSITIVE_KEYWORDS if kw.lower() in title_lower]
        neg_found = [kw for kw in self.NEGATIVE_KEYWORDS if kw.lower() in title_lower]

        raw_score = len(pos_found) - len(neg_found)
        return raw_score, pos_found, neg_found

    def get_aggregate_score(self) -> dict:
        """뉴스 종합 감성 점수 (-100 ~ +100)"""
        if not self.news_data:
            return {
                'score': 0.0,
                'label': '데이터 없음',
                'total_news': 0,
                'positive_count': 0,
                'negative_count': 0,
                'neutral_count': 0,
            }

        scores = [n['score'] for n in self.news_data]
        total = len(scores)
        positive_count = sum(1 for s in scores if s > 0)
        negative_count = sum(1 for s in scores if s < 0)
        neutral_count = sum(1 for s in scores if s == 0)

        # 평균 점수를 -100 ~ +100으로 정규화
        avg_score = np.mean(scores)
        # 최대 가능 점수 기준으로 정규화 (클리핑)
        max_possible = 3.0  # 보통 헤드라인에서 3개 이상 키워드 매칭은 드묾
        normalized = np.clip(avg_score / max_possible * 100, -100, 100)

        if normalized > 30:
            label = '긍정적 뉴스 우세'
        elif normalized > 10:
            label = '약간 긍정적'
        elif normalized > -10:
            label = '중립적'
        elif normalized > -30:
            label = '약간 부정적'
        else:
            label = '부정적 뉴스 우세'

        return {
            'score': round(float(normalized), 1),
            'label': label,
            'total_news': total,
            'positive_count': positive_count,
            'negative_count': negative_count,
            'neutral_count': neutral_count,
        }


# ====================================================================
# 2. 기술적 심리 분석기
# ====================================================================
class TechnicalSentimentAnalyzer:

    def __init__(self, ticker: str):
        self.ticker = ticker
        self.df = None
        self.signals = {}

    def fetch_price_data(self) -> pd.DataFrame:
        """6개월 일봉 데이터 수집"""
        print("  |- 주가 데이터 수집 중 (6개월)...")
        stock = yf.Ticker(self.ticker)
        df = stock.history(period="6mo")
        if df.empty or len(df) < 30:
            print("  |  [오류] 충분한 주가 데이터가 없습니다.")
            return pd.DataFrame()

        if df.index.tz is not None:
            df.index = df.index.tz_localize(None)
        self.df = df
        print(f"  |  수집 완료: {len(df)}일")
        return df

    def calculate_rsi(self, period: int = 14) -> dict:
        """RSI(14) 계산"""
        if self.df is None or len(self.df) < period + 1:
            return {'value': None, 'signal': '데이터 부족', 'score': 0}

        delta = self.df['Close'].diff()
        gain = delta.clip(lower=0)
        loss = (-delta.clip(upper=0))

        avg_gain = gain.rolling(window=period, min_periods=period).mean()
        avg_loss = loss.rolling(window=period, min_periods=period).mean()

        rs = avg_gain / avg_loss.replace(0, np.nan)
        rsi = 100 - (100 / (1 + rs))
        current_rsi = rsi.iloc[-1]

        if np.isnan(current_rsi):
            return {'value': None, 'signal': '계산 불가', 'score': 0}

        # 심리 점수 매핑
        if current_rsi >= 80:
            score, signal = -2, '극도 과매수'
        elif current_rsi >= 70:
            score, signal = -1, '과매수'
        elif current_rsi >= 55:
            score, signal = 1, '강세 영역'
        elif current_rsi >= 45:
            score, signal = 0, '중립'
        elif current_rsi >= 30:
            score, signal = -1, '약세 영역'
        elif current_rsi >= 20:
            score, signal = 1, '과매도 (반등 기대)'
        else:
            score, signal = 2, '극도 과매도 (강한 반등 기대)'

        result = {'value': round(current_rsi, 1), 'signal': signal, 'score': score}
        self.signals['RSI'] = result
        return result

    def calculate_macd(self) -> dict:
        """MACD 계산 (12, 26, 9)"""
        if self.df is None or len(self.df) < 35:
            return {'value': None, 'signal_line': None, 'signal': '데이터 부족', 'score': 0}

        close = self.df['Close']
        ema12 = close.ewm(span=12, adjust=False).mean()
        ema26 = close.ewm(span=26, adjust=False).mean()
        macd_line = ema12 - ema26
        signal_line = macd_line.ewm(span=9, adjust=False).mean()
        histogram = macd_line - signal_line

        current_macd = macd_line.iloc[-1]
        current_signal = signal_line.iloc[-1]
        current_hist = histogram.iloc[-1]
        prev_hist = histogram.iloc[-2] if len(histogram) >= 2 else 0

        # 시그널 판단
        if current_hist > 0 and prev_hist <= 0:
            score, signal = 2, '골든크로스 (매수 시그널)'
        elif current_hist < 0 and prev_hist >= 0:
            score, signal = -2, '데드크로스 (매도 시그널)'
        elif current_hist > 0 and current_hist > prev_hist:
            score, signal = 1, '상승 모멘텀 강화'
        elif current_hist > 0 and current_hist <= prev_hist:
            score, signal = 0, '상승 모멘텀 둔화'
        elif current_hist < 0 and current_hist < prev_hist:
            score, signal = -1, '하락 모멘텀 강화'
        else:
            score, signal = 0, '하락 모멘텀 둔화'

        result = {
            'value': round(current_macd, 4),
            'signal_line': round(current_signal, 4),
            'histogram': round(current_hist, 4),
            'signal': signal,
            'score': score,
        }
        self.signals['MACD'] = result
        return result

    def calculate_bollinger(self, period: int = 20, std_mult: float = 2.0) -> dict:
        """볼린저 밴드 (20, 2)"""
        if self.df is None or len(self.df) < period:
            return {'value': None, 'signal': '데이터 부족', 'score': 0}

        close = self.df['Close']
        sma = close.rolling(window=period).mean()
        std = close.rolling(window=period).std()
        upper = sma + std_mult * std
        lower = sma - std_mult * std

        current_price = close.iloc[-1]
        current_upper = upper.iloc[-1]
        current_lower = lower.iloc[-1]
        current_sma = sma.iloc[-1]

        band_width = current_upper - current_lower
        if band_width == 0:
            position_pct = 50.0
        else:
            position_pct = (current_price - current_lower) / band_width * 100

        # 심리 점수
        if position_pct >= 100:
            score, signal = -2, '상단밴드 돌파 (과매수)'
        elif position_pct >= 80:
            score, signal = -1, '상단밴드 근접'
        elif position_pct >= 55:
            score, signal = 1, '중심선 위 (강세)'
        elif position_pct >= 45:
            score, signal = 0, '중심선 부근 (중립)'
        elif position_pct >= 20:
            score, signal = -1, '중심선 아래 (약세)'
        elif position_pct >= 0:
            score, signal = 1, '하단밴드 근접 (반등 기대)'
        else:
            score, signal = 2, '하단밴드 이탈 (강한 반등 기대)'

        result = {
            'value': round(position_pct, 1),
            'upper': round(current_upper, 2),
            'lower': round(current_lower, 2),
            'middle': round(current_sma, 2),
            'signal': signal,
            'score': score,
        }
        self.signals['Bollinger'] = result
        return result

    def calculate_volume_trend(self) -> dict:
        """거래량 추세 (20일 평균 vs 5일 평균)"""
        if self.df is None or len(self.df) < 20:
            return {'value': None, 'signal': '데이터 부족', 'score': 0}

        vol = self.df['Volume']
        avg_20 = vol.tail(20).mean()
        avg_5 = vol.tail(5).mean()

        if avg_20 == 0:
            ratio = 1.0
        else:
            ratio = avg_5 / avg_20

        # 가격 방향과 결합
        price_change_5d = (self.df['Close'].iloc[-1] / self.df['Close'].iloc[-5] - 1) * 100

        if ratio > 1.5 and price_change_5d > 0:
            score, signal = 2, '거래량 급증 + 상승 (강한 매수세)'
        elif ratio > 1.2 and price_change_5d > 0:
            score, signal = 1, '거래량 증가 + 상승 (매수세 유입)'
        elif ratio > 1.5 and price_change_5d < 0:
            score, signal = -2, '거래량 급증 + 하락 (강한 매도세)'
        elif ratio > 1.2 and price_change_5d < 0:
            score, signal = -1, '거래량 증가 + 하락 (매도세 유입)'
        elif ratio < 0.7:
            score, signal = 0, '거래량 감소 (관망세)'
        else:
            score, signal = 0, '거래량 보통'

        result = {
            'value': round(ratio, 2),
            'avg_5d': int(avg_5),
            'avg_20d': int(avg_20),
            'signal': signal,
            'score': score,
        }
        self.signals['Volume'] = result
        return result

    def calculate_ma_alignment(self) -> dict:
        """이동평균 배열 (5MA, 20MA, 60MA, 120MA)"""
        if self.df is None or len(self.df) < 120:
            # 데이터 부족시 가능한 범위로 계산
            if self.df is not None and len(self.df) >= 20:
                close = self.df['Close']
                ma5 = close.rolling(5).mean().iloc[-1]
                ma20 = close.rolling(20).mean().iloc[-1]
                if ma5 > ma20:
                    return {'value': '단기 정배열', 'signal': '단기 강세', 'score': 1,
                            'ma5': round(ma5, 2), 'ma20': round(ma20, 2),
                            'ma60': None, 'ma120': None}
                else:
                    return {'value': '단기 역배열', 'signal': '단기 약세', 'score': -1,
                            'ma5': round(ma5, 2), 'ma20': round(ma20, 2),
                            'ma60': None, 'ma120': None}
            return {'value': None, 'signal': '데이터 부족', 'score': 0,
                    'ma5': None, 'ma20': None, 'ma60': None, 'ma120': None}

        close = self.df['Close']
        ma5 = close.rolling(5).mean().iloc[-1]
        ma20 = close.rolling(20).mean().iloc[-1]
        ma60 = close.rolling(60).mean().iloc[-1]
        ma120 = close.rolling(120).mean().iloc[-1]

        # 정배열: 5 > 20 > 60 > 120
        # 역배열: 5 < 20 < 60 < 120
        if ma5 > ma20 > ma60 > ma120:
            score, signal, value = 2, '완전 정배열 (강한 상승 추세)', '정배열'
        elif ma5 > ma20 > ma60:
            score, signal, value = 1, '부분 정배열 (상승 추세)', '부분 정배열'
        elif ma5 < ma20 < ma60 < ma120:
            score, signal, value = -2, '완전 역배열 (강한 하락 추세)', '역배열'
        elif ma5 < ma20 < ma60:
            score, signal, value = -1, '부분 역배열 (하락 추세)', '부분 역배열'
        else:
            score, signal, value = 0, '혼조세 (방향성 없음)', '혼조'

        result = {
            'value': value,
            'signal': signal,
            'score': score,
            'ma5': round(ma5, 2),
            'ma20': round(ma20, 2),
            'ma60': round(ma60, 2),
            'ma120': round(ma120, 2),
        }
        self.signals['MA_Alignment'] = result
        return result

    def calculate_momentum(self) -> dict:
        """가격 모멘텀 (1주, 1개월, 3개월 수익률)"""
        if self.df is None or len(self.df) < 5:
            return {'value': None, 'signal': '데이터 부족', 'score': 0,
                    'return_1w': None, 'return_1m': None, 'return_3m': None}

        close = self.df['Close']
        current = close.iloc[-1]

        # 1주 (5거래일)
        return_1w = (current / close.iloc[-5] - 1) * 100 if len(close) >= 5 else 0
        # 1개월 (약 21거래일)
        return_1m = (current / close.iloc[-21] - 1) * 100 if len(close) >= 21 else 0
        # 3개월 (약 63거래일)
        return_3m = (current / close.iloc[-63] - 1) * 100 if len(close) >= 63 else 0

        # 종합 모멘텀 점수
        avg_momentum = (return_1w * 0.4 + return_1m * 0.35 + return_3m * 0.25)

        if avg_momentum > 10:
            score, signal = 2, '강한 상승 모멘텀'
        elif avg_momentum > 3:
            score, signal = 1, '상승 모멘텀'
        elif avg_momentum > -3:
            score, signal = 0, '모멘텀 중립'
        elif avg_momentum > -10:
            score, signal = -1, '하락 모멘텀'
        else:
            score, signal = -2, '강한 하락 모멘텀'

        result = {
            'value': round(avg_momentum, 2),
            'signal': signal,
            'score': score,
            'return_1w': round(return_1w, 2),
            'return_1m': round(return_1m, 2),
            'return_3m': round(return_3m, 2),
        }
        self.signals['Momentum'] = result
        return result

    def run_all(self) -> dict:
        """모든 기술적 지표 계산"""
        print("  |- 기술적 심리 지표 계산 중...")
        self.calculate_rsi()
        self.calculate_macd()
        self.calculate_bollinger()
        self.calculate_volume_trend()
        self.calculate_ma_alignment()
        self.calculate_momentum()

        # 종합 기술적 심리 점수 (-100 ~ +100)
        scores = [s['score'] for s in self.signals.values()]
        if scores:
            # 각 지표의 점수 범위: -2 ~ +2 -> 합산 후 정규화
            max_possible = 2 * len(scores)
            raw_sum = sum(scores)
            normalized = (raw_sum / max_possible) * 100 if max_possible > 0 else 0
        else:
            normalized = 0

        print(f"  |  기술적 지표 {len(self.signals)}개 계산 완료")
        return {
            'composite_score': round(normalized, 1),
            'signals': self.signals,
        }


# ====================================================================
# 3. 시장 심리 분석기 (VIX + 상대강도 + 종합)
# ====================================================================
class MarketSentimentAnalyzer:

    # 종합 심리 라벨 매핑
    SENTIMENT_LABELS = [
        (-100, -60, '극도공포'),
        (-60, -30, '공포'),
        (-30, -10, '약세'),
        (-10, 10, '중립'),
        (10, 30, '약간긍정'),
        (30, 60, '긍정'),
        (60, 101, '극도낙관'),
    ]

    def __init__(self, ticker: str, is_korean: bool = False, is_crypto: bool = False):
        self.ticker = ticker
        self.is_korean = is_korean
        self.is_crypto = is_crypto
        self.vix_data = {}
        self.relative_strength = {}

    def fetch_vix(self) -> dict:
        """VIX (공포지수) 수집"""
        print("  |- VIX (공포지수) 수집 중...")
        try:
            vix = yf.Ticker('^VIX')
            vix_hist = vix.history(period='1mo')
            if vix_hist.empty:
                print("  |  [경고] VIX 데이터 수집 실패")
                return {'value': None, 'signal': '데이터 없음', 'score': 0}

            current_vix = vix_hist['Close'].iloc[-1]

            if current_vix < 15:
                score, signal = 2, '낮은 공포 (안정)'
            elif current_vix < 20:
                score, signal = 1, '보통 수준'
            elif current_vix < 25:
                score, signal = 0, '경계 수준'
            elif current_vix < 30:
                score, signal = -1, '높은 공포'
            elif current_vix < 35:
                score, signal = -2, '매우 높은 공포'
            else:
                score, signal = -2, '극도 공포 (패닉)'

            # VIX 점수를 -100~+100으로 변환
            # VIX 15 이하 = +100, VIX 40 이상 = -100
            vix_normalized = np.clip((25 - current_vix) / 15 * 100, -100, 100)

            self.vix_data = {
                'value': round(current_vix, 1),
                'signal': signal,
                'score': round(float(vix_normalized), 1),
                'raw_score': score,
            }
            print(f"  |  VIX: {current_vix:.1f} ({signal})")
            return self.vix_data

        except Exception as e:
            print(f"  |  [경고] VIX 수집 실패: {e}")
            return {'value': None, 'signal': '수집 실패', 'score': 0, 'raw_score': 0}

    def fetch_relative_strength(self) -> dict:
        """시장 대비 상대 강도"""
        print("  |- 상대 강도 계산 중...")
        try:
            # 비교 지수 선택
            if self.is_korean:
                index_ticker = '^KS11'  # KOSPI
                index_name = 'KOSPI'
            else:
                index_ticker = '^GSPC'  # S&P 500
                index_name = 'S&P500'

            stock = yf.Ticker(self.ticker)
            index = yf.Ticker(index_ticker)

            stock_hist = stock.history(period='3mo')
            index_hist = index.history(period='3mo')

            if stock_hist.empty or index_hist.empty:
                print("  |  [경고] 상대 강도 계산 실패")
                return {'score': 0, 'signal': '데이터 부족'}

            if stock_hist.index.tz is not None:
                stock_hist.index = stock_hist.index.tz_localize(None)
            if index_hist.index.tz is not None:
                index_hist.index = index_hist.index.tz_localize(None)

            # 1개월 수익률
            stock_close = stock_hist['Close']
            index_close = index_hist['Close']

            stock_1m = (stock_close.iloc[-1] / stock_close.iloc[-21] - 1) * 100 if len(stock_close) >= 21 else 0
            index_1m = (index_close.iloc[-1] / index_close.iloc[-21] - 1) * 100 if len(index_close) >= 21 else 0

            # 3개월 수익률
            stock_3m = (stock_close.iloc[-1] / stock_close.iloc[0] - 1) * 100
            index_3m = (index_close.iloc[-1] / index_close.iloc[0] - 1) * 100

            # 상대 강도: 종목 수익률 - 지수 수익률
            rs_1m = stock_1m - index_1m
            rs_3m = stock_3m - index_3m
            rs_avg = rs_1m * 0.6 + rs_3m * 0.4

            # -100 ~ +100 정규화
            rs_normalized = np.clip(rs_avg / 10 * 100, -100, 100)

            if rs_avg > 10:
                signal = f'{index_name} 대비 강한 아웃퍼폼'
            elif rs_avg > 3:
                signal = f'{index_name} 대비 아웃퍼폼'
            elif rs_avg > -3:
                signal = f'{index_name} 대비 유사'
            elif rs_avg > -10:
                signal = f'{index_name} 대비 언더퍼폼'
            else:
                signal = f'{index_name} 대비 강한 언더퍼폼'

            self.relative_strength = {
                'score': round(float(rs_normalized), 1),
                'signal': signal,
                'stock_1m': round(stock_1m, 2),
                'stock_3m': round(stock_3m, 2),
                'index_1m': round(index_1m, 2),
                'index_3m': round(index_3m, 2),
                'rs_1m': round(rs_1m, 2),
                'rs_3m': round(rs_3m, 2),
                'index_name': index_name,
            }
            print(f"  |  상대강도: {rs_avg:+.2f}%p ({signal})")
            return self.relative_strength

        except Exception as e:
            print(f"  |  [경고] 상대 강도 계산 실패: {e}")
            return {'score': 0, 'signal': '계산 실패', 'index_name': 'N/A'}

    def compute_composite(self, news_score: float, tech_score: float) -> dict:
        """
        종합 심리 점수 계산 (가중 평균)
        - 뉴스 심리: 20%
        - 기술적 심리: 50%
        - VIX (시장 공포): 15%
        - 상대 강도: 15%
        """
        vix_score = self.vix_data.get('score', 0)
        rs_score = self.relative_strength.get('score', 0)

        composite = (
            news_score * 0.20 +
            tech_score * 0.50 +
            vix_score * 0.15 +
            rs_score * 0.15
        )
        composite = np.clip(composite, -100, 100)

        # 라벨 매핑
        label = '중립'
        for low, high, lbl in self.SENTIMENT_LABELS:
            if low <= composite < high:
                label = lbl
                break

        return {
            'composite_score': round(float(composite), 1),
            'label': label,
            'news_score': round(news_score, 1),
            'tech_score': round(tech_score, 1),
            'vix_score': round(vix_score, 1),
            'rs_score': round(rs_score, 1),
        }

    @staticmethod
    def get_sentiment_description(composite: dict) -> str:
        """종합 판단 텍스트 생성"""
        score = composite['composite_score']
        label = composite['label']
        parts = []

        # 종합 판단
        if score > 30:
            parts.append(f"현재 시장 심리는 [{label}] 상태로, 전반적으로 긍정적인 흐름입니다.")
        elif score > 0:
            parts.append(f"현재 시장 심리는 [{label}] 상태로, 소폭 긍정적인 분위기입니다.")
        elif score > -30:
            parts.append(f"현재 시장 심리는 [{label}] 상태로, 다소 부정적인 흐름이 감지됩니다.")
        else:
            parts.append(f"현재 시장 심리는 [{label}] 상태로, 시장 전반에 비관적 분위기가 팽배합니다.")

        # 기술적 심리 코멘트
        tech = composite['tech_score']
        if tech > 20:
            parts.append("기술적 지표들이 매수 시그널을 보이고 있습니다.")
        elif tech < -20:
            parts.append("기술적 지표들이 매도 시그널을 보이고 있습니다.")
        else:
            parts.append("기술적 지표는 뚜렷한 방향성 없이 혼조세입니다.")

        # VIX 코멘트
        vix = composite['vix_score']
        if vix < -30:
            parts.append("VIX가 높아 시장 전반의 불안감이 커진 상황입니다.")
        elif vix > 30:
            parts.append("VIX가 낮아 시장의 안정성이 유지되고 있습니다.")

        return ' '.join(parts)

    @staticmethod
    def get_contrarian_signal(composite: dict) -> str:
        """역발상 시그널"""
        score = composite['composite_score']
        if score <= -60:
            return "극도 공포 구간: 역발상 매수 관점에서 강한 관심 구간. 역사적으로 극도 공포 시기가 중장기 저점 형성 구간인 경우가 많음."
        elif score <= -30:
            return "공포 구간: 추가 하락 가능성 있으나, 점진적 분할 매수 고려 구간."
        elif score >= 60:
            return "극도 낙관 구간: 역발상 매도 관점에서 주의 필요. 과열 시그널로 차익 실현 검토."
        elif score >= 30:
            return "낙관 구간: 추세 추종은 유효하나, 포지션 과다 확대 주의."
        else:
            return "중립 구간: 뚜렷한 역발상 시그널 없음. 기본 전략 유지."

    @staticmethod
    def get_action_recommendation(composite: dict) -> str:
        """투자 행동 권고"""
        score = composite['composite_score']
        if score <= -60:
            return "강력 관심 구간 (역발상 매수). 분할 매수 시작 검토. 단, 추세 반전 확인 필요."
        elif score <= -30:
            return "관심 구간. 점진적 매수 검토. 현금 비중 유지하며 분할 접근."
        elif score <= -10:
            return "방어적 운용. 포지션 축소 검토. 손절 라인 점검."
        elif score <= 10:
            return "관망. 뚜렷한 방향성 부재. 기존 포지션 유지, 신규 진입은 보류."
        elif score <= 30:
            return "소폭 긍정. 기존 포지션 유지. 눌림목 매수 기회 모색."
        elif score <= 60:
            return "긍정적. 추세 추종 매매. 단, 과열 경계 필요."
        else:
            return "극도 낙관. 차익 실현 검토. 신규 매수는 자제. 리스크 관리 강화."


# ====================================================================
# 4. Excel 보고서 생성기
# ====================================================================
class SentimentExcelBuilder:

    # 색상 팔레트 (financial_analyzer.py 동일)
    COLORS = {
        'header_dark': '0D1B2A',
        'header_blue': '1B4F72',
        'accent': '2E86C1',
        'light_blue': 'D6EAF8',
        'white': 'FFFFFF',
        'light_gray': 'F2F3F4',
        'positive': 'EAFAF1',
        'negative': 'FDEDEC',
        'gold': 'F39C12',
    }

    # 심리 라벨별 색상
    SENTIMENT_COLORS = {
        '극도공포': '922B21',
        '공포': 'E74C3C',
        '약세': 'F0B27A',
        '중립': 'F7DC6F',
        '약간긍정': '82E0AA',
        '긍정': '2ECC71',
        '극도낙관': '1A5276',
    }

    SENTIMENT_FONT_COLORS = {
        '극도공포': 'FFFFFF',
        '공포': 'FFFFFF',
        '약세': '0D1B2A',
        '중립': '0D1B2A',
        '약간긍정': '0D1B2A',
        '긍정': 'FFFFFF',
        '극도낙관': 'FFFFFF',
    }

    def __init__(self, ticker: str, company_name: str):
        self.wb = Workbook()
        self.ticker = ticker
        self.company_name = company_name
        self.wb.remove(self.wb.active)

    def _style_header(self, ws, row, col, value, color='header_dark', font_color='FFFFFF', bold=True, size=11):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = PatternFill("solid", fgColor=self.COLORS.get(color, color))
        cell.font = Font(bold=bold, color=font_color, size=size, name='맑은 고딕')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return cell

    def _style_data(self, ws, row, col, value, bg_color=None, number_format=None, bold=False):
        cell = ws.cell(row=row, column=col, value=value)
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=self.COLORS.get(bg_color, bg_color))
        cell.font = Font(size=10, name='맑은 고딕', bold=bold)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if number_format:
            cell.number_format = number_format
        return cell

    def _score_color(self, score):
        """점수에 따른 배경색"""
        if score > 30:
            return 'positive'
        elif score > 0:
            return 'EAFAF1'
        elif score > -30:
            return 'light_gray'
        else:
            return 'negative'

    def add_dashboard_sheet(self, composite: dict, description: str,
                            contrarian: str, recommendation: str):
        """시트 1: 종합 심리 대시보드"""
        ws = self.wb.create_sheet("종합 심리 대시보드")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = f"{self.company_name} ({self.ticker}) - 심리 분석 대시보드"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 날짜
        ws['B3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B3'].font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 종합 심리 점수 (대형 표시)
        ws.merge_cells('B5:H5')
        self._style_header(ws, 5, 2, "종합 심리 점수", 'header_dark', size=12)
        ws.row_dimensions[5].height = 30

        label = composite['label']
        score = composite['composite_score']
        sc = self.SENTIMENT_COLORS.get(label, 'F7DC6F')
        sf = self.SENTIMENT_FONT_COLORS.get(label, '0D1B2A')

        ws.merge_cells('B6:H6')
        score_cell = ws.cell(row=6, column=2, value=f"[{label}]  {score:+.1f}점")
        score_cell.fill = PatternFill("solid", fgColor=sc)
        score_cell.font = Font(bold=True, size=18, color=sf, name='맑은 고딕')
        score_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[6].height = 50

        # 게이지 표시 (텍스트 기반)
        gauge = self._build_gauge(score)
        ws.merge_cells('B7:H7')
        gauge_cell = ws.cell(row=7, column=2, value=gauge)
        gauge_cell.font = Font(size=10, name='Consolas', color='0D1B2A')
        gauge_cell.alignment = Alignment(horizontal='center', vertical='center')
        gauge_cell.fill = PatternFill("solid", fgColor='F2F3F4')
        ws.row_dimensions[7].height = 25

        # 서브 점수 테이블
        row = 9
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "세부 심리 점수", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        sub_headers = ['지표', '점수', '비중', '가중 기여']
        for col, h in enumerate(sub_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        sub_items = [
            ('뉴스 심리', composite['news_score'], '20%', composite['news_score'] * 0.2),
            ('기술적 심리', composite['tech_score'], '50%', composite['tech_score'] * 0.5),
            ('시장 공포 (VIX)', composite['vix_score'], '15%', composite['vix_score'] * 0.15),
            ('상대 강도', composite['rs_score'], '15%', composite['rs_score'] * 0.15),
        ]

        for name, sc_val, weight, contrib in sub_items:
            bg = self._score_color(sc_val)
            ws.cell(row=row, column=2, value=name).font = Font(bold=True, size=10, name='맑은 고딕')
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            self._style_data(ws, row, 3, f"{sc_val:+.1f}", bg_color=bg)
            self._style_data(ws, row, 4, weight)
            self._style_data(ws, row, 5, f"{contrib:+.1f}", bg_color=bg)
            ws.row_dimensions[row].height = 22
            row += 1

        # 종합 판단
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "종합 판단", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f'B{row}:H{row+1}')
        desc_cell = ws.cell(row=row, column=2, value=description)
        desc_cell.font = Font(size=10, name='맑은 고딕')
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
        ws.row_dimensions[row].height = 45

        # 역발상 시그널
        row += 3
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "역발상 시그널", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f'B{row}:H{row+1}')
        cont_cell = ws.cell(row=row, column=2, value=contrarian)
        cont_cell.font = Font(size=10, name='맑은 고딕', color='1B4F72')
        cont_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
        cont_cell.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
        ws.row_dimensions[row].height = 40

        # 투자 행동 권고
        row += 3
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "투자 행동 권고", 'accent', size=11)
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f'B{row}:H{row+1}')
        rec_cell = ws.cell(row=row, column=2, value=recommendation)
        rec_cell.font = Font(size=10, name='맑은 고딕', bold=True)
        rec_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
        rec_cell.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
        ws.row_dimensions[row].height = 40

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18
        for c in 'CDEFGH':
            ws.column_dimensions[c].width = 14

    def _build_gauge(self, score: float) -> str:
        """텍스트 기반 게이지 생성"""
        # -100 ~ +100 을 0 ~ 40 위치로 매핑
        pos = int((score + 100) / 200 * 40)
        pos = max(0, min(40, pos))
        bar = ['-'] * 41
        bar[pos] = '*'
        return '[극도공포 ' + ''.join(bar) + ' 극도낙관]'

    def add_news_sheet(self, news_data: list, aggregate: dict):
        """시트 2: 뉴스 분석"""
        ws = self.wb.create_sheet("뉴스 분석")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "뉴스 감성 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 뉴스 요약
        ws.merge_cells('B4:G4')
        self._style_header(ws, 4, 2, "뉴스 심리 요약", 'header_dark', size=11)
        ws.row_dimensions[4].height = 28

        summary_items = [
            ('종합 뉴스 심리', f"{aggregate['score']:+.1f}점 ({aggregate['label']})"),
            ('분석 뉴스 수', f"{aggregate['total_news']}건"),
            ('긍정 뉴스', f"{aggregate['positive_count']}건"),
            ('부정 뉴스', f"{aggregate['negative_count']}건"),
            ('중립 뉴스', f"{aggregate['neutral_count']}건"),
        ]

        row = 5
        for key, val in summary_items:
            ws.cell(row=row, column=2, value=key).font = Font(bold=True, size=10, name='맑은 고딕')
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(f'C{row}:G{row}')
            self._style_data(ws, row, 3, val)
            ws.row_dimensions[row].height = 22
            row += 1

        # 뉴스 상세 테이블
        row += 1
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "뉴스 상세 목록", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        detail_headers = ['날짜', '제목', '감성점수', '긍정 키워드', '부정 키워드', '출처']
        for col, h in enumerate(detail_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        if not news_data:
            ws.merge_cells(f'B{row}:G{row}')
            ws.cell(row=row, column=2, value="뉴스 데이터 없음").font = Font(
                size=10, name='맑은 고딕', color='7F8C8D')
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        else:
            for item in news_data:
                score_val = item['score']
                bg = 'positive' if score_val > 0 else ('negative' if score_val < 0 else 'light_gray')

                ws.cell(row=row, column=2, value=item['date']).font = Font(size=10, name='맑은 고딕')
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')

                # 제목 (cp949 안전하게 처리)
                title_safe = item['title'].encode('cp949', errors='replace').decode('cp949')
                title_cell = ws.cell(row=row, column=3, value=title_safe)
                title_cell.font = Font(size=9, name='맑은 고딕')
                title_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

                self._style_data(ws, row, 4, f"{score_val:+d}", bg_color=bg)

                pos_kw = ', '.join(item['positive_keywords'][:3]) if item['positive_keywords'] else '-'
                neg_kw = ', '.join(item['negative_keywords'][:3]) if item['negative_keywords'] else '-'
                self._style_data(ws, row, 5, pos_kw, bg_color='positive' if pos_kw != '-' else None)
                self._style_data(ws, row, 6, neg_kw, bg_color='negative' if neg_kw != '-' else None)
                self._style_data(ws, row, 7, item.get('publisher', 'N/A'))

                ws.row_dimensions[row].height = 28
                row += 1

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 16

    def add_technical_sheet(self, signals: dict):
        """시트 3: 기술적 심리"""
        ws = self.wb.create_sheet("기술적 심리")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:G2')
        title = ws['B2']
        title.value = "기술적 심리 지표 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 헤더
        row = 4
        ws.merge_cells(f'B{row}:G{row}')
        self._style_header(ws, row, 2, "기술적 지표 상세", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        detail_headers = ['지표명', '현재값', '시그널', '심리점수', '해석', '수치 상세']
        for col, h in enumerate(detail_headers, 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        # 지표 라벨 매핑
        indicator_names = {
            'RSI': 'RSI(14)',
            'MACD': 'MACD(12,26,9)',
            'Bollinger': '볼린저 밴드(20,2)',
            'Volume': '거래량 추세',
            'MA_Alignment': '이동평균 배열',
            'Momentum': '가격 모멘텀',
        }

        for key, name in indicator_names.items():
            sig = signals.get(key, {})
            score = sig.get('score', 0)
            value = sig.get('value', 'N/A')
            signal_text = sig.get('signal', 'N/A')

            # 심리 점수에 따른 색상
            if score >= 2:
                bg = '1A5276'
                fc = 'FFFFFF'
                score_label = f"+{score} (매우 긍정)"
            elif score == 1:
                bg = '2ECC71'
                fc = 'FFFFFF'
                score_label = f"+{score} (긍정)"
            elif score == 0:
                bg = 'F7DC6F'
                fc = '0D1B2A'
                score_label = f"{score} (중립)"
            elif score == -1:
                bg = 'F0B27A'
                fc = '0D1B2A'
                score_label = f"{score} (부정)"
            else:
                bg = 'E74C3C'
                fc = 'FFFFFF'
                score_label = f"{score} (매우 부정)"

            # 지표명
            ws.cell(row=row, column=2, value=name).font = Font(bold=True, size=10, name='맑은 고딕')
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)

            # 현재값
            if isinstance(value, float):
                self._style_data(ws, row, 3, f"{value:.2f}")
            else:
                self._style_data(ws, row, 3, str(value) if value is not None else 'N/A')

            # 시그널
            sig_cell = ws.cell(row=row, column=4, value=signal_text)
            sig_cell.font = Font(size=10, name='맑은 고딕')
            sig_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 심리점수 (색상)
            score_cell = ws.cell(row=row, column=5, value=score_label)
            score_cell.fill = PatternFill("solid", fgColor=bg)
            score_cell.font = Font(size=10, name='맑은 고딕', bold=True, color=fc)
            score_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 해석
            interp = self._get_interpretation(key, sig)
            interp_cell = ws.cell(row=row, column=6, value=interp)
            interp_cell.font = Font(size=9, name='맑은 고딕')
            interp_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

            # 수치 상세
            detail = self._get_detail_string(key, sig)
            detail_cell = ws.cell(row=row, column=7, value=detail)
            detail_cell.font = Font(size=9, name='맑은 고딕', color='7F8C8D')
            detail_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

            ws.row_dimensions[row].height = 35
            row += 1

        # 범례
        row += 2
        ws.merge_cells(f'B{row}:G{row}')
        ws.cell(row=row, column=2,
                value="* 심리점수: -2(매우부정) ~ +2(매우긍정). 색상: 파랑=긍정, 초록=약긍정, 노랑=중립, 주황=약부정, 빨강=부정"
                ).font = Font(size=9, color='7F8C8D', name='맑은 고딕')

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 28

    def _get_interpretation(self, key: str, sig: dict) -> str:
        """지표별 해석"""
        score = sig.get('score', 0)
        if key == 'RSI':
            v = sig.get('value')
            if v is not None:
                if v >= 70:
                    return f"RSI {v}로 과매수 영역. 단기 조정 가능성."
                elif v <= 30:
                    return f"RSI {v}로 과매도 영역. 반등 가능성."
                else:
                    return f"RSI {v}로 중립 영역."
        elif key == 'MACD':
            return sig.get('signal', '')
        elif key == 'Bollinger':
            v = sig.get('value')
            if v is not None:
                return f"밴드 내 위치 {v}%. 50% 기준 상위/하위 판단."
        elif key == 'Volume':
            ratio = sig.get('value')
            if ratio is not None:
                return f"최근5일/20일 거래량비 {ratio}x"
        elif key == 'MA_Alignment':
            return sig.get('signal', '')
        elif key == 'Momentum':
            return sig.get('signal', '')
        return ''

    def _get_detail_string(self, key: str, sig: dict) -> str:
        """지표별 수치 상세"""
        if key == 'RSI':
            return f"RSI: {sig.get('value', 'N/A')}"
        elif key == 'MACD':
            return f"MACD: {sig.get('value', 'N/A')}, Signal: {sig.get('signal_line', 'N/A')}, Hist: {sig.get('histogram', 'N/A')}"
        elif key == 'Bollinger':
            return f"Upper: {sig.get('upper', 'N/A')}, Mid: {sig.get('middle', 'N/A')}, Lower: {sig.get('lower', 'N/A')}"
        elif key == 'Volume':
            return f"5일평균: {sig.get('avg_5d', 'N/A'):,}, 20일평균: {sig.get('avg_20d', 'N/A'):,}" if isinstance(sig.get('avg_5d'), int) else 'N/A'
        elif key == 'MA_Alignment':
            parts = []
            for ma in ['ma5', 'ma20', 'ma60', 'ma120']:
                v = sig.get(ma)
                if v is not None:
                    parts.append(f"{ma.upper()}: {v:,.2f}")
            return ', '.join(parts) if parts else 'N/A'
        elif key == 'Momentum':
            return f"1W: {sig.get('return_1w', 'N/A')}%, 1M: {sig.get('return_1m', 'N/A')}%, 3M: {sig.get('return_3m', 'N/A')}%"
        return ''

    def add_investment_sentiment_sheet(self, composite: dict, vix_data: dict,
                                        relative_strength: dict, tech_signals: dict,
                                        contrarian: str, recommendation: str):
        """시트 4: 투자 심리 종합"""
        ws = self.wb.create_sheet("투자 심리 종합")
        ws.sheet_view.showGridLines = False

        # 제목
        ws.merge_cells('B2:H2')
        title = ws['B2']
        title.value = "투자 심리 종합 분석"
        title.font = Font(bold=True, size=14, color='1B4F72', name='맑은 고딕')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 35

        # 현재 심리 레짐
        row = 4
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "현재 심리 레짐", 'header_dark', size=12)
        ws.row_dimensions[row].height = 30
        row += 1

        label = composite['label']
        score = composite['composite_score']
        sc = self.SENTIMENT_COLORS.get(label, 'F7DC6F')
        sf = self.SENTIMENT_FONT_COLORS.get(label, '0D1B2A')

        regime_items = [
            ('종합 심리', f"[{label}] {score:+.1f}점"),
            ('뉴스 심리', f"{composite['news_score']:+.1f}점"),
            ('기술적 심리', f"{composite['tech_score']:+.1f}점"),
            ('VIX (공포지수)', f"{vix_data.get('value', 'N/A')} ({vix_data.get('signal', 'N/A')})"),
            ('상대 강도', f"{relative_strength.get('signal', 'N/A')}"),
        ]

        for key, val in regime_items:
            ws.cell(row=row, column=2, value=key).font = Font(bold=True, size=10, name='맑은 고딕')
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.merge_cells(f'C{row}:H{row}')
            data_cell = ws.cell(row=row, column=3, value=val)
            data_cell.font = Font(size=10, name='맑은 고딕')
            data_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            if key == '종합 심리':
                data_cell.fill = PatternFill("solid", fgColor=sc)
                data_cell.font = Font(size=12, name='맑은 고딕', bold=True, color=sf)
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                bg = 'light_gray' if row % 2 == 0 else 'white'
                data_cell.fill = PatternFill("solid", fgColor=self.COLORS[bg])

            ws.row_dimensions[row].height = 24
            row += 1

        # 상대 강도 상세
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "시장 대비 상대 강도 상세", 'header_blue', size=11)
        ws.row_dimensions[row].height = 28
        row += 1

        index_name = relative_strength.get('index_name', 'Index')
        rs_items = [
            ('기간', '종목 수익률', f'{index_name} 수익률', '초과 수익률'),
        ]
        for col, h in enumerate(rs_items[0], 2):
            self._style_header(ws, row, col, h, 'header_blue', size=10)
        ws.row_dimensions[row].height = 24
        row += 1

        periods = [
            ('1개월', relative_strength.get('stock_1m', 'N/A'),
             relative_strength.get('index_1m', 'N/A'), relative_strength.get('rs_1m', 'N/A')),
            ('3개월', relative_strength.get('stock_3m', 'N/A'),
             relative_strength.get('index_3m', 'N/A'), relative_strength.get('rs_3m', 'N/A')),
        ]

        for period_name, stock_ret, idx_ret, excess in periods:
            self._style_data(ws, row, 2, period_name, bold=True)
            if isinstance(stock_ret, (int, float)):
                bg = 'positive' if stock_ret > 0 else 'negative'
                self._style_data(ws, row, 3, f"{stock_ret:+.2f}%", bg_color=bg)
            else:
                self._style_data(ws, row, 3, str(stock_ret))
            if isinstance(idx_ret, (int, float)):
                self._style_data(ws, row, 4, f"{idx_ret:+.2f}%")
            else:
                self._style_data(ws, row, 4, str(idx_ret))
            if isinstance(excess, (int, float)):
                bg = 'positive' if excess > 0 else 'negative'
                self._style_data(ws, row, 5, f"{excess:+.2f}%p", bg_color=bg)
            else:
                self._style_data(ws, row, 5, str(excess))
            ws.row_dimensions[row].height = 22
            row += 1

        # 역발상 시그널
        row += 1
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "역발상 시그널", 'header_dark', size=11)
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f'B{row}:H{row+1}')
        cont_cell = ws.cell(row=row, column=2, value=contrarian)
        cont_cell.font = Font(size=10, name='맑은 고딕', color='1B4F72')
        cont_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
        cont_cell.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
        ws.row_dimensions[row].height = 40

        # 행동 권고
        row += 3
        ws.merge_cells(f'B{row}:H{row}')
        self._style_header(ws, row, 2, "투자 행동 권고", 'accent', size=11)
        ws.row_dimensions[row].height = 28
        row += 1
        ws.merge_cells(f'B{row}:H{row+1}')
        rec_cell = ws.cell(row=row, column=2, value=recommendation)
        rec_cell.font = Font(size=10, name='맑은 고딕', bold=True)
        rec_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', indent=1)
        rec_cell.fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])
        ws.row_dimensions[row].height = 40

        # 면책조항
        row += 4
        ws.merge_cells(f'B{row}:H{row}')
        disclaimer = ws.cell(row=row, column=2,
            value="* 본 분석은 참고용이며 투자 판단의 최종 책임은 투자자 본인에게 있습니다. 키워드 기반 감성 분석은 한계가 있으며 실제 시장 상황과 다를 수 있습니다.")
        disclaimer.font = Font(size=9, color='7F8C8D', name='맑은 고딕')
        disclaimer.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left', indent=1)

        # 열 너비
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18
        for c in 'CDEFGH':
            ws.column_dimensions[c].width = 16

    def save(self, filepath: str):
        self.wb.save(filepath)
        print(f"\n  [저장 완료] {filepath}")


# ====================================================================
# 5. 콘솔 출력
# ====================================================================
def print_console_report(ticker_raw: str, company_name: str,
                         composite: dict, tech_signals: dict,
                         vix_data: dict, relative_strength: dict,
                         description: str, recommendation: str):
    """콘솔에 심리 분석 결과 출력"""
    print("\n" + "=" * 60)
    print(f"  [{ticker_raw}] 심리 분석 결과 - {company_name}")
    print("=" * 60)

    label = composite['label']
    score = composite['composite_score']
    print(f"\n  종합 심리: [{label}] ({score:+.1f}점)")

    print(f"\n  뉴스 심리:     {composite['news_score']:+6.1f}")
    print(f"  기술적 심리:   {composite['tech_score']:+6.1f}", end='')

    # 핵심 기술적 시그널 요약
    tech_details = []
    rsi = tech_signals.get('RSI', {})
    if rsi.get('value') is not None:
        tech_details.append(f"RSI {rsi['value']}")
    macd = tech_signals.get('MACD', {})
    if macd.get('signal'):
        tech_details.append(macd['signal'])
    if tech_details:
        print(f" ({', '.join(tech_details)})")
    else:
        print()

    vix_val = vix_data.get('value', 'N/A')
    vix_sig = vix_data.get('signal', '')
    print(f"  시장 공포:     {composite['vix_score']:+6.1f} (VIX {vix_val}, {vix_sig})")
    print(f"  상대 강도:     {composite['rs_score']:+6.1f} ({relative_strength.get('signal', 'N/A')})")

    # 핵심 시그널 라인
    key_signals = []
    if rsi.get('value') is not None:
        rsi_v = rsi['value']
        if rsi_v >= 70:
            key_signals.append(f"RSI {rsi_v} (과매수)")
        elif rsi_v <= 30:
            key_signals.append(f"RSI {rsi_v} (과매도)")
        else:
            key_signals.append(f"RSI {rsi_v}")
    ma = tech_signals.get('MA_Alignment', {})
    if ma.get('value') and ma['value'] != 'N/A':
        key_signals.append(f"{ma['value']} 진행중")

    if key_signals:
        print(f"\n  핵심 시그널: {', '.join(key_signals)}")

    print(f"  권고: {recommendation}")
    print("=" * 60)


# ====================================================================
# 6. 메인 실행
# ====================================================================
def main():
    parser = argparse.ArgumentParser(description="뉴스 및 시장 심리 분석 시스템")
    parser.add_argument('--ticker', required=True, help="종목 코드 (예: 005930, AAPL, BTC-USD)")
    args = parser.parse_args()

    ticker_input = args.ticker.strip()

    print("\n" + "=" * 60)
    print("  뉴스 및 시장 심리 분석 시스템")
    print("=" * 60)

    # -- 티커 감지 및 시장 분류 --
    is_korean = False
    is_crypto = False
    ticker_raw = ticker_input

    if ticker_input.isdigit():
        is_korean = True
        ticker = _detect_market(ticker_input)
    elif '-' in ticker_input and ticker_input.endswith('USD'):
        is_crypto = True
        ticker = ticker_input.upper()
    else:
        ticker = ticker_input.upper()

    print(f"\n  분석 대상: {ticker}")
    if is_korean:
        print("  시장: 한국 (KOSPI/KOSDAQ)")
    elif is_crypto:
        print("  시장: 암호화폐")
    else:
        print("  시장: 해외 주식")

    # -- 기업/종목 정보 수집 --
    print(f"\n  종목 정보 수집 중...")
    stock = yf.Ticker(ticker)
    info = stock.info
    company_name = info.get('longName') or info.get('shortName', ticker_raw)
    print(f"  종목명: {company_name}")

    # -- 1. 뉴스 감성 분석 --
    print(f"\n[1/4] 뉴스 감성 분석")
    news_collector = NewsSentimentCollector(ticker, is_crypto=is_crypto)
    news_data = news_collector.fetch_news()
    news_aggregate = news_collector.get_aggregate_score()
    news_score = news_aggregate['score']
    print(f"  |  뉴스 심리: {news_score:+.1f}점 ({news_aggregate['label']})")

    # -- 2. 기술적 심리 분석 --
    print(f"\n[2/4] 기술적 심리 분석")
    tech_analyzer = TechnicalSentimentAnalyzer(ticker)
    price_df = tech_analyzer.fetch_price_data()
    if price_df.empty:
        print("  [오류] 주가 데이터 부족으로 분석을 중단합니다.")
        sys.exit(1)
    tech_result = tech_analyzer.run_all()
    tech_score = tech_result['composite_score']
    print(f"  |  기술적 심리: {tech_score:+.1f}점")

    # -- 3. 시장 심리 (VIX + 상대강도) --
    print(f"\n[3/4] 시장 심리 분석")
    market_analyzer = MarketSentimentAnalyzer(ticker, is_korean=is_korean, is_crypto=is_crypto)
    vix_data = market_analyzer.fetch_vix()
    relative_strength = market_analyzer.fetch_relative_strength()

    # -- 4. 종합 점수 계산 --
    print(f"\n[4/4] 종합 심리 점수 계산")
    composite = market_analyzer.compute_composite(news_score, tech_score)
    description = MarketSentimentAnalyzer.get_sentiment_description(composite)
    contrarian = MarketSentimentAnalyzer.get_contrarian_signal(composite)
    recommendation = MarketSentimentAnalyzer.get_action_recommendation(composite)
    print(f"  종합: [{composite['label']}] {composite['composite_score']:+.1f}점")

    # -- 콘솔 리포트 --
    print_console_report(
        ticker_raw, company_name, composite,
        tech_result['signals'], vix_data, relative_strength,
        description, recommendation,
    )

    # -- Excel 저장 --
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"output/{ticker_raw}_심리분석_{date_str}.xlsx"

    print(f"\n  Excel 보고서 생성 중...")
    builder = SentimentExcelBuilder(ticker_raw, company_name)
    builder.add_dashboard_sheet(composite, description, contrarian, recommendation)
    builder.add_news_sheet(news_data, news_aggregate)
    builder.add_technical_sheet(tech_result['signals'])
    builder.add_investment_sentiment_sheet(
        composite, vix_data, relative_strength,
        tech_result['signals'], contrarian, recommendation,
    )
    builder.save(filename)

    print(f"\n  다음 단계: output 폴더의 Excel 파일을 확인하세요.")
    print(f"  파일 경로: {filename}\n")


def _detect_market(ticker_code):
    """코스피(.KS) / 코스닥(.KQ) 자동 감지"""
    for suffix in [".KS", ".KQ"]:
        try:
            t = yf.Ticker(ticker_code + suffix)
            hist = t.history(period="5d")
            if not hist.empty and len(hist) > 0:
                info = t.info or {}
                name = info.get('longName') or info.get('shortName', '')
                if name and not name.startswith(ticker_code):
                    return ticker_code + suffix
        except Exception:
            pass
    print(f"  [!] 시장 자동 감지 실패, 코스피(.KS)로 시도합니다.")
    return ticker_code + ".KS"


if __name__ == "__main__":
    main()
